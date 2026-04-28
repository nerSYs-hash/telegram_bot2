from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Header, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse, FileResponse
import io
import re
import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn
import os
import sys
import json
import hmac
import hashlib
import time
import jwt
from datetime import datetime, timedelta
import logging
from decimal import Decimal

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("PulseApi")

# Определение путей (api.py в корне бота)
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

app = FastAPI(
    title="Pulse Pro API",
    description="Backend API для панели управления (из корня бота)",
    version="1.6.0"
)

# CORS настройки: критически важны для связи с React
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://puls-chat.ru", "http://puls-chat.ru", "https://www.puls-chat.ru", "http://www.puls-chat.ru"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- БЕЗОПАСНАЯ ИНИЦИАЛИЗАЦИЯ МОДУЛЕЙ БОТА ---
db = None
calculate_health = None

try:
    from database.db_manager import Database
    db = Database(db_path='database/bot_database.db')
    logger.info("✅ База данных подключена: database/bot_database.db")
except Exception as e:
    logger.warning(f"⚠️ Ошибка подключения БД: {e}")

# ── Инициализация permissions (создаёт таблицу role_permissions, сидит дефолты) ──
try:
    from permissions import init_permissions_db
    init_permissions_db()
    logger.info("✅ permissions: таблица role_permissions готова")
except Exception as e:
    logger.warning(f"⚠️ Ошибка инициализации permissions: {e}")

# ── Economy: WebSocket + роутер ──
try:
    from api.economy_ws import manager as _economy_ws_manager
    from api.economy_routes import router as economy_router, _setup as _economy_setup
    _economy_setup(db, lambda auth: _require_auth(auth), lambda uid: _resolve_user_role(uid), _economy_ws_manager)
    app.include_router(economy_router)
    logger.info("✅ economy: роутер подключён")
except Exception as e:
    logger.warning(f"⚠️ Ошибка подключения economy router: {e}")
    _economy_ws_manager = None

try:
    # Явно указываем поиск в текущей папке для stats_calculators
    if os.path.exists(os.path.join(current_dir, "stats_calculators.py")):
        import stats_calculators
        calculate_health = stats_calculators.calculate_chat_health_indices
        logger.info("✅ Модуль калькулятора здоровья подключен")
    else:
        logger.warning("⚠️ Файл stats_calculators.py не найден в корне")
except Exception as e:
    logger.warning(f"⚠️ Ошибка импорта stats_calculators: {e}")

# ── Auth config ──
_BOT_TOKEN    = os.getenv("BOT_TOKEN", "")
_BOT_USERNAME = "Pulse_On_bot"
_JWT_SECRET   = os.getenv("JWT_SECRET", "pulse-jwt-secret-2026")
_JWT_DAYS     = 7


def _verify_tg_hash(data: dict) -> bool:
    received = data.get("hash", "")
    check = {k: v for k, v in data.items() if k != "hash"}
    data_str = "\n".join(f"{k}={v}" for k, v in sorted(check.items()))
    secret = hashlib.sha256(_BOT_TOKEN.encode()).digest()
    computed = hmac.new(secret, data_str.encode(), hashlib.sha256).hexdigest()
    return hmac.compare_digest(computed, received)


def _make_jwt(payload: dict) -> str:
    payload["exp"] = int(time.time()) + _JWT_DAYS * 86400
    return jwt.encode(payload, _JWT_SECRET, algorithm="HS256")


def _decode_jwt(token: str) -> dict:
    return jwt.decode(token, _JWT_SECRET, algorithms=["HS256"])


@app.get("/api/auth/config")
async def auth_config():
    return {"bot_username": _BOT_USERNAME}


@app.post("/api/auth/telegram")
async def auth_telegram(request: Request):
    data = await request.json()
    if not _verify_tg_hash(data):
        raise HTTPException(status_code=401, detail="Подпись Telegram не совпадает")
    if time.time() - int(data.get("auth_date", 0)) > 86400:
        raise HTTPException(status_code=401, detail="Данные авторизации устарели")
    user_id = int(data["id"])
    _main_admin_id = int(os.getenv('MAIN_ADMIN_ID', 0))
    _developer_id  = int(os.getenv('DEVELOPER_ID', 0))
    udata = db.get_user(user_id) if db else None
    is_owner    = bool((udata and udata["is_owner"]) or (_main_admin_id and user_id == _main_admin_id))
    is_developer = bool(_developer_id and user_id == _developer_id)
    is_admin    = bool(is_owner or is_developer or (udata and udata["is_admin"]))
    token = _make_jwt({
        "user_id":    user_id,
        "username":   data.get("username", ""),
        "first_name": data.get("first_name", ""),
        "photo_url":  data.get("photo_url", ""),
        "is_admin":   is_admin,
        "is_owner":   is_owner,
    })
    return {"token": token, "is_admin": is_admin, "is_owner": is_owner}


@app.get("/api/auth/me")
async def auth_me(authorization: str = Header(default=None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        return _decode_jwt(authorization[7:])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Токен истёк")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Неверный токен")


def _get_user_role_meta(user_id: int) -> dict:
    """Читает роль + регистрационные мета-данные из pulse_bot.db (источник истины по ролям)."""
    pulse_db_path = os.path.join(current_dir, 'pulse_bot.db')
    if not os.path.exists(pulse_db_path):
        return {}
    try:
        import sqlite3
        conn = sqlite3.connect(pulse_db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            "SELECT role, status, q_name, created_at, last_seen, last_activity "
            "FROM users WHERE tg_id = ?",
            (user_id,)
        )
        row = cur.fetchone()
        conn.close()
        return dict(row) if row else {}
    except Exception as e:
        logger.warning(f"⚠️ Ошибка чтения pulse_bot.db: {e}")
        return {}


@app.get("/api/admin/profile/me")
async def admin_profile_me(authorization: str = Header(default=None)):
    """Полные данные профиля для страницы «Профиль» в админ-панели."""
    payload = _require_auth(authorization)
    user_id = int(payload.get("user_id", 0))
    if not user_id:
        raise HTTPException(status_code=400, detail="ID не определён")

    from permissions import ROLE_LABELS, normalize_role, get_role_accesses
    role_raw = normalize_role(_resolve_user_role(user_id))
    role_label = ROLE_LABELS.get(role_raw, "Без статуса")
    role_meta = _get_user_role_meta(user_id)

    # Сообщения и последняя активность из bot_database.db.user_stats
    total_messages = 0
    last_msg_date  = None
    if db:
        try:
            db.cursor.execute(
                "SELECT COALESCE(SUM(total_messages), 0) AS total, MAX(date) AS last_date "
                "FROM user_stats WHERE user_id = ?",
                (user_id,)
            )
            r = db.cursor.fetchone()
            if r:
                total_messages = int(r['total'])
                last_msg_date  = r['last_date']
        except Exception as e:
            logger.warning(f"⚠️ Ошибка чтения user_stats: {e}")

    # joined_at: bot_database.db.users.joined_at, фолбэк → pulse_bot.db.users.created_at
    joined_at = None
    if db:
        try:
            db.cursor.execute("SELECT joined_at FROM users WHERE user_id = ?", (user_id,))
            r = db.cursor.fetchone()
            if r and r['joined_at']:
                joined_at = r['joined_at']
        except Exception:
            pass
    if not joined_at:
        joined_at = role_meta.get("created_at")

    has_chat = role_raw in ("owner", "deputy", "admin")
    accesses = get_role_accesses(role_raw)
    from permissions import get_role_permissions
    permissions_flat = get_role_permissions(role_raw)

    return {
        "user_id":        user_id,
        "username":       payload.get("username", ""),
        "first_name":     payload.get("first_name", ""),
        "photo_url":      payload.get("photo_url", ""),
        "role":           role_raw,
        "role_raw":       role_raw,
        "role_label":     role_label,
        "joined_at":      joined_at,
        "last_message":   last_msg_date,
        "total_messages": total_messages,
        "has_q_name":     bool(role_meta.get("q_name")),
        "status":         role_meta.get("status") or "unknown",
        "has_chat":       has_chat,
        "bot_username":   _BOT_USERNAME,
        "accesses":       accesses,
        "permissions":    permissions_flat,
    }


def _require_auth(authorization: str) -> dict:
    """Декодирует JWT и возвращает payload. Бросает HTTPException при ошибке."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        return _decode_jwt(authorization[7:])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Токен истёк")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Неверный токен")


def _resolve_user_role(user_id: int) -> str:
    """Возвращает роль пользователя (owner/developer/deputy/admin/user)."""
    main_admin_id = int(os.getenv('MAIN_ADMIN_ID', 0))
    if main_admin_id and user_id == main_admin_id:
        return "owner"
    developer_id = int(os.getenv('DEVELOPER_ID', 0))
    if developer_id and user_id == developer_id:
        return "developer"
    meta = _get_user_role_meta(user_id)
    return (meta.get("role") or "user").lower()


def _require_owner(authorization: str) -> int:
    """Проверяет что пользователь — owner. Возвращает user_id или 403."""
    payload = _require_auth(authorization)
    user_id = int(payload.get("user_id", 0))
    if not user_id:
        raise HTTPException(status_code=400, detail="ID не определён")
    if _resolve_user_role(user_id) != "owner":
        raise HTTPException(status_code=403, detail="Доступно только владельцу")
    return user_id


# ── PERMISSIONS API ──
@app.get("/api/admin/permissions/catalog")
async def permissions_catalog(authorization: str = Header(default=None)):
    """Полный каталог: ресурсы, действия, роли. Доступно любому авторизованному."""
    _require_auth(authorization)
    from permissions import get_catalog
    return get_catalog()


@app.get("/api/admin/permissions/roles")
async def permissions_roles(authorization: str = Header(default=None)):
    """Текущая раскладка прав по всем ролям. Доступно любому авторизованному."""
    _require_auth(authorization)
    from permissions import get_all_role_permissions
    return get_all_role_permissions()


class PermissionsUpdate(BaseModel):
    permissions: list[str]


@app.put("/api/admin/permissions/roles/{role}")
async def permissions_roles_update(
    role: str,
    body: PermissionsUpdate,
    authorization: str = Header(default=None),
):
    """Перезаписать список permissions для роли. Только owner."""
    _require_owner(authorization)
    from permissions import set_role_permissions, EDITABLE_ROLES, get_role_permissions
    if role not in EDITABLE_ROLES:
        raise HTTPException(
            status_code=400,
            detail=f"Роль '{role}' нельзя редактировать (доступны: {list(EDITABLE_ROLES)})",
        )
    try:
        set_role_permissions(role, body.permissions)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"role": role, "permissions": get_role_permissions(role)}


# --- ЭНДПОИНТЫ ДЛЯ AdminDashboard.jsx ---

@app.get("/")
async def root():
    return {
        "status": "online",
        "message": "API Pulse Pro работает!",
        "time": _now_msk().strftime("%H:%M:%S"),
        "db_connected": db is not None,
        "calculators_loaded": calculate_health is not None
    }

RU_MONTHS = {1:'Янв',2:'Фев',3:'Мар',4:'Апр',5:'Май',6:'Июн',
             7:'Июл',8:'Авг',9:'Сен',10:'Окт',11:'Ноя',12:'Дек'}

PERIOD_LABELS = {
    'today':     'Сегодня',
    'yesterday': 'Вчера',
    'week':      'Неделя',
    'month':     'Месяц',
    'year':      'Год',
}

# Московское время (UTC+3) — сервер может работать в UTC
MSK_OFFSET = timedelta(hours=3)

def _now_msk() -> datetime:
    """Текущее время по Москве (UTC+3)."""
    return datetime.utcnow() + MSK_OFFSET

def _today_msk():
    """Сегодняшняя дата по Москве."""
    return _now_msk().date()


def _clean_journal_html(text: str) -> str:
    """
    Подготавливает HTML из text_preview для отображения в браузере:
    - убирает незакрытый тег в конце (обрыв по 200 символов)
    - переносы строк → <br>
    - ссылки открываются в новой вкладке
    """
    if not text:
        return ''
    text = re.sub(r'<[^>]*$', '', text)                                  # незакрытый тег в конце
    text = re.sub(r'<a\s', '<a target="_blank" rel="noopener" ', text)   # ссылки в новой вкладке
    text = text.replace('\n', '<br>')                                     # переносы строк
    return text.strip()


def _build_daily_history(start_date, end_date):
    history = []
    cur = start_date
    while cur <= end_date:
        msgs = 0
        if db:
            db.cursor.execute(
                "SELECT COALESCE(SUM(total_messages),0) as s FROM user_stats WHERE date=?",
                (cur.isoformat(),)
            )
            r = db.cursor.fetchone()
            msgs = int(r['s']) if r else 0
        history.append({"day": cur.strftime("%d.%m"), "val": msgs})
        cur += timedelta(days=1)
    return history


def _build_monthly_history(start_date, end_date):
    if not db:
        return []
    db.cursor.execute('''
        SELECT strftime('%Y-%m', date) as mon, COALESCE(SUM(total_messages),0) as val
        FROM user_stats WHERE date >= ? AND date <= ?
        GROUP BY mon ORDER BY mon
    ''', (start_date.isoformat(), end_date.isoformat()))
    history = []
    for r in db.cursor.fetchall():
        r = dict(r)
        month_num = int(r['mon'].split('-')[1])
        history.append({"day": RU_MONTHS.get(month_num, r['mon']), "val": int(r['val'])})
    return history


def _compute_stats(period: str) -> dict:
    today = _today_msk()

    if period == 'yesterday':
        start_date = end_date = today - timedelta(days=1)
        hist_start = today - timedelta(days=6)
        hist_end   = today
        history    = _build_daily_history(hist_start, hist_end)
    elif period == 'week':
        start_date = today - timedelta(days=6)
        end_date   = today
        history    = _build_daily_history(start_date, end_date)
    elif period == 'month':
        start_date = today - timedelta(days=29)
        end_date   = today
        history    = _build_daily_history(start_date, end_date)
    elif period == 'year':
        start_date = today - timedelta(days=364)
        end_date   = today
        history    = _build_monthly_history(start_date, end_date)
    else:  # today
        start_date = end_date = today
        hist_start = today - timedelta(days=6)
        hist_end   = today
        history    = _build_daily_history(hist_start, hist_end)

    bank         = float(db.get_bank_balance()) if db else 0
    rate         = float(db.get_setting('pulse_rate', '1.42')) if db else 1.42
    difficulty_k = float(db.get_setting('difficulty_k', '5.0')) if db else 5.0

    # Активные юзеры за ВЕСЬ период (DISTINCT user_id, хотя бы 1 сообщение)
    active = 0
    if db:
        db.cursor.execute(
            "SELECT COUNT(DISTINCT user_id) as c FROM user_stats "
            "WHERE date >= ? AND date <= ? AND total_messages > 0",
            (start_date.isoformat(), end_date.isoformat())
        )
        r = db.cursor.fetchone()
        active = int(r['c']) if r else 0

    total_users = messages = pulses = 0
    if db:
        db.cursor.execute(
            "SELECT COUNT(*) as c FROM users WHERE is_left=0 AND is_admin=0 AND is_owner=0"
        )
        r = db.cursor.fetchone(); total_users = r['c'] if r else 0

        db.cursor.execute(
            "SELECT COALESCE(SUM(total_messages),0) as s FROM user_stats WHERE date>=? AND date<=?",
            (start_date.isoformat(), end_date.isoformat())
        )
        r = db.cursor.fetchone(); messages = int(r['s']) if r else 0

        db.cursor.execute(
            "SELECT COALESCE(SUM(pulses_mined),0) as s FROM user_stats WHERE date>=? AND date<=?",
            (start_date.isoformat(), end_date.isoformat())
        )
        r = db.cursor.fetchone(); pulses = float(r['s']) if r else 0.0

    dynamics = db.get_user_dynamics_stats(start_date.isoformat(), end_date.isoformat()) if db else {}

    # ── Субиндексы здоровья чата ──
    indices = {"oksp": 0.0, "sdsp": 0.0, "cho": 0.0, "media": 0.0, "korp": 0.0, "kopyup": 0.0}
    health = 0.0
    if db and messages > 0:
        try:
            db.cursor.execute(
                "SELECT COALESCE(SUM(replies_sent),0) as rpl, "
                "COALESCE(SUM(reactions_given),0) as rea, "
                "COALESCE(SUM(media_sent),0) as med "
                "FROM user_stats WHERE date>=? AND date<=?",
                (start_date.isoformat(), end_date.isoformat())
            )
            sr = db.cursor.fetchone()
            replies   = int(sr['rpl'])  if sr else 0
            reactions = int(sr['rea'])  if sr else 0
            media_cnt = int(sr['med'])  if sr else 0
            joined    = dynamics.get('joined', 0)
            left      = dynamics.get('left',   0)

            oksp       = round(min((messages / max(active, 1)) * 10, 100.0), 1)
            sdsp       = round(replies   / messages * 100, 1)
            cho        = round(reactions / messages * 100, 1)
            media_idx  = round(media_cnt / messages * 100, 1)
            korp       = round((replies + reactions) / messages * 100, 1)
            kopyup     = round((joined - left) / max(total_users, 1) * 100, 1)

            indices = {"oksp": oksp, "sdsp": sdsp, "cho": cho,
                       "media": media_idx, "korp": korp, "kopyup": kopyup}
            health = round(min(
                oksp * 0.25 + sdsp * 0.15 + cho * 0.15 +
                media_idx * 0.10 + korp * 0.20 + max(0, kopyup) * 0.15,
                100.0
            ), 1)
        except Exception as _he:
            logger.warning(f"health indices: {_he}")

    return {
        "period":       period,
        "periodLabel":  PERIOD_LABELS.get(period, period),
        "bankBalance":  bank,
        "pulseRate":    rate,
        "difficultyK":  difficulty_k,
        "activeUsers":  active,
        "totalUsers":   total_users,
        "messages":     messages,
        "pulsesMined":  pulses,
        "joined":       dynamics.get('joined', 0),
        "left":         dynamics.get('left', 0),
        "history":      history,
        "healthIndex":  health,
        "indices":      indices,
    }


@app.get("/api/stats")
async def get_stats(period: str = Query('today')):
    """Статистика за период: today / yesterday / week / month / year"""
    try:
        return _compute_stats(period)
    except Exception as e:
        logger.error(f"Error in /api/stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/stats/export")
async def export_stats(period: str = Query('week')):
    """Экспорт статистики в Excel"""
    try:
        data = _compute_stats(period)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Статистика"
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18

        header_font  = Font(bold=True, size=13, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1E293B")
        bold_font    = Font(bold=True, size=11)
        label_fill   = PatternFill("solid", fgColor="F1F5F9")
        center       = Alignment(horizontal='center')

        # ── Заголовок ──
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Pulse Pro — Статистика ({data['periodLabel']})"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center

        ws.merge_cells('A2:B2')
        ws['A2'] = f"Экспорт: {_now_msk().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = center

        # ── Метрики ──
        ws['A4'] = "Показатель";  ws['A4'].font = bold_font; ws['A4'].fill = label_fill
        ws['B4'] = "Значение";    ws['B4'].font = bold_font; ws['B4'].fill = label_fill

        metrics = [
            ("Сообщений за период", data['messages']),
            ("Активных пользователей", data['activeUsers']),
            ("Всего пользователей", data['totalUsers']),
            ("Вступило за период", data['joined']),
            ("Вышло за период", data['left']),
            ("Баланс банка", data['bankBalance']),
            ("Курс пульса", data['pulseRate']),
            ("Пульсов намайнено", data['pulsesMined']),
        ]
        for i, (label, value) in enumerate(metrics, start=5):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value

        # ── История ──
        row = 5 + len(metrics) + 2
        ws[f'A{row}'] = "История активности"
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = label_fill
        ws[f'B{row}'] = "Сообщений"
        ws[f'B{row}'].font = bold_font
        ws[f'B{row}'].fill = label_fill
        row += 1
        for item in data['history']:
            ws[f'A{row}'] = item['day']
            ws[f'B{row}'] = item['val']
            row += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"pulse_stats_{period}_{_now_msk().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        logger.error(f"Error in /api/stats/export: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/top")
async def get_top():
    """Топы пользователей: баланс, майнеры, активисты"""
    try:
        today = _today_msk().isoformat()

        def fmt(row):
            r = dict(row)
            uname = r.get('username') or r.get('first_name') or str(r.get('user_id', '?'))
            r['display_name'] = f"@{uname}" if r.get('username') else uname
            return r

        top_balance   = [fmt(r) for r in db.get_top_users_by_balance(5)]   if db else []
        top_miners    = [fmt(r) for r in db.get_top_daily_earners(today, 5)] if db else []
        top_activists = [fmt(r) for r in db.get_top_activists(today, 5)]    if db else []

        return {
            "topBalance":   top_balance,
            "topMiners":    top_miners,
            "topActivists": top_activists,
        }
    except Exception as e:
        logger.error(f"Error in /api/top: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/shipper")
async def get_shipper():
    """Данные Шиппера"""
    return {
        "enabled": True,
        "mode": "active_48",
        "categories": [
            {"id": "hot18", "name": "🔥 Горячие / 18+", "count": 42, "active": True},
            {"id": "funny", "name": "😂 Смешные / Подколы", "count": 28, "active": True},
            {"id": "romantic", "name": "💘 Милые / Романтика", "count": 15, "active": False}
        ],
        "minHours": 2,
        "maxHours": 5
    }

@app.get("/api/system")
async def get_system():
    """Системные настройки"""
    try:
        pulse_rate = db.get_setting('pulse_rate', '1.42') if db else "1.42"
        return {
            "pulseRate":   float(pulse_rate),
            "difficultyK": 5.0,
            "admins":      ["@vitya_owner", "@alex_admin"],
            "blacklist":   []
        }
    except Exception as e:
        return {"error": str(e)}

def _resolve_target_user_raw(raw) -> str:
    """Приводит trigger.target_user к числовому user_id.
    Принимает: int, "12345" или "@username". Для @username — ищет в users, возвращает user_id.
    Если не найден — оставляет как есть (бот отработает fallback)."""
    if raw is None:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    if s.lstrip('-').isdigit():
        return s
    if s.startswith('@') and db is not None:
        uname = s[1:].lstrip('@').strip()
        if uname:
            try:
                db.cursor.execute('SELECT user_id FROM users WHERE username = ? LIMIT 1', (uname,))
                row = db.cursor.fetchone()
                if row:
                    uid = row['user_id'] if hasattr(row, 'keys') else row[0]
                    if uid:
                        return str(uid)
            except Exception:
                pass
    return s


class TriggerIn(BaseModel):
    name: str
    condition: str = 'contains'
    keywords: str = ''
    probability: int = 100
    where_fires: str = 'all'
    initiator: str = 'all'
    target: str = 'nobody'
    target_user: str = ''
    actions: list = []                 # ["msg_chat", "pin", "delete"]
    action_configs: dict = {}          # {"msg_chat": {...}, "pin": {...}}
    condition_groups: list = []        # исходные группы условий с сайта (JSON)
    bot_msg_delete: str = 'no'
    bot_msg_delete_after: int = 60
    fire_limit: int = 0
    auto_pin: int = 0
    warn_period: int = 0
    is_enabled: bool = True


# ── Маппинг типов действий: сайт ↔ бот (явный для всех активных типов) ──
_SITE_TO_BOT = {
    'send_text': 'msg_chat',
    'dm':        'msg_dm',
    'pin':       'pin',
    'delete':    'delete',
    'warn':      'warn',
    'mute':      'mute',
    'ban':       'ban',
    'emoji':     'emoji',
}
_BOT_TO_SITE = {v: k for k, v in _SITE_TO_BOT.items()}


def _keyboard_to_buttons(keyboard: list) -> list:
    """Конвертирует кнопки из формата сайта в формат бота (только URL-кнопки)."""
    buttons = []
    for btn in keyboard:
        btn_type = btn.get('type', '')
        text = (btn.get('text') or '').strip()
        if not text:
            continue
        if btn_type == 'link':
            url = (btn.get('url') or '').strip()
            if url:
                buttons.append({'text': text, 'url': url})
        # trigger, share, reaction — пока не поддерживаются ботом, пропускаем
    return buttons


def _buttons_to_keyboard(buttons: list) -> list:
    """Конвертирует кнопки из формата бота обратно в формат сайта."""
    keyboard = []
    for btn in buttons:
        text = btn.get('text', '')
        url = btn.get('url', '')
        if text and url:
            keyboard.append({'id': hash(text + url) & 0xFFFFFF, 'type': 'link', 'text': text, 'url': url})
    return keyboard


_TIME_UNIT_TO_SEC = {
    'seconds': 1, 'minutes': 60, 'hours': 3600,
    'days': 86400, 'weeks': 604800, 'months': 2592000,
    # русские (из send_text settings)
    'секунда': 1, 'минута': 60, 'час': 3600,
    'день': 86400, 'неделя': 604800, 'месяц': 2592000,
}

def _sec_to_val_unit(seconds: int) -> tuple:
    """Обратное: секунды → (value, unit_key)."""
    for unit, sec in [('days', 86400), ('hours', 3600), ('minutes', 60)]:
        if seconds % sec == 0 and seconds >= sec:
            return seconds // sec, unit
    return seconds, 'seconds'


def _site_to_bot(site_actions: list, site_configs: dict) -> tuple:
    """Переводит типы и формат из сайта в формат бота перед записью в БД."""
    bot_actions = []
    bot_configs = {}
    for site_type in site_actions:
        if site_type not in _SITE_TO_BOT:
            logger.warning(f"[TRIGGERS] неизвестный тип действия с сайта: {site_type!r} — пропускаю")
            continue
        bot_type = _SITE_TO_BOT[site_type]
        bot_actions.append(bot_type)
        cfg = site_configs.get(site_type, {})

        if site_type in ('send_text', 'dm'):
            variants = cfg.get('variants') or [{}]
            first = variants[0] if variants else {}
            settings = cfg.get('settings', {})
            def _to_sec(val_key, unit_key):
                try:
                    val = int(settings.get(val_key) or 1)
                    unit = settings.get(unit_key) or 'секунда'
                    return val * _TIME_UNIT_TO_SEC.get(unit, 1)
                except Exception:
                    return None
            bot_configs[bot_type] = {
                'text':                 first.get('text', ''),
                'media_type':           first.get('media_type', 'none'),
                'media_id':             first.get('media_id'),
                'media_server_path':    first.get('media_server_path'),
                'media_pos':            first.get('media_pos', 'above'),
                'reply_target':         cfg.get('reply_target', 'none'),
                'target_thread_id':     cfg.get('target_thread_id'),
                'buttons':              _keyboard_to_buttons(cfg.get('keyboard', [])),
                'link_preview':         not settings.get('disable_preview', False),
                'disable_notification': bool(settings.get('disable_notify', False)),
                'protect_content':      bool(settings.get('content_protection', False)),
                'pin':                  bool(settings.get('pin', False)),
                'delete_previous':      bool(settings.get('delete_previous', False)),
                'delete_after_seconds': _to_sec('delete_after_val','delete_after_unit') if settings.get('delete_after') else None,
                'send_delay_seconds':   _to_sec('send_delayed_val','send_delayed_unit') if settings.get('send_delayed') else None,
                '_variants':    variants,
                '_settings':    settings,
            }
            if len(variants) > 1 and site_type == 'send_text':
                bot_configs['rotation'] = {
                    'items': [
                        {
                            'text':              v.get('text', ''),
                            'media_id':          v.get('media_id'),
                            'media_type':        v.get('media_type', 'none'),
                            'media_server_path': v.get('media_server_path'),
                        }
                        for v in variants
                    ],
                    'next_idx': 0,
                }

        elif site_type == 'mute':
            # Конвертируем value+unit → секунды
            dur_sec = 3600
            if cfg.get('mute_time_enabled'):
                try:
                    val = int(cfg.get('mute_time_value') or 1)
                    unit = cfg.get('mute_time_unit') or 'hours'
                    dur_sec = val * _TIME_UNIT_TO_SEC.get(unit, 3600)
                except Exception:
                    pass
            bot_configs['mute'] = {
                'duration_seconds': dur_sec,
                'mute_target':      cfg.get('mute_target', 'initiator'),
                'mute_type':        cfg.get('mute_type', 'all'),
            }

        elif site_type == 'ban':
            # Конвертируем ban_time_value+unit → duration_seconds (0 = permanent)
            dur_sec = 0
            if cfg.get('ban_time_enabled'):
                try:
                    val = int(cfg.get('ban_time_value') or 1)
                    unit = cfg.get('ban_time_unit') or 'hours'
                    dur_sec = val * _TIME_UNIT_TO_SEC.get(unit, 3600)
                except Exception:
                    pass
            bot_configs['ban'] = {
                'ban_target':       cfg.get('ban_target', 'initiator'),
                'duration_seconds': int(dur_sec),
                'revoke_messages':  bool(cfg.get('ban_revoke_messages', False)),
                # round-trip поля для UI
                'ban_time_enabled': bool(cfg.get('ban_time_enabled', False)),
                'ban_time_value':   int(cfg.get('ban_time_value', 1) or 1),
                'ban_time_unit':    cfg.get('ban_time_unit', 'hours') or 'hours',
            }

        elif site_type == 'emoji':
            # сайт шлёт ключ 'emoji' (action.emoji в UI), а не 'emoji_reaction'
            bot_configs['emoji'] = {
                'emoji':        cfg.get('emoji') or cfg.get('emoji_reaction') or '👍',
                'emoji_target': cfg.get('emoji_target', 'initiator'),
            }

        elif site_type == 'warn':
            bot_configs['warn'] = {
                'warn_target':  cfg.get('warn_target', 'initiator'),
                'warn_notify':  cfg.get('warn_notify', True),
                'warn_count':   int(cfg.get('warn_count') or 1),
                'warn_period':  cfg.get('warn_period', 0),
                'warn_mute_duration_seconds': int(cfg.get('warn_mute_duration_seconds') or 3600),
            }

        elif site_type == 'delete':
            del_target = cfg.get('delete_target', 'initiator')
            if del_target == 'both':
                what = ['trigger_word', 'quoted_message']
            elif del_target == 'replied':
                what = ['quoted_message']
            else:
                what = ['trigger_word']
            bot_configs['delete'] = {
                'what':             what,
                'delete_delay':     cfg.get('delete_delay', 0),
                'delete_delay_unit': cfg.get('delete_delay_unit', 'seconds'),
            }

        elif site_type == 'pin':
            # конвертируем pin_time_value+pin_time_unit в секунды (0 = без автоотпкрепления)
            pin_time_enabled = bool(cfg.get('pin_time_enabled', False))
            pin_time_value   = int(cfg.get('pin_time_value', 0) or 0)
            pin_time_unit    = cfg.get('pin_time_unit', 'seconds') or 'seconds'
            _PIN_UNIT_SEC = {
                'seconds': 1, 'minutes': 60, 'hours': 3600,
                'days': 86400, 'weeks': 604800, 'months': 2592000,
            }
            unpin_after = pin_time_value * _PIN_UNIT_SEC.get(pin_time_unit, 1) if pin_time_enabled else 0
            bot_configs['pin'] = {
                'notify':       bool(cfg.get('pin_notify', False)),
                'pin_target':   cfg.get('pin_target', '') or '',
                'unpin_after':  int(unpin_after),
                # сохраняем исходные поля для round-trip обратно в UI
                'pin_time_enabled': pin_time_enabled,
                'pin_time_value':   pin_time_value,
                'pin_time_unit':    pin_time_unit,
            }

        else:
            bot_configs[bot_type] = cfg

        # action_probability — шестерёнка ⚙️ в UI, применяется ко всем типам действий
        if 'action_probability' in cfg:
            try:
                bot_configs[bot_type]['action_probability'] = int(cfg.get('action_probability') or 100)
            except (TypeError, ValueError):
                pass

    return bot_actions, bot_configs


def _bot_to_site(bot_actions: list, bot_configs: dict) -> tuple:
    """Переводит типы и формат из БД (бот) в формат сайта."""
    site_actions = []
    site_configs = {}
    for bot_type in bot_actions:
        site_type = _BOT_TO_SITE.get(bot_type, bot_type)
        site_actions.append(site_type)
        cfg = bot_configs.get(bot_type, {})

        if bot_type in ('msg_chat', 'msg_dm'):
            rot_items = (bot_configs.get('rotation') or {}).get('items', [])
            if cfg.get('_variants'):
                variants = cfg['_variants']
            elif rot_items:
                variants = [
                    {'id': i+1, 'text': it.get('text',''), 'media_type': it.get('media_type','none'), 'media_id': it.get('media_id')}
                    for i, it in enumerate(rot_items)
                ]
            else:
                variants = [{'id': 1, 'text': cfg.get('text', ''), 'media_type': cfg.get('media_type', 'none'), 'media_id': cfg.get('media_id'), 'media_pos': cfg.get('media_pos', 'above')}]
            site_configs[site_type] = {
                'variants':       variants,
                'currentVariant': 0,
                'msgTab':         'editor',
                'reply_target':   cfg.get('reply_target', 'none'),
                'keyboard':       _buttons_to_keyboard(cfg.get('buttons', [])),
                'settings':       cfg.get('_settings', {}),
            }

        elif bot_type == 'mute':
            # duration_seconds → value+unit для UI
            dur_sec = cfg.get('duration_seconds')
            if dur_sec is None:
                # Обратная совместимость со старым ключом '5m'/'60m' из бота
                _legacy = {'5m': 300, '15m': 900, '30m': 1800, '60m': 3600, '24h': 86400}
                dur_sec = _legacy.get(cfg.get('duration', '60m'), 3600)
            val, unit = _sec_to_val_unit(int(dur_sec))
            site_configs['mute'] = {
                'mute_target':      cfg.get('mute_target', 'initiator'),
                'mute_time_enabled': True,
                'mute_time_value':  val,
                'mute_time_unit':   unit,
                'mute_type':        cfg.get('mute_type', 'all'),
            }

        elif bot_type == 'ban':
            # duration_seconds → value+unit для UI. 0 = permanent → enabled=false
            dur_sec = int(cfg.get('duration_seconds', 0) or 0)
            if dur_sec > 0:
                val, unit = _sec_to_val_unit(dur_sec)
                ban_time_enabled = True
            else:
                val, unit = cfg.get('ban_time_value', 1), cfg.get('ban_time_unit', 'hours')
                ban_time_enabled = False
            site_configs['ban'] = {
                'ban_target':          cfg.get('ban_target', 'initiator'),
                'ban_time_enabled':    ban_time_enabled,
                'ban_time_value':      val,
                'ban_time_unit':       unit,
                'ban_revoke_messages': bool(cfg.get('revoke_messages', False)),
            }

        elif bot_type == 'emoji':
            # UI ждёт ключ 'emoji' (action.emoji); дублируем 'emoji_reaction' для совместимости со старым фронтом
            _em = cfg.get('emoji', '👍')
            site_configs['emoji'] = {
                'emoji':          _em,
                'emoji_reaction': _em,
                'emoji_target':   cfg.get('emoji_target', 'initiator'),
            }

        elif bot_type == 'warn':
            site_configs['warn'] = {
                'warn_target': cfg.get('warn_target', 'initiator'),
                'warn_notify': cfg.get('warn_notify', True),
                'warn_count':  cfg.get('warn_count', 1),
                'warn_period': cfg.get('warn_period', 0),
                'warn_mute_duration_seconds': int(cfg.get('warn_mute_duration_seconds') or 3600),
            }

        elif bot_type == 'delete':
            what = cfg.get('what', ['trigger_word'])
            if isinstance(what, list):
                if 'trigger_word' in what and 'quoted_message' in what:
                    del_target = 'both'
                elif 'quoted_message' in what:
                    del_target = 'replied'
                else:
                    del_target = 'initiator'
            else:
                del_target = 'initiator'
            site_configs['delete'] = {
                'delete_target':    del_target,
                'delete_delay':     cfg.get('delete_delay', 0),
                'delete_delay_unit': cfg.get('delete_delay_unit', 'seconds'),
            }

        elif bot_type == 'pin':
            site_configs['pin'] = {
                'pin_notify':       bool(cfg.get('notify', False)),
                'pin_target':       cfg.get('pin_target', '') or '',
                'pin_time_enabled': bool(cfg.get('pin_time_enabled', False)),
                'pin_time_value':   int(cfg.get('pin_time_value', 10) or 10),
                'pin_time_unit':    cfg.get('pin_time_unit', 'seconds') or 'seconds',
            }

        else:
            site_configs[site_type] = cfg

        # action_probability — round-trip обратно в UI
        if 'action_probability' in cfg and site_type in site_configs:
            site_configs[site_type]['action_probability'] = cfg.get('action_probability', 100)

    return site_actions, site_configs


def _row_to_trigger(row: dict) -> dict:
    actions = []
    try:
        actions = json.loads(row.get('actions') or '[]')
    except Exception:
        # backward compat: старый формат с одним action
        old_action = row.get('action')
        if old_action:
            actions = [old_action]

    action_configs = {}
    try:
        action_configs = json.loads(row.get('action_configs') or '{}')
    except Exception:
        pass

    site_actions, site_configs = _bot_to_site(actions, action_configs)

    cond_groups = []
    try:
        cg_raw = row.get('condition_groups')
        if cg_raw:
            cond_groups = json.loads(cg_raw)
    except Exception:
        cond_groups = []

    return {
        'id':                  row['id'],
        'name':                row['name'],
        'condition':           row.get('condition', 'contains'),
        'keywords':            row.get('keywords', ''),
        'probability':         row.get('probability', 100),
        'where_fires':         row.get('where_fires', 'all'),
        'initiator':           row.get('initiator', 'all'),
        'target':              row.get('target', 'nobody'),
        'target_user':         row.get('target_user', '') or '',
        'actions':             site_actions,
        'action_configs':      site_configs,
        'conditionGroups':     cond_groups,
        'bot_msg_delete':      row.get('bot_msg_delete', 'no'),
        'bot_msg_delete_after':row.get('bot_msg_delete_after') or 60,
        'fire_limit':          row.get('fire_limit') or 0,
        'auto_pin':            bool(row.get('auto_pin', 0)),
        'is_enabled':          bool(row.get('is_enabled', 1)),
    }


@app.get("/api/placeholders")
async def get_custom_placeholders():
    """Кастомные плейсхолдеры из БД (созданные вручную через бота)"""
    try:
        db.cursor.execute(
            "SELECT name, value, description FROM custom_placeholders ORDER BY name"
        )
        rows = db.cursor.fetchall()
        return [{'name': r['name'], 'value': r['value'], 'description': r['description'] or ''} for r in rows]
    except Exception as e:
        return []  # Таблица может не существовать — возвращаем пустой список


@app.get("/api/topics")
async def get_topics():
    """Ветки (топики) группы из БД"""
    try:
        chat_id = int(os.getenv('TARGET_CHAT_ID', 0))
        rows = db.get_all_topics(chat_id)
        result = []
        for r in rows:
            result.append({
                'thread_id':   r['thread_id'],
                'name':        r['thread_name'] or f'Ветка #{r["thread_id"]}',
                'is_main':     bool(r['is_main_thread']),
                'total_msgs':  int(r['total_messages'] or 0),
            })
        return result
    except Exception as e:
        logger.warning(f"get_topics error: {e}")
        return []


# ── Медиа-загрузка ──────────────────────────────────────────────
MEDIA_DIR = os.path.join(current_dir, 'Admin_SITE', 'media_uploads')
try:
    os.makedirs(MEDIA_DIR, exist_ok=True)
    from fastapi.staticfiles import StaticFiles
    app.mount("/media", StaticFiles(directory=MEDIA_DIR), name="media")
    logger.info(f"✅ Медиа-хранилище: {MEDIA_DIR}")
except Exception as _me:
    logger.warning(f"⚠️ Медиа-хранилище недоступно: {_me}")


@app.post("/api/media/upload")
async def upload_media(file: UploadFile = File(...)):
    """Загрузка медиафайла. Возвращает url для предпросмотра и server_path."""
    import uuid
    allowed = {'image/jpeg', 'image/png', 'image/gif', 'image/webp',
               'video/mp4', 'video/mpeg', 'video/quicktime'}
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Недопустимый тип файла")
    ext = os.path.splitext(file.filename or '')[1] or '.bin'
    fname = f"{uuid.uuid4().hex}{ext}"
    fpath = os.path.join(MEDIA_DIR, fname)
    try:
        contents = await file.read()
        with open(fpath, 'wb') as fh:
            fh.write(contents)
        # Определяем тип
        ct = file.content_type or ''
        if ct.startswith('video'):
            media_type = 'video'
        elif ct == 'image/gif':
            media_type = 'animation'
        else:
            media_type = 'photo'
        return {
            'url':         f'/media/{fname}',
            'server_path': fpath,
            'media_type':  media_type,
            'filename':    fname,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/triggers")
async def get_triggers():
    """Список триггеров из БД"""
    try:
        db.cursor.execute("SELECT * FROM triggers ORDER BY id DESC")
        return [_row_to_trigger(dict(r)) for r in db.cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/triggers")
async def create_trigger(t: TriggerIn):
    """Создать триггер"""
    try:
        bot_actions, bot_configs = _site_to_bot(t.actions, t.action_configs)
        db.cursor.execute('''
            INSERT INTO triggers
                (name, keywords, condition, probability,
                 where_fires, initiator, target, target_user,
                 actions, action_configs,
                 bot_msg_delete, bot_msg_delete_after,
                 fire_limit, auto_pin, warn_period, is_enabled,
                 condition_groups)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            t.name, t.keywords, t.condition, t.probability,
            t.where_fires, t.initiator, t.target, _resolve_target_user_raw(t.target_user),
            json.dumps(bot_actions), json.dumps(bot_configs),
            t.bot_msg_delete, t.bot_msg_delete_after,
            t.fire_limit, int(t.auto_pin), t.warn_period, int(t.is_enabled),
            json.dumps(t.condition_groups) if t.condition_groups else None,
        ))
        db.conn.commit()
        return {'id': db.cursor.lastrowid, 'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/triggers/{trigger_id}")
async def update_trigger(trigger_id: int, t: TriggerIn):
    """Обновить триггер"""
    try:
        bot_actions, bot_configs = _site_to_bot(t.actions, t.action_configs)
        db.cursor.execute('''
            UPDATE triggers SET
                name=?, keywords=?, condition=?, probability=?,
                where_fires=?, initiator=?, target=?, target_user=?,
                actions=?, action_configs=?,
                bot_msg_delete=?, bot_msg_delete_after=?,
                fire_limit=?, auto_pin=?, warn_period=?, is_enabled=?,
                condition_groups=?
            WHERE id=?
        ''', (
            t.name, t.keywords, t.condition, t.probability,
            t.where_fires, t.initiator, t.target, _resolve_target_user_raw(t.target_user),
            json.dumps(bot_actions), json.dumps(bot_configs),
            t.bot_msg_delete, t.bot_msg_delete_after,
            t.fire_limit, int(t.auto_pin), t.warn_period, int(t.is_enabled),
            json.dumps(t.condition_groups) if t.condition_groups else None,
            trigger_id,
        ))
        db.conn.commit()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/triggers/{trigger_id}")
async def delete_trigger(trigger_id: int):
    """Удалить триггер"""
    try:
        db.cursor.execute("DELETE FROM triggers WHERE id=?", (trigger_id,))
        db.conn.commit()
        return {'success': True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/triggers/{trigger_id}/toggle")
async def toggle_trigger(trigger_id: int):
    """Переключить активность триггера"""
    try:
        db.cursor.execute("SELECT is_enabled FROM triggers WHERE id=?", (trigger_id,))
        row = db.cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Trigger not found")
        new_val = 0 if row[0] else 1
        db.cursor.execute("UPDATE triggers SET is_enabled=? WHERE id=?", (new_val, trigger_id))
        db.conn.commit()
        return {'id': trigger_id, 'is_enabled': bool(new_val)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/triggers/{trigger_id}/copy")
async def copy_trigger(trigger_id: int):
    """Копировать триггер (создаёт выключенную копию)"""
    try:
        db.cursor.execute("SELECT * FROM triggers WHERE id=?", (trigger_id,))
        row = db.cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Trigger not found")
        r = dict(row)
        db.cursor.execute('''
            INSERT INTO triggers
                (name, keywords, condition, probability,
                 where_fires, initiator, target, target_user,
                 actions, action_configs,
                 bot_msg_delete, bot_msg_delete_after,
                 fire_limit, auto_pin, warn_period, condition_groups, is_enabled)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)
        ''', (
            r['name'] + ' (копия)',
            r.get('keywords', ''), r.get('condition', 'contains'),
            r.get('probability', 100),
            r.get('where_fires', 'all'), r.get('initiator', 'all'),
            r.get('target', 'nobody'), r.get('target_user', '') or '',
            r.get('actions', '[]'), r.get('action_configs', '{}'),
            r.get('bot_msg_delete', 'no'), r.get('bot_msg_delete_after', 60),
            r.get('fire_limit', 0), r.get('auto_pin', 0),
            r.get('warn_period', 0), r.get('condition_groups'),
        ))
        db.conn.commit()
        return {'id': db.cursor.lastrowid, 'success': True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/journal")
async def get_journal():
    """Журнал событий из journal_messages"""
    try:
        EVENT_TAG = {
            'join':    ('#Вход',    'join'),
            'leave':   ('#Выход',   'leave'),
            'mute':    ('#Мут',     'mute'),
            'ban':     ('#Бан',     'ban'),
            'warn':    ('#Варн',    'warn'),
            'kick':    ('#Кик',     'ban'),
            'unban':   ('#Разбан',  'unban'),
            'unmute':  ('#Размут',  'unban'),
            'trigger': ('#Триггер', 'trigger'),
        }

        db.cursor.execute('''
            SELECT jm.id, jm.event_type, jm.user_id, jm.text_preview, jm.created_at,
                   u.username, u.first_name
            FROM journal_messages jm
            LEFT JOIN users u ON jm.user_id = u.user_id
            ORDER BY jm.id DESC
            LIMIT 100
        ''')
        entries = []
        for r in (dict(x) for x in db.cursor.fetchall()):
            ev = r.get('event_type') or 'other'
            tag, typ = EVENT_TAG.get(ev, (f'#{ev}', ev))
            uname = r.get('username') or r.get('first_name') or (str(r['user_id']) if r['user_id'] else '—')
            entries.append({
                'id':      r['id'],
                'time':    (r['created_at'] or '')[:16],
                'type':    typ,
                'tag':     tag,
                'user':    f"@{uname}" if r.get('username') else uname,
                'user_id': r.get('user_id', 0),
                'text':    _clean_journal_html(r.get('text_preview') or ev),
            })

        return entries
    except Exception as e:
        logger.error(f"Error in /api/journal: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Кэш аватаров (user_id -> (bytes, timestamp)) ──────────────
_avatar_cache: dict = {}
_AVATAR_TTL = 86400  # 24 часа


class JournalActionIn(BaseModel):
    user_id: int
    action: str  # ban | kick | mute_forever | unmute | unban


@app.post("/api/journal/action")
async def journal_action(body: JournalActionIn, authorization: str = Header(default=None)):
    """Действия модерации из записей журнала: ban/kick/mute_forever/unmute/unban."""
    payload = _require_auth(authorization)
    caller_id = int(payload.get("user_id", 0))
    caller_role = _resolve_user_role(caller_id)
    if caller_role not in ("owner", "deputy", "admin"):
        raise HTTPException(status_code=403, detail="Нет прав")

    target_chat_id = int(os.getenv("TARGET_CHAT_ID", 0))
    if not target_chat_id:
        raise HTTPException(status_code=500, detail="TARGET_CHAT_ID не задан")

    uid = body.user_id
    action = body.action
    tg_url = f"https://api.telegram.org/bot{_BOT_TOKEN}"

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            if action == "ban":
                r = await client.post(f"{tg_url}/banChatMember", json={"chat_id": target_chat_id, "user_id": uid})
            elif action == "kick":
                r = await client.post(f"{tg_url}/banChatMember", json={"chat_id": target_chat_id, "user_id": uid})
                if r.json().get("ok"):
                    await client.post(f"{tg_url}/unbanChatMember", json={"chat_id": target_chat_id, "user_id": uid, "only_if_banned": True})
            elif action == "mute_forever":
                r = await client.post(f"{tg_url}/restrictChatMember", json={
                    "chat_id": target_chat_id, "user_id": uid,
                    "permissions": {"can_send_messages": False, "can_send_audios": False,
                                    "can_send_documents": False, "can_send_photos": False,
                                    "can_send_videos": False, "can_send_video_notes": False,
                                    "can_send_voice_notes": False, "can_send_polls": False,
                                    "can_send_other_messages": False, "can_add_web_page_previews": False},
                })
            elif action == "unmute":
                r = await client.post(f"{tg_url}/restrictChatMember", json={
                    "chat_id": target_chat_id, "user_id": uid,
                    "permissions": {"can_send_messages": True, "can_send_audios": True,
                                    "can_send_documents": True, "can_send_photos": True,
                                    "can_send_videos": True, "can_send_video_notes": True,
                                    "can_send_voice_notes": True, "can_send_polls": True,
                                    "can_send_other_messages": True, "can_add_web_page_previews": True},
                })
            elif action == "unban":
                r = await client.post(f"{tg_url}/unbanChatMember", json={"chat_id": target_chat_id, "user_id": uid, "only_if_banned": False})
            else:
                raise HTTPException(status_code=400, detail=f"Неизвестное действие: {action}")

        result = r.json()
        if result.get("ok"):
            return {"ok": True}
        else:
            return {"ok": False, "error": result.get("description", "Ошибка Telegram")}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"journal_action error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/user/{user_id}/avatar")
async def get_user_avatar(user_id: int):
    """Прокси-аватар пользователя через TG Bot API. Кэш 24ч."""
    now = time.time()
    cached = _avatar_cache.get(user_id)
    if cached and now - cached[1] < _AVATAR_TTL:
        from fastapi.responses import Response
        return Response(content=cached[0], media_type="image/jpeg",
                        headers={"Cache-Control": "public, max-age=86400"})

    try:
        tg_url = f"https://api.telegram.org/bot{_BOT_TOKEN}"
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"{tg_url}/getUserProfilePhotos", params={"user_id": user_id, "limit": 1})
            data = r.json()
            if not data.get("ok") or not data["result"]["total_count"]:
                from fastapi.responses import Response
                return Response(status_code=204)
            file_id = data["result"]["photos"][0][-1]["file_id"]

            r2 = await client.get(f"{tg_url}/getFile", params={"file_id": file_id})
            file_path = r2.json()["result"]["file_path"]

            r3 = await client.get(f"https://api.telegram.org/file/bot{_BOT_TOKEN}/{file_path}")
            img_bytes = r3.content

        _avatar_cache[user_id] = (img_bytes, now)
        from fastapi.responses import Response
        return Response(content=img_bytes, media_type="image/jpeg",
                        headers={"Cache-Control": "public, max-age=86400"})
    except Exception as e:
        logger.warning(f"get_user_avatar error uid={user_id}: {e}")
        from fastapi.responses import Response
        return Response(status_code=204)


# ── UI-настройки (цитаты журнала и др.) ───────────────────────
_UI_SETTINGS_KEYS = {
    "journal_quote_bg",
    "journal_quote_stripe_mode",
    "journal_quote_stripe_color1",
    "journal_quote_stripe_color2",
}

@app.get("/api/ui_settings")
async def get_ui_settings():
    """Публичные UI-настройки: цвета цитат журнала."""
    defaults = {
        "journal_quote_bg": "#fff7ed",
        "journal_quote_stripe_mode": "solid",
        "journal_quote_stripe_color1": "#fdba74",
        "journal_quote_stripe_color2": "#f87171",
    }
    if not db:
        return defaults
    result = {}
    for key, default in defaults.items():
        try:
            result[key] = db.get_setting(key, default)
        except Exception:
            result[key] = default
    return result


@app.post("/api/ui_settings")
async def save_ui_settings(request: Request, authorization: str = Header(default=None)):
    """Сохраняет UI-настройки. Только owner."""
    _require_owner(authorization)
    body = await request.json()
    saved = []
    errors = []
    for key, val in body.items():
        if key in _UI_SETTINGS_KEYS and isinstance(val, str):
            try:
                db.set_setting(key, val)
                saved.append(key)
            except Exception as e1:
                # Fallback: INSERT OR REPLACE без updated_at (старые схемы БД)
                try:
                    db.cursor.execute(
                        'INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
                        (key, val)
                    )
                    db.conn.commit()
                    saved.append(key)
                except Exception as e2:
                    logger.warning(f"ui_settings save error {key}: {e1} / {e2}")
                    errors.append(key)
    if errors:
        return {"ok": False, "saved": saved, "errors": errors}
    return {"ok": True, "saved": saved}


# --- GEMINI AI ---

GEMINI_KEY = os.environ.get('GEMINI_API_KEY', '')
GEMINI_URL = 'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent'

GEMINI_SYSTEM = """Ты — умный ИИ-ассистент встроенный в панель управления Telegram-чатом "PULSE 4ever 18+".
Помогаешь владельцу чата:
- Писать тексты для триггеров (авто-ответы на сообщения)
- Составлять тексты предупреждений, мутов, банов
- Готовить рассылки для участников
- Анализировать ситуации в чате и давать советы по модерации
- Отвечать на любые вопросы по управлению чатом

Отвечай кратко, по делу, на русском языке. Форматируй ответ — используй списки и абзацы."""

class AiRequest(BaseModel):
    prompt: str
    context: str = ''   # опциональный контекст (например, текущая статистика)

@app.post("/api/ai")
async def ai_chat(req: AiRequest):
    if not GEMINI_KEY:
        raise HTTPException(status_code=503, detail="GEMINI_API_KEY не задан на сервере")
    try:
        full_prompt = GEMINI_SYSTEM
        if req.context:
            full_prompt += f"\n\nТекущий контекст панели:\n{req.context}"
        full_prompt += f"\n\nЗапрос пользователя:\n{req.prompt}"

        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                f"{GEMINI_URL}?key={GEMINI_KEY}",
                json={"contents": [{"parts": [{"text": full_prompt}]}]}
            )
            data = resp.json()
            logger.info(f"Gemini response status: {resp.status_code}, keys: {list(data.keys())}")
            if resp.status_code != 200:
                logger.error(f"Gemini error body: {data}")
                raise HTTPException(status_code=502, detail=data.get('error', {}).get('message', f'HTTP {resp.status_code}'))

        if 'candidates' not in data:
            logger.error(f"Gemini unexpected response: {data}")
            raise HTTPException(status_code=502, detail=f"Gemini: {data.get('error', {}).get('message', 'нет candidates')}")
        text = data['candidates'][0]['content']['parts'][0]['text']
        return {"result": text}
    except HTTPException:
        raise
    except httpx.HTTPStatusError as e:
        logger.error(f"Gemini API HTTP error {e.response.status_code}: {e.response.text}")
        raise HTTPException(status_code=502, detail=f"Gemini HTTP {e.response.status_code}")
    except Exception as e:
        logger.error(f"AI error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- АВТООБНОВЛЕНИЯ ---

AI_UPDATES_FILE = 'ai_updates.json'
DEPLOY_SECRET   = os.environ.get('DEPLOY_SECRET', 'pulse-deploy-secret')

def _load_ai_updates() -> list:
    try:
        if os.path.exists(AI_UPDATES_FILE):
            with open(AI_UPDATES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return []

def _save_ai_updates(data: list):
    with open(AI_UPDATES_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.get("/api/updates")
async def get_updates():
    """Список AI-сгенерированных заметок об обновлениях"""
    return _load_ai_updates()

@app.delete("/api/updates/{update_id}")
async def delete_update(update_id: int):
    """Удалить запись об обновлении"""
    updates = _load_ai_updates()
    new_list = [u for u in updates if u.get('id') != update_id]
    if len(new_list) == len(updates):
        raise HTTPException(status_code=404, detail="Not found")
    _save_ai_updates(new_list)
    return {"ok": True}

class DeployEvent(BaseModel):
    commit_message: str
    secret: str

@app.post("/api/updates/generate")
async def generate_update(event: DeployEvent):
    """Вызывается после деплоя — генерирует user-friendly заметку через Gemini"""
    if event.secret != DEPLOY_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    if not GEMINI_KEY:
        raise HTTPException(status_code=503, detail="GEMINI_API_KEY не задан")

    # Определяем версию из коммита (feat(V1.11) → V1.11, feat(V1.11.0a) → V1.11.0a)
    ver_match = re.search(r'V[\d]+\.[\d]+(?:\.[\d]+)?[a-z]?', event.commit_message)
    version = ver_match.group(0) if ver_match else ''

    # Определяем тег по ключевым словам
    msg_lower = event.commit_message.lower()
    if any(w in msg_lower for w in ['сайт', 'site', 'панел', 'dashboard', 'ui']):
        tag = 'site'
    elif any(w in msg_lower for w in ['триггер', 'trigger']):
        tag = 'triggers'
    elif any(w in msg_lower for w in ['журнал', 'journal']):
        tag = 'journal'
    elif any(w in msg_lower for w in ['статист', 'stat']):
        tag = 'statistics'
    else:
        tag = 'bot'

    # Определяем тип (fix/feat/improve)
    if event.commit_message.startswith('fix'):
        upd_type = 'fix'
    elif event.commit_message.startswith('feat'):
        upd_type = 'new'
    else:
        upd_type = 'improve'

    prompt = f"""Ты пишешь краткие заметки об обновлениях для Telegram-бота и веб-панели управления чатом "PULSE 4ever 18+".

Техническое описание коммита: "{event.commit_message}"

Напиши 2-4 коротких пункта на русском языке, понятных обычному пользователю (без технических терминов).
Каждый пункт начинай с одного из слов: «Добавлено», «Исправлено», «Улучшено» или «Теперь».
ВАЖНО: не используй markdown-разметку, звёздочки **, решётки # и другие спецсимволы.
Верни ТОЛЬКО список пунктов, каждый на новой строке, без заголовков и нумерации."""

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                f"{GEMINI_URL}?key={GEMINI_KEY}",
                json={"contents": [{"parts": [{"text": prompt}]}]}
            )
            try:
                data = resp.json()
            except Exception:
                logger.error(f"Gemini non-JSON response {resp.status_code}: {resp.text[:500]}")
                raise HTTPException(status_code=502, detail=f"Gemini вернул не-JSON (HTTP {resp.status_code})")
            if resp.status_code != 200:
                err_msg = (data.get('error') or {}).get('message') or f"HTTP {resp.status_code}"
                logger.error(f"Gemini API error: {err_msg} | raw: {str(data)[:300]}")
                raise HTTPException(status_code=502, detail=f"Gemini: {err_msg}")

        candidates = data.get('candidates') or []
        if not candidates:
            block = (data.get('promptFeedback') or {}).get('blockReason', 'пустой ответ')
            logger.error(f"Gemini без candidates: {str(data)[:300]}")
            raise HTTPException(status_code=502, detail=f"Gemini не ответил: {block}")
        try:
            raw = candidates[0]['content']['parts'][0]['text']
        except (KeyError, IndexError, TypeError) as pe:
            logger.error(f"Gemini parse error {pe}: {str(candidates[0])[:300]}")
            raise HTTPException(status_code=502, detail=f"Неожиданная структура ответа Gemini")
        # Убираем markdown: **, *, #, _ и т.д.
        raw = re.sub(r'\*{1,2}|_{1,2}|#{1,3}', '', raw)
        lines = [re.sub(r'^[-•\d.]+\s*', '', l).strip() for l in raw.split('\n') if l.strip()]

        # Определяем тип каждой строки по первому слову
        def _line_type(line: str) -> str:
            l = line.lower()
            if l.startswith('исправл'):
                return 'fix'
            elif l.startswith('улучш'):
                return 'improve'
            else:
                return 'new'  # «Добавлено», «Теперь» и всё остальное

        items = [{"type": _line_type(l), "text": l, "tag": tag} for l in lines]

        # Очищаем заголовок: убираем «feat(V1.11):» и лишние пробелы
        clean_title = re.sub(r'^(?:feat|fix|improve|chore|docs|refactor)\([^)]+\):\s*', '', event.commit_message).strip()

        entry = {
            "id":          int(_now_msk().timestamp()),
            "date":        _now_msk().strftime('%d.%m.%Y'),
            "version":     version,
            "title":       clean_title[:100],
            "tag":         tag,
            "type":        upd_type,
            "items":       items,
            "aiGenerated": True,
        }
        updates = [entry] + _load_ai_updates()
        _save_ai_updates(updates[:50])
        return entry
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"generate_update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ─── УПРАВЛЕНИЕ ФУНКЦИЯМИ ───────────────────────────────────────────────────

FEATURES_LIST = [
    {'id': 'shipper',       'name': '💘 Шиппер (Рулетка пар)'},
    {'id': 'profile',       'name': '👤 Личный кабинет (Профиль)'},
    {'id': 'statistics',    'name': '📊 Статистика'},
    {'id': 'top',           'name': '🏆 ТОП-5 и команды'},
    {'id': 'bank',          'name': '🏦 Центробанк'},
    {'id': 'activities',    'name': '🎯 Активности'},
    {'id': 'detalization',  'name': '📋 Детализация'},
    {'id': 'lottery',       'name': '🎰 Лотерея'},
    {'id': 'bingo',         'name': '🎱 Бинго'},
    {'id': 'referral',      'name': '👥 Рефералы'},
    {'id': 'donate',        'name': '🎁 Донаты'},
    {'id': 'monthly_gift',  'name': '🎁 Подарок Месяца'},
    {'id': 'horoscope',     'name': '🔮 Гороскоп'},
    {'id': 'bbs',           'name': '❣️ Pulse BBS'},
    {'id': 'bbs_other',     'name': '📦 BBS: Другое'},
    {'id': 'bbs_edit',      'name': '✏️ Редактирование анкет BBS'},
    {'id': 'registration',  'name': '📝 Регистрация новых участников'},
]

def _get_feature_enabled(feature_id: str) -> bool:
    if not db:
        return False
    if feature_id == 'shipper':
        return db.get_setting('shipper_enabled', '0') == '1'
    return db.is_feature_enabled(feature_id)

def _set_feature_enabled(feature_id: str, enabled: bool):
    val = '1' if enabled else '0'
    db.set_setting(f'feature_{feature_id}', val)
    if feature_id == 'top':
        db.set_setting('feature_top_commands', val)
    if feature_id == 'shipper':
        db.set_setting('shipper_enabled', val)

@app.get("/api/features")
async def get_features():
    """Список функций бота с их состоянием вкл/выкл"""
    if not db:
        raise HTTPException(status_code=503, detail="DB unavailable")
    result = []
    for f in FEATURES_LIST:
        result.append({**f, 'enabled': _get_feature_enabled(f['id'])})
    return result

@app.post("/api/features/{feature_id}/toggle")
async def toggle_feature_api(feature_id: str):
    """Переключить функцию вкл/выкл"""
    if not db:
        raise HTTPException(status_code=503, detail="DB unavailable")
    if not any(f['id'] == feature_id for f in FEATURES_LIST):
        raise HTTPException(status_code=404, detail="Unknown feature")
    current = _get_feature_enabled(feature_id)
    _set_feature_enabled(feature_id, not current)
    return {'id': feature_id, 'enabled': not current}


# ─────────────── STAFF (администраторы) ────────────────

class StaffAddRequest(BaseModel):
    user_id: str  # принимаем строку — может быть числом или @username


@app.get("/api/staff")
async def get_staff():
    """Список владельца и всех администраторов"""
    if not db:
        raise HTTPException(status_code=503, detail="DB unavailable")
    try:
        conn = db.get_connection()
        rows = conn.execute(
            "SELECT user_id, username, first_name, is_owner, is_admin "
            "FROM users WHERE is_admin = 1 OR is_owner = 1 "
            "ORDER BY is_owner DESC, first_name"
        ).fetchall()
        result = []
        for r in rows:
            result.append({
                "user_id":    r["user_id"],
                "username":   r["username"] or "",
                "first_name": r["first_name"] or "",
                "is_owner":   bool(r["is_owner"]),
                "is_admin":   bool(r["is_admin"]),
            })
        return result
    except Exception as e:
        logger.error(f"get_staff error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/staff")
async def add_staff(req: StaffAddRequest):
    """Назначить пользователя администратором по user_id или @username"""
    if not db:
        raise HTTPException(status_code=503, detail="DB unavailable")
    try:
        conn = db.get_connection()
        uid_str = req.user_id.strip().lstrip('@')
        # Пробуем найти по числовому ID или по username
        if uid_str.isdigit():
            row = conn.execute(
                "SELECT user_id, username, first_name, is_owner FROM users WHERE user_id = ?",
                (int(uid_str),)
            ).fetchone()
        else:
            row = conn.execute(
                "SELECT user_id, username, first_name, is_owner FROM users "
                "WHERE lower(username) = lower(?)", (uid_str,)
            ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Пользователь не найден в базе")
        if row["is_owner"]:
            raise HTTPException(status_code=400, detail="Нельзя изменить права Владельца")
        conn.execute("UPDATE users SET is_admin = 1 WHERE user_id = ?", (row["user_id"],))
        conn.commit()
        return {"ok": True, "user_id": row["user_id"]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"add_staff error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/staff/{user_id}")
async def remove_staff(user_id: int):
    """Снять права администратора"""
    if not db:
        raise HTTPException(status_code=503, detail="DB unavailable")
    try:
        conn = db.get_connection()
        row = conn.execute(
            "SELECT is_owner FROM users WHERE user_id = ?", (user_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        if row["is_owner"]:
            raise HTTPException(status_code=400, detail="Нельзя снять права Владельца")
        conn.execute("UPDATE users SET is_admin = 0 WHERE user_id = ?", (user_id,))
        conn.commit()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"remove_staff error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Economy WebSocket ──────────────────────────────────────────────────────────

@app.websocket("/api/ws/economy")
async def ws_economy(websocket: WebSocket, token: str = Query(...)):
    try:
        payload = _decode_jwt(token)
    except Exception:
        await websocket.close(code=4003)
        return

    user_id = int(payload.get("user_id", 0))
    role = _resolve_user_role(user_id)
    from permissions import has_permission
    if not has_permission(role, "economy.view"):
        await websocket.close(code=4003)
        return

    if _economy_ws_manager is None:
        await websocket.close(code=4503)
        return

    await _economy_ws_manager.connect(websocket)
    try:
        while True:
            try:
                await websocket.receive_text()
            except WebSocketDisconnect:
                break
    finally:
        await _economy_ws_manager.disconnect(websocket)


async def _metrics_broadcaster():
    """Раз в 30 сек рассылает live-метрики всем подключённым клиентам."""
    import asyncio
    while True:
        await asyncio.sleep(30)
        try:
            if _economy_ws_manager and db:
                from database.db_economy import get_economy_metrics
                metrics = get_economy_metrics(db)
                await _economy_ws_manager.broadcast({"event": "metrics_update", **metrics})
        except Exception as e:
            logger.error(f"_metrics_broadcaster error: {e}")


@app.on_event("startup")
async def _on_startup():
    import asyncio
    asyncio.create_task(_metrics_broadcaster())
    logger.info("✅ economy metrics broadcaster запущен")


# --- ЗАПУСК ---
if __name__ == "__main__":
    # Запуск uvicorn напрямую. reload=True включен для удобства разработки.
    uvicorn.run(app, host="0.0.0.0", port=8000)