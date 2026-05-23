"""
Сканер проекта → Pulse_Auto_Map.xlsx.

Обходит репозиторий, для каждого .py / .js / .jsx достаёт описание
(docstring для Python, // или /* */ коммент-шапка для JS) и собирает
Excel-карту: Категория · Путь к файлу · Тип · Описание.

Запускается воркфлоу update_map.yml на каждый push в main с изменениями
кода — результат коммитится обратно в main, поэтому свежая карта приезжает
обычным `git pull`.
"""
import os
import ast
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Папки и расширения, которые мы сканируем
IGNORED_DIRS = {'.git', 'venv', '.venv', '__pycache__', 'node_modules', 'dist', 'build', 'logs'}
ALLOWED_EXTENSIONS = {'.py', '.js', '.jsx'}

def get_python_docstring(filepath):
    """Вытаскивает описание из Python-файла (docstrings или первые комментарии)."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Пытаемся вытащить официальный docstring (текст в """ """ в начале файла)
        try:
            tree = ast.parse(content)
            doc = ast.get_docstring(tree)
            if doc:
                # Возвращаем первые 500 символов, чтобы не порвать Эксель
                return doc.strip()[:500] + ('...' if len(doc) > 500 else '')
        except SyntaxError:
            pass
            
        # Если docstring нет, ищем обычные комментарии # в первых 15 строках
        lines = content.split('\n')[:15]
        comments =[l.replace('#', '').strip() for l in lines if l.strip().startswith('#') and not l.startswith('#!')]
        if comments:
            return " ".join(comments)[:500]
            
    except Exception as e:
        print(f"⚠️ Ошибка при чтении {filepath}: {e}")
    return "⚠️ Описание в коде не найдено. Добавьте \"\"\"описание\"\"\" в начало файла."

def get_js_docstring(filepath):
    """Вытаскивает описание из JS/JSX файлов."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()[:20] # Читаем первые 20 строк
        comments =[]
        in_block = False
        for line in lines:
            l = line.strip()
            if l.startswith('/*'): in_block = True
            if in_block:
                clean = l.replace('/*', '').replace('*/', '').replace('*', '').strip()
                if clean: comments.append(clean)
            if '*/' in l: in_block = False
            
            if not in_block and l.startswith('//'):
                comments.append(l.replace('//', '').strip())
                
        if comments:
            return " ".join(comments)[:500]
    except Exception as e:
        print(f"⚠️ Ошибка при чтении {filepath}: {e}")
    return "⚠️ Описание в коде не найдено."

def determine_category(rel_path):
    """Определяет логическую категорию файла по его папке."""
    parts = rel_path.replace('\\', '/').split('/')
    if 'database' in parts: return "🗄 База Данных"
    if 'BBS' in parts: return "💘 Знакомства (BBS)"
    if 'Admin_SITE' in parts: return "🌐 Frontend (React Сайт)"
    if 'utils' in parts: return "🛠 Утилиты и ИИ"
    if 'commands' in parts: return "⌨️ Команды"
    if 'handlers' in parts: return "⚙️ Хендлеры (Логика)"
    if 'api' in parts or 'api.py' in parts: return "🔌 Бэкенд (API)"
    if rel_path in ['bot.py', 'main.py']: return "🚀 Ядро (Core)"
    return "📁 Прочее"

def _safe(v):
    """openpyxl интерпретирует строки, начинающиеся с =/+/-/@, как формулы.
    Excel при открытии видит «битую формулу» → ругается «Удалены формулы».
    Префикс zero-width-space безопасно отключает это: для глаз не видно,
    парсер формул больше не срабатывает."""
    if isinstance(v, str) and v and v[0] in '=+-@':
        return '​' + v
    return v


def create_dynamic_project_map():
    print("🔍 Начинаю сканирование проекта...")
    
    current_dir = os.getcwd()
    files_data =[]

    # Обходим все папки и файлы
    for root, dirs, files in os.walk(current_dir):
        # Исключаем мусорные папки
        dirs[:] =[d for d in dirs if d not in IGNORED_DIRS]
        
        for file in files:
            ext = os.path.splitext(file)[1].lower()
            if ext in ALLOWED_EXTENSIONS:
                try:
                    filepath = os.path.join(root, file)
                    rel_path = os.path.relpath(filepath, current_dir)
                    
                    # Извлекаем данные
                    category = determine_category(rel_path)
                    if ext == '.py':
                        desc = get_python_docstring(filepath)
                    else:
                        desc = get_js_docstring(filepath)
                        
                    files_data.append((category, rel_path, ext, desc))
                except Exception as e:
                    print(f"❌ Ошибка при обработке {file}: {e}")
                    continue

    # Сортируем по категориям для красоты
    files_data.sort(key=lambda x: x[0])

    print(f"✅ Найдено файлов: {len(files_data)}. Формирую Excel...")

    # --- Создание Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Архитектура Проекта"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Шапка
    headers =["Категория", "Путь к файлу", "Тип", "Описание (вытащено из файла)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    # Размеры колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 80

    # Заполнение
    for i, data in enumerate(files_data, start=2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=i, column=col_idx, value=_safe(value))
            cell.border = border_thin
            cell.alignment = align_left if col_idx in [2, 4] else align_center

    # Включаем фильтры
    ws.auto_filter.ref = f"A1:D{len(files_data)+1}"
    ws.freeze_panes = "A2"

    wb.save("Pulse_Auto_Map.xlsx")
    print("🎉 Готово! Файл 'Pulse_Auto_Map.xlsx' успешно создан.")

if __name__ == "__main__":
    create_dynamic_project_map()
    # trigger action +
