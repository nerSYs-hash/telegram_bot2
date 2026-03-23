#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import math
from datetime import datetime, timedelta
import pytz

def calculate_mining_reward(active_core_count, difficulty_k=5.0):
    """
    Calculate mining reward based on active core
    Formula: K / sqrt(Active Core)
    """
    if active_core_count == 0:
        active_core_count = 1
    
    reward = difficulty_k / math.sqrt(active_core_count)
    return round(reward, 2)

def calculate_user_activity_score(user_stats, chat_stats):
    """
    Calculate user activity score (АктП) based on formula
    """
    if not chat_stats or chat_stats['total_messages'] == 0:
        return 0.0
    
    score = 0.0
    
    # 1. ОКСП - Total characters (1%)
    if user_stats.get('total_chars', 0) > 0 and chat_stats.get('total_chars', 0) > 0:
        score += 0.01 * (chat_stats['total_chars'] / user_stats['total_chars'])
    
    # 2. СДСП - Average message length (7%)
    user_avg_len = user_stats.get('total_chars', 0) / max(user_stats.get('total_messages', 1), 1)
    chat_avg_len = chat_stats.get('total_chars', 0) / max(chat_stats.get('total_messages', 1), 1)
    if user_avg_len > 0 and chat_avg_len > 0:
        score += 0.07 * (chat_avg_len / user_avg_len)
    
    # 3. КСП - Word count (1%)
    if user_stats.get('total_words', 0) > 0 and chat_stats.get('total_words', 0) > 0:
        score += 0.01 * (chat_stats['total_words'] / user_stats['total_words'])
    
    # 4. КОРП - Reactions given (8%)
    if user_stats.get('reactions_given', 0) > 0 and chat_stats.get('total_reactions', 0) > 0:
        score += 0.08 * (chat_stats['total_reactions'] / user_stats['reactions_given'])
    
    # 5. КПРП - Reactions received (8%)
    if user_stats.get('reactions_received', 0) > 0 and chat_stats.get('total_reactions', 0) > 0:
        score += 0.08 * (chat_stats['total_reactions'] / user_stats['reactions_received'])
    
    # 6. КОПЮП - Replies received (20%)
    if user_stats.get('replies_received', 0) > 0 and chat_stats.get('total_replies', 0) > 0:
        score += 0.20 * (chat_stats['total_replies'] / user_stats['replies_received'])
    
    # 7. КОПЯП - Replies sent (20%)
    if user_stats.get('replies_sent', 0) > 0 and chat_stats.get('total_replies', 0) > 0:
        score += 0.20 * (chat_stats['total_replies'] / user_stats['replies_sent'])
    
    # 8. КУПП - Mentions received (20%)
    if user_stats.get('mentions_received', 0) > 0 and chat_stats.get('total_mentions', 0) > 0:
        score += 0.20 * (chat_stats['total_mentions'] / user_stats['mentions_received'])
    
    # 9. МедиаП - Media sent (7%)
    if user_stats.get('media_sent', 0) > 0 and chat_stats.get('total_media', 0) > 0:
        score += 0.07 * (chat_stats['total_media'] / user_stats['media_sent'])
    
    # 10. ПИВДВП - Other threads posts (8%)
    if user_stats.get('other_threads_posts', 0) > 0 and chat_stats.get('other_threads_posts', 0) > 0:
        score += 0.08 * (chat_stats['other_threads_posts'] / user_stats['other_threads_posts'])
    
    # 11. ПЗП - Warnings penalty (REMOVED - no longer affects activity)
    # Warnings are tracked but do NOT reduce activity score
    
    return round(score, 2)

def format_number(num):
    """Format number with thousands separator and 2 decimal places for pulses"""
    num = float(num)
    # Format with 2 decimal places and space as thousand separator
    formatted = "{:,.2f}".format(num).replace(',', ' ')
    return formatted

def get_moscow_time():
    """Get current Moscow time"""
    moscow_tz = pytz.timezone('Europe/Moscow')
    return datetime.now(moscow_tz)

def get_today_date_msk():
    """Get today's date in MSK"""
    return get_moscow_time().date()

def is_media_message(message):
    """Check if message contains media"""
    return any([
        message.photo,
        message.video,
        message.animation,  # GIF
        message.voice,
        message.video_note,  # Video circle
        message.audio,
        message.document,
        message.sticker,
        bool(message.entities and any(e.type == 'url' for e in message.entities))
    ])

def count_words(text):
    """Count words in text"""
    if not text:
        return 0
    return len(text.split())

def generate_referral_link(bot_username, user_id):
    """Generate referral link"""
    import hashlib
    code = f"ref_{hashlib.md5(str(user_id).encode()).hexdigest()[:8]}"
    return f"https://t.me/{bot_username}?start={code}"

def format_duration(seconds):
    """Format duration in human-readable format"""
    if seconds < 60:
        return f"{seconds} сек"
    elif seconds < 3600:
        minutes = seconds // 60
        return f"{minutes} мин"
    else:
        hours = seconds // 3600
        return f"{hours} ч"

def calculate_lottery_chances(user_tickets, total_tickets):
    """Calculate lottery winning chances"""
    if total_tickets == 0:
        return 0.0
    return round((user_tickets / total_tickets) * 100, 2)

def escape_markdown(text):
    """Escape markdown special characters"""
    special_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    for char in special_chars:
        text = text.replace(char, f'\\{char}')
    return text

class Statistics:
    """Helper class for statistics calculations"""
    
    @staticmethod
    def calculate_percentage(part, total):
        """Calculate percentage"""
        if total == 0:
            return 0.0
        return round((part / total) * 100, 2)
    
    @staticmethod
    def get_rank_emoji(rank):
        """Get emoji for rank"""
        emojis = {
            1: '🥇',
            2: '🥈',
            3: '🥉',
            4: '4️⃣',
            5: '5️⃣'
        }
        return emojis.get(rank, '▪️')
    
    @staticmethod
    def format_activity_score(score, as_percentage=False, total_sum=None):
        """Format activity score"""
        if as_percentage and total_sum and total_sum > 0:
            percentage = (score / total_sum) * 100
            return f"{percentage:.2f}%"
        return f"{score:.2f}"

def create_html_report(data, title, headers, filename):
    """Create HTML report file"""
    html_content = f"""<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        h1 {{
            color: #2c3e50;
            text-align: center;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background-color: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background-color: #3498db;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f1f1f1;
        }}
        .footer {{
            text-align: center;
            margin-top: 20px;
            color: #7f8c8d;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <table>
        <thead>
            <tr>
"""
    
    for header in headers:
        html_content += f"                <th>{header}</th>\n"
    
    html_content += """            </tr>
        </thead>
        <tbody>
"""
    
    for row in data:
        html_content += "            <tr>\n"
        for cell in row:
            html_content += f"                <td>{cell}</td>\n"
        html_content += "            </tr>\n"
    
    html_content += """        </tbody>
    </table>
    <div class="footer">
        <p>Сгенерировано: """ + get_moscow_time().strftime('%d.%m.%Y %H:%M:%S') + """ МСК</p>
    </div>
</body>
</html>"""
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    return filename

def check_referral_qualification(db, user_id):
    """Check if referred user qualifies"""
    # Check if user has:
    # 1. Been in chat for 24 hours
    # 2. Sent 5+ messages OR received 3+ reactions
    
    user = db.get_user(user_id)
    if not user:
        return False
    
    # Check time in chat (24 hours)
    joined_at = datetime.fromisoformat(user['joined_at'])
    if datetime.now() - joined_at < timedelta(hours=24):
        return False
    
    # Check activity
    today = get_today_date_msk()
    db.cursor.execute('''
        SELECT SUM(total_messages) as msgs, SUM(reactions_received) as reactions
        FROM user_stats
        WHERE user_id = ?
    ''', (user_id,))
    
    result = db.cursor.fetchone()
    if not result:
        return False
    
    messages = result['msgs'] or 0
    reactions = result['reactions'] or 0
    
    return messages >= 5 or reactions >= 3


def export_stats_to_excel(stats_data, filename):
    """Export statistics to Excel file - beautiful single sheet"""
    import logging
    try:
        logging.info(f"Starting Excel export to: {filename}")
        logging.info(f"Stats data keys: {stats_data.keys() if stats_data else 'None'}")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Ensure directory exists
        import os
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"
        
        # ============ СТИЛИ ============
        
        # Главный заголовок
        title_font = Font(bold=True, size=18, color="FFFFFF")
        title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки секций
        section_font = Font(bold=True, size=14, color="FFFFFF")
        section_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        section_alignment = Alignment(horizontal='left', vertical='center')
        
        # Заголовки таблиц
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные (четные строки)
        data_even_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        
        # Данные (нечетные строки)
        data_odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Границы
        thin_border = Border(
            left=Side(style='thin', color="BDC3C7"),
            right=Side(style='thin', color="BDC3C7"),
            top=Side(style='thin', color="BDC3C7"),
            bottom=Side(style='thin', color="BDC3C7")
        )
        
        thick_border = Border(
            left=Side(style='medium', color="34495E"),
            right=Side(style='medium', color="34495E"),
            top=Side(style='medium', color="34495E"),
            bottom=Side(style='medium', color="34495E")
        )
        
        # Выравнивание
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # ============ ГЛАВНЫЙ ЗАГОЛОВОК ============
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = stats_data['title']
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = title_alignment
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 30
        
        # Период
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        period_text = f"Период: {stats_data['start_date']} - {stats_data['end_date']}"
        ws[f'A{row}'] = period_text
        ws[f'A{row}'].font = Font(size=11, italic=True)
        ws[f'A{row}'].alignment = center_align
        ws.row_dimensions[row].height = 20
        
        row += 2  # Пустая строка
        
        # ============ СЕКЦИЯ 1: ОБЩАЯ СТАТИСТИКА ============
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "📋 ОБЩАЯ СТАТИСТИКА"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = section_alignment
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 25
        
        row += 1
        
        # Заголовки для общей статистики
        ws[f'A{row}'] = "Параметр"
        ws[f'B{row}'] = "Значение"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 20
        
        row += 1
        start_general = row
        
        # Данные общей статистики (зебра)
        idx = 0
        for key, value in stats_data['general'].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            
            # Чередование цветов
            fill = data_even_fill if idx % 2 == 0 else data_odd_fill
            ws[f'A{row}'].fill = fill
            ws[f'B{row}'].fill = fill
            
            ws[f'A{row}'].alignment = left_align
            ws[f'B{row}'].alignment = right_align
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            
            row += 1
            idx += 1
        
        row += 1  # Пустая строка
        
        # ============ СЕКЦИЯ 2: ТОП-10 ПО СООБЩЕНИЯМ ============
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "🏆 ТОП-10 ПО СООБЩЕНИЯМ"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = section_alignment
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 25
        
        row += 1
        
        # Заголовки таблицы
        top_headers = ['№', 'Пользователь', 'Сообщений', 'Заработано 💎']
        top_cols = ['A', 'B', 'C', 'D']
        for col, header in zip(top_cols, top_headers):
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 20
        
        row += 1
        
        # Данные топа по сообщениям
        for idx, user_data in enumerate(stats_data.get('top_messages', [])[:10]):
            ws[f'A{row}'] = user_data.get('rank', idx + 1)
            ws[f'B{row}'] = user_data.get('username', 'Unknown')
            ws[f'C{row}'] = user_data.get('messages', 0)
            ws[f'D{row}'] = user_data.get('earned', 0)
            
            # Чередование цветов
            fill = data_even_fill if idx % 2 == 0 else data_odd_fill
            for col in top_cols:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = thin_border
            
            ws[f'A{row}'].alignment = center_align
            ws[f'B{row}'].alignment = left_align
            ws[f'C{row}'].alignment = center_align
            ws[f'D{row}'].alignment = center_align
            
            # Медали для топ-3
            if idx == 0:
                ws[f'A{row}'].font = Font(bold=True, color="FFD700", size=12)  # Золото
            elif idx == 1:
                ws[f'A{row}'].font = Font(bold=True, color="C0C0C0", size=12)  # Серебро
            elif idx == 2:
                ws[f'A{row}'].font = Font(bold=True, color="CD7F32", size=12)  # Бронза
            
            row += 1
        
        row += 1  # Пустая строка
        
        # ============ СЕКЦИЯ 3: ТОП-10 ПО ЗАРАБОТКУ ============
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "💎 ТОП-10 ПО ЗАРАБОТКУ"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = section_alignment
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 25
        
        row += 1
        
        # Заголовки таблицы
        earners_headers = ['№', 'Пользователь', 'Заработано 💎']
        earners_cols = ['A', 'B', 'C']
        for col, header in zip(earners_cols, earners_headers):
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 20
        
        row += 1
        
        # Данные топа по заработку
        for idx, user_data in enumerate(stats_data.get('top_earners', [])[:10]):
            ws[f'A{row}'] = user_data.get('rank', idx + 1)
            ws[f'B{row}'] = user_data.get('username', 'Unknown')
            ws[f'C{row}'] = user_data.get('earned', 0)
            
            # Чередование цветов
            fill = data_even_fill if idx % 2 == 0 else data_odd_fill
            for col in earners_cols:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = thin_border
            
            ws[f'A{row}'].alignment = center_align
            ws[f'B{row}'].alignment = left_align
            ws[f'C{row}'].alignment = center_align
            
            # Медали для топ-3
            if idx == 0:
                ws[f'A{row}'].font = Font(bold=True, color="FFD700", size=12)
            elif idx == 1:
                ws[f'A{row}'].font = Font(bold=True, color="C0C0C0", size=12)
            elif idx == 2:
                ws[f'A{row}'].font = Font(bold=True, color="CD7F32", size=12)
            
            row += 1
        
        row += 1  # Пустая строка
        
        # ============ СЕКЦИЯ 4: ДЕТАЛЬНАЯ СТАТИСТИКА (если есть) ============
        if stats_data.get('detailed_stats'):
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = "📊 ДЕТАЛЬНАЯ СТАТИСТИКА ВСЕХ ПОЛЬЗОВАТЕЛЕЙ"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].alignment = section_alignment
            ws[f'A{row}'].border = thick_border
            ws.row_dimensions[row].height = 25
            
            row += 1
            
            # Заголовки детальной таблицы
            detail_headers = ['№', 'Пользователь', 'Сообщений', 'Символов', 'Слов', 'Реакций', 'Ответов', 'Медиа', 'Дней в чате', 'Заработано 💎']
            detail_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            for col, header in zip(detail_cols, detail_headers):
                ws[f'{col}{row}'] = header
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].alignment = center_align
                ws[f'{col}{row}'].border = thin_border
            ws.row_dimensions[row].height = 20
            
            row += 1
            
            # Данные детальной статистики
            for idx, (user, data) in enumerate(stats_data['detailed_stats'].items()):
                ws[f'A{row}'] = idx + 1
                ws[f'B{row}'] = user
                ws[f'C{row}'] = data.get('messages', 0)
                ws[f'D{row}'] = data.get('chars', 0)
                ws[f'E{row}'] = data.get('words', 0)
                ws[f'F{row}'] = data.get('reactions', 0)
                ws[f'G{row}'] = data.get('replies', 0)
                ws[f'H{row}'] = data.get('media', 0)
                ws[f'I{row}'] = data.get('days_in_chat', 0)
                ws[f'J{row}'] = data.get('earned', 0)
                
                # Чередование цветов
                fill = data_even_fill if idx % 2 == 0 else data_odd_fill
                for col in detail_cols:
                    ws[f'{col}{row}'].fill = fill
                    ws[f'{col}{row}'].border = thin_border
                
                ws[f'A{row}'].alignment = center_align
                ws[f'B{row}'].alignment = left_align
                for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    ws[f'{col}{row}'].alignment = center_align
                
                row += 1
        
        # ============ АВТОШИРИНА КОЛОНОК ============
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        
        # ============ ЗАКРЕПИТЬ ПЕРВУЮ СТРОКУ ============
        ws.freeze_panes = 'A3'
        
        # ============ ДЕТАЛИЗАЦИЯ ПО ДНЯМ/МЕСЯЦАМ (НОВЫЙ ЛИСТ) ============
        if 'daily_stats' in stats_data and stats_data['daily_stats'] and 'period_type' in stats_data:
            period_type = stats_data['period_type']
            daily_data = stats_data['daily_stats']
            
            # Создаем новый лист
            ws_detail = wb.create_sheet("Детализация")
            
            # Заголовок
            row = 1
            num_cols = len(daily_data) + 2  # Итого + Параметр + дни
            last_col = get_column_letter(num_cols)
            ws_detail.merge_cells(f'A{row}:{last_col}{row}')
            
            if period_type == 'week':
                detail_title = f"📊 ДЕТАЛИЗАЦИЯ ПО ДНЯМ - НЕДЕЛЯ"
            elif period_type == 'month':
                detail_title = f"📊 ДЕТАЛИЗАЦИЯ ПО ДНЯМ - МЕСЯЦ"
            elif period_type == 'year':
                detail_title = f"📊 ДЕТАЛИЗАЦИЯ ПО МЕСЯЦАМ - ГОД"
            else:
                detail_title = "📊 ДЕТАЛЬНАЯ СТАТИСТИКА"
            
            ws_detail[f'A{row}'] = detail_title
            ws_detail[f'A{row}'].font = title_font
            ws_detail[f'A{row}'].fill = title_fill
            ws_detail[f'A{row}'].alignment = title_alignment
            ws_detail[f'A{row}'].border = thick_border
            ws_detail.row_dimensions[row].height = 30
            
            # Период
            row += 1
            ws_detail.merge_cells(f'A{row}:{last_col}{row}')
            period_text = f"Период: {stats_data['start_date']} - {stats_data['end_date']}"
            ws_detail[f'A{row}'] = period_text
            ws_detail[f'A{row}'].font = Font(size=11, italic=True)
            ws_detail[f'A{row}'].alignment = center_align
            ws_detail.row_dimensions[row].height = 20
            
            row += 2
            
            # Заголовки столбцов
            ws_detail['A{}'.format(row)] = "Параметр"
            ws_detail['B{}'.format(row)] = "ИТОГО"
            
            # Заголовки для каждого дня/месяца
            for idx, day_stat in enumerate(daily_data):
                col = get_column_letter(idx + 3)  # C, D, E, F, ...
                ws_detail[f'{col}{row}'] = day_stat['date']
                ws_detail[f'{col}{row}'].font = header_font
                ws_detail[f'{col}{row}'].fill = header_fill
                ws_detail[f'{col}{row}'].alignment = center_align
                ws_detail[f'{col}{row}'].border = thin_border
            
            ws_detail[f'A{row}'].font = header_font
            ws_detail[f'A{row}'].fill = header_fill
            ws_detail[f'A{row}'].alignment = center_align
            ws_detail[f'A{row}'].border = thin_border
            
            ws_detail[f'B{row}'].font = header_font
            ws_detail[f'B{row}'].fill = header_fill
            ws_detail[f'B{row}'].alignment = center_align
            ws_detail[f'B{row}'].border = thin_border
            
            ws_detail.row_dimensions[row].height = 20
            row += 1
            
            # Строки с данными - ВСЕ ПАРАМЕТРЫ
            params = [
                ('КСП - Количество слов', 'words'),
                ('Количество реакций', 'reactions'),
                ('ОКСП - Общее кол-во символов', 'chars'),
                ('СДСП - Средняя длина сообщения', 'avg_msg_length'),
                ('💬 Сообщений с администраторами', 'msgs_with_admins'),
                ('💬 Сообщений без администраторов', 'msgs_without_admins'),
                ('💬 Всего сообщений', 'messages'),
                ('МедиаП - Медиа контент', 'media'),
                ('👥 Активных пользователей', 'active_users'),
                ('💎 Добыто Пульсов', 'pulses'),
                ('👥 Пользователей в чате', 'total_users'),
                ('🆕 Вступило за период', 'joined'),
                ('👋 Вышло за период', 'left_users'),
                ('📊 Коэффициент вовлеченности', 'engagement'),
                ('КОРП - Реакции оставленные', 'reactions_given'),
                ('КПРП - Реакции полученные', 'reactions_received'),
                ('КОПЮП - Ответы пользователю', 'replies_received'),
                ('КОПЯП - Ответы пользователя', 'replies_sent'),
                ('КУПП - Упоминания @', 'mentions'),
                ('ПИВДВП - Публ. в других ветках', 'other_threads')
            ]
            
            for param_idx, (param_name, param_key) in enumerate(params):
                # Название параметра
                ws_detail[f'A{row}'] = param_name
                
                # Подсчет итого (для engagement - среднее, для остальных - сумма)
                if param_key == 'engagement':
                    # Среднее значение вовлеченности
                    total = sum(day_stat.get(param_key, 0) for day_stat in daily_data) / len(daily_data) if daily_data else 0
                    ws_detail[f'B{row}'] = f"{total:.1f}%"
                elif param_key == 'avg_msg_length':
                    # Среднее значение длины сообщения
                    total = sum(day_stat.get(param_key, 0) for day_stat in daily_data) / len(daily_data) if daily_data else 0
                    ws_detail[f'B{row}'] = f"{total:.1f}"
                else:
                    # Сумма для остальных параметров
                    total = sum(day_stat.get(param_key, 0) for day_stat in daily_data)
                    ws_detail[f'B{row}'] = format_number(total)
                
                # Данные по дням
                for idx, day_stat in enumerate(daily_data):
                    col = get_column_letter(idx + 3)
                    value = day_stat.get(param_key, 0)
                    
                    # Форматирование в зависимости от типа
                    if param_key == 'engagement':
                        ws_detail[f'{col}{row}'] = f"{value:.1f}%"
                    elif param_key == 'avg_msg_length':
                        ws_detail[f'{col}{row}'] = f"{value:.1f}"
                    else:
                        ws_detail[f'{col}{row}'] = format_number(value)
                    
                    # Стиль
                    fill = data_even_fill if param_idx % 2 == 0 else data_odd_fill
                    ws_detail[f'{col}{row}'].fill = fill
                    ws_detail[f'{col}{row}'].alignment = center_align
                    ws_detail[f'{col}{row}'].border = thin_border
                
                # Стили для первых двух колонок
                fill = data_even_fill if param_idx % 2 == 0 else data_odd_fill
                ws_detail[f'A{row}'].fill = fill
                ws_detail[f'A{row}'].alignment = left_align
                ws_detail[f'A{row}'].border = thin_border
                
                ws_detail[f'B{row}'].fill = fill
                ws_detail[f'B{row}'].alignment = center_align
                ws_detail[f'B{row}'].border = thin_border
                
                row += 1
            
            # Ширина колонок
            ws_detail.column_dimensions['A'].width = 30
            ws_detail.column_dimensions['B'].width = 15
            
            for idx in range(len(daily_data)):
                col = get_column_letter(idx + 3)
                ws_detail.column_dimensions[col].width = 13
            
            # Закрепить заголовки
            ws_detail.freeze_panes = 'C4'
        
        # Сохранение
        logging.info(f"Saving Excel file to: {filename}")
        wb.save(filename)
        logging.info(f"✅ Excel file created successfully: {filename}")
        return filename
        
    except Exception as e:
        import logging
        logging.error(f"❌ Error creating Excel: {e}")
        import traceback
        logging.error(traceback.format_exc())
        traceback.print_exc()
        return None


def export_stats_to_csv(stats_data, filename):
    """Export statistics to CSV file"""
    try:
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            
            # Title
            writer.writerow([stats_data['title']])
            writer.writerow([])
            
            # General stats
            writer.writerow(['📊 ОБЩАЯ СТАТИСТИКА'])
            for key, value in stats_data['general'].items():
                writer.writerow([key, value])
            
            writer.writerow([])
            
            # Top messages
            writer.writerow(['🏆 ТОП-5 ПО СООБЩЕНИЯМ'])
            writer.writerow(['Место', 'Пользователь', 'Сообщений', 'Заработано'])
            for user_data in stats_data['top_messages']:
                writer.writerow([
                    user_data['rank'],
                    user_data['username'],
                    user_data['messages'],
                    user_data.get('earned', 0)
                ])
            
            writer.writerow([])
            
            # Top earners
            writer.writerow(['💰 ТОП-5 ПО ЗАРАБОТКУ'])
            writer.writerow(['Место', 'Пользователь', 'Заработано'])
            for user_data in stats_data['top_earners']:
                writer.writerow([
                    user_data['rank'],
                    user_data['username'],
                    user_data['earned']
                ])
            
            writer.writerow([])
            writer.writerow([f'Сгенерировано: {get_moscow_time().strftime("%d.%m.%Y %H:%M:%S")} МСК'])
        
        return filename
    except Exception as e:
        print(f"Error creating CSV: {e}")
        return None


def export_stats_to_pdf(stats_data, filename):
    """Export statistics to PDF file with Cyrillic support"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Register DejaVu fonts for Cyrillic support
        try:
            # Try to register DejaVu fonts (common on Linux)
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
            font_name = 'DejaVu'
            font_bold = 'DejaVu-Bold'
        except:
            # Fallback to Helvetica (no Cyrillic, but won't crash)
            font_name = 'Helvetica'
            font_bold = 'Helvetica-Bold'
        
        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName=font_bold
        )
        
        # Normal style
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=font_name
        )
        
        elements.append(Paragraph(stats_data['title'], title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # General stats
        general_data = [[Paragraph('ОБЩАЯ СТАТИСТИКА', title_style), '']]
        for key, value in stats_data['general'].items():
            general_data.append([
                Paragraph(str(key), normal_style),
                Paragraph(str(value), normal_style)
            ])
        
        general_table = Table(general_data, colWidths=[10*cm, 6*cm])
        general_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(general_table)
        elements.append(Spacer(1, 1*cm))
        
        # Top messages header
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Heading2'],
            fontSize=14,
            fontName=font_bold,
            textColor=colors.whitesmoke
        )
        
        messages_data = [[Paragraph('ТОП-5 ПО СООБЩЕНИЯМ', header_style), '', '', '']]
        messages_data.append([
            Paragraph('Место', normal_style),
            Paragraph('Пользователь', normal_style),
            Paragraph('Сообщений', normal_style),
            Paragraph('Заработано', normal_style)
        ])
        
        for user_data in stats_data['top_messages']:
            messages_data.append([
                Paragraph(str(user_data['rank']), normal_style),
                Paragraph(str(user_data['username']), normal_style),
                Paragraph(str(user_data['messages']), normal_style),
                Paragraph(str(user_data.get('earned', 0)), normal_style)
            ])
        
        messages_table = Table(messages_data, colWidths=[2*cm, 7*cm, 4*cm, 4*cm])
        messages_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTNAME', (0, 0), (-1, 1), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e8f4f8')),
            ('BOTTOMPADDING', (0, 0), (-1, 1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(messages_table)
        elements.append(Spacer(1, 1*cm))
        
        # Top earners
        earners_data = [[Paragraph('ТОП-5 ПО ЗАРАБОТКУ', header_style), '', '']]
        earners_data.append([
            Paragraph('Место', normal_style),
            Paragraph('Пользователь', normal_style),
            Paragraph('Заработано', normal_style)
        ])
        
        for user_data in stats_data['top_earners']:
            earners_data.append([
                Paragraph(str(user_data['rank']), normal_style),
                Paragraph(str(user_data['username']), normal_style),
                Paragraph(str(user_data['earned']), normal_style)
            ])
        
        earners_table = Table(earners_data, colWidths=[3*cm, 8*cm, 6*cm])
        earners_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTNAME', (0, 0), (-1, 1), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e8f4f8')),
            ('BOTTOMPADDING', (0, 0), (-1, 1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(earners_table)
        
        # Footer
        elements.append(Spacer(1, 2*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7f8c8d'),
            alignment=TA_CENTER,
            fontName=font_name
        )
        footer_text = f"Сгенерировано: {get_moscow_time().strftime('%d.%m.%Y %H:%M:%S')} МСК"
        elements.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(elements)
        return filename
    except Exception as e:
        print(f"Error creating PDF: {e}")
        import traceback
        traceback.print_exc()
        return None

def calculate_days_in_chat(joined_at):
    """Calculate days user has been in chat"""
    if not joined_at:
        return 0
    
    from datetime import datetime
    
    if isinstance(joined_at, str):
        joined_at = datetime.fromisoformat(joined_at.replace('Z', '+00:00'))
    
    now = datetime.now(joined_at.tzinfo) if joined_at.tzinfo else datetime.now()
    delta = now - joined_at
    
    return delta.days


def export_users_stats_to_excel(users_data, filename, period_name):
    """
    Export ALL users statistics to Excel with ALL new parameters
    
    20 columns total:
    №, Имя, @username, ID, Сообщ., ОКСП, КСП, СДСП, 
    КОРП, КПРП, КОПЮП, КОПЯП, КУПП, МедиаП, ПИВДВП,
    Дней в чате, Когда пришёл, Последняя активность, Онлайн, Заработано 💎
    """
    import logging
    
    def calculate_online_time(last_active_str, total_messages):
        """
        Рассчитывает примерное время онлайн на основе последней активности
        
        Возвращает строку вида:
        - "5 мин назад" 
        - "2 ч 30 мин назад"
        - "3 дня назад"
        - "~45 мин" (если был сегодня, показываем примерное время в чате)
        """
        from datetime import datetime
        
        if not last_active_str:
            return 'Не известно'
        
        try:
            # Парсим last_active
            last_active = datetime.fromisoformat(last_active_str.replace('Z', '+00:00'))
            now = datetime.now(last_active.tzinfo) if last_active.tzinfo else datetime.now()
            
            # Разница во времени
            delta = now - last_active
            total_seconds = int(delta.total_seconds())
            
            # Если был онлайн менее минуты назад
            if total_seconds < 60:
                return "Только что"
            
            # Если был онлайн менее часа назад
            elif total_seconds < 3600:
                minutes = total_seconds // 60
                return f"{minutes} мин назад"
            
            # Если был онлайн менее суток назад
            elif total_seconds < 86400:
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                if minutes > 0:
                    return f"{hours} ч {minutes} мин назад"
                else:
                    return f"{hours} ч назад"
            
            # Если был онлайн несколько дней назад
            else:
                days = total_seconds // 86400
                if days == 1:
                    return "Вчера"
                elif days < 30:
                    return f"{days} дн назад"
                else:
                    # Для старых данных показываем примерное время в чате
                    # на основе сообщений (1 сообщение ≈ 2 минуты)
                    estimated_minutes = total_messages * 2
                    if estimated_minutes < 60:
                        return f"~{estimated_minutes} мин"
                    else:
                        hours = estimated_minutes // 60
                        return f"~{hours} ч"
        
        except Exception as e:
            # Если не удалось распарсить - показываем примерное время
            # на основе количества сообщений
            if total_messages > 0:
                estimated_minutes = total_messages * 2
                if estimated_minutes < 60:
                    return f"~{estimated_minutes} мин"
                else:
                    hours = estimated_minutes // 60
                    minutes = estimated_minutes % 60
                    if minutes > 0:
                        return f"~{hours} ч {minutes} мин"
                    else:
                        return f"~{hours} ч"
            return 'Не известно'
    
    try:
        logging.info(f"📊 Starting users Excel export to: {filename}")
        logging.info(f"👥 Total users to export: {len(users_data)}")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Ensure directory exists
        import os
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Все пользователи"
        
        # === СТИЛИ ===
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        
        data_even_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        data_odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color="BDC3C7"),
            right=Side(style='thin', color="BDC3C7"),
            top=Side(style='thin', color="BDC3C7"),
            bottom=Side(style='thin', color="BDC3C7")
        )
        
        # === ЗАГОЛОВОК ===
        ws.merge_cells('A1:T1')
        ws['A1'] = f'📊 СТАТИСТИКА ВСЕХ ПОЛЬЗОВАТЕЛЕЙ ЧАТА - {period_name.upper()}'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        ws.row_dimensions[1].height = 30
        
        # === ЗАГОЛОВКИ КОЛОНОК (20 штук) ===
        headers = [
            '№',                                # A - Номер
            'Имя',                              # B - Имя пользователя
            '@username',                        # C - Username
            'ID',                               # D - User ID
            'Сообщений',                        # E - Количество сообщений
            'Общее кол-во\nсимволов',          # F - Общее кол-во символов
            'Количество\nслов',                 # G - Количество слов
            'Средняя\nдлина',                   # H - Средняя длина сообщения
            'Реакции\nоставленные',            # I - Реакции оставленные
            'Реакции\nполученные',             # J - Реакции полученные
            'Ответы\nпользователю',            # K - Ответы пользователю
            'Ответы\nпользователя',            # L - Ответы пользователя
            'Упоминания\n@',                    # M - Упоминания @
            'Медиа',                            # N - Медиа контент
            'Др. ветки',                        # O - Публ. в др. ветках
            'Дней\nв чате',                     # P - Дней в чате
            'Когда\nпришёл',                    # Q - Дата присоединения
            'Последняя\nактивность',            # R - Последнее действие
            'Онлайн',                           # S - Статус онлайн
            'Заработано\n💎'                    # T - Заработанные пульсы
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        ws.row_dimensions[2].height = 35
        
        # === ДАННЫЕ ВСЕХ ПОЛЬЗОВАТЕЛЕЙ ===
        logging.info(f"📝 Writing {len(users_data)} users to Excel...")
        
        for idx, user in enumerate(users_data, 1):
            row = idx + 2
            fill = data_even_fill if idx % 2 == 0 else data_odd_fill
            
            # Рассчитываем СДСП (средняя длина сообщения)
            total_msgs = user.get('total_messages', 0)
            total_chars = user.get('total_chars', 0)
            avg_chars = total_chars / total_msgs if total_msgs > 0 else 0
            
            # Форматируем дату присоединения
            joined_at = user.get('joined_at', '')
            joined_date = joined_at[:10] if joined_at else '-'
            
            # Последняя активность
            last_active = user.get('last_active', '')
            last_active_date = last_active[:10] if last_active else '-'
            
            # Рассчитываем время онлайн
            online_time = calculate_online_time(last_active, total_msgs)
            
            # Все значения для строки
            values = [
                idx,                                                      # A - №
                user.get('first_name', 'Unknown'),                        # B - Имя
                f"@{user['username']}" if user.get('username') else '-', # C - @username
                user.get('user_id', 0),                                   # D - ID
                total_msgs,                                               # E - Сообщений
                total_chars,                                              # F - ОКСП (символы)
                user.get('total_words', 0),                              # G - КСП (слова)
                f"{avg_chars:.1f}",                                      # H - СДСП (средняя длина)
                user.get('reactions_given', 0),                          # I - КОРП (реакции отданные)
                user.get('reactions_received', 0),                       # J - КПРП (реакции полученные)
                user.get('replies_received', 0),                         # K - КОПЮП (ответы ЕМУ)
                user.get('replies_sent', 0),                             # L - КОПЯП (ответы ОТ НЕГО)
                user.get('mentions_received', 0),                        # M - КУПП (упоминания)
                user.get('media_sent', 0),                               # N - МедиаП (медиа)
                user.get('other_threads_posts', 0),                      # O - ПИВДВП (другие ветки)
                user.get('days_in_chat', 0),                             # P - Дней в чате
                joined_date,                                              # Q - Когда пришёл
                last_active_date,                                         # R - Последняя активность
                online_time,                                              # S - Онлайн (время/активность)
                f"{user.get('pulses_mined', 0):.2f}"                    # T - Заработано 💎
            ]
            
            # Записываем значения в ячейки
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = value
                cell.fill = fill
                cell.border = thin_border
                
                # Выравнивание: имя - слева, остальное - по центру
                if col_num == 2:  # Имя
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
        
        # === ШИРИНА КОЛОНОК (оптимизировано для читаемости) ===
        widths = {
            'A': 5,   # №
            'B': 20,  # Имя
            'C': 15,  # @username
            'D': 12,  # ID
            'E': 12,  # Сообщений
            'F': 14,  # Общее кол-во символов
            'G': 12,  # Количество слов
            'H': 10,  # Средняя длина
            'I': 12,  # Реакции оставленные
            'J': 12,  # Реакции полученные
            'K': 14,  # Ответы пользователю
            'L': 12,  # Ответы пользователя
            'M': 16,  # Упоминания @
            'N': 10,  # Медиа
            'O': 10,  # Др. ветки
            'P': 10,  # Дней в чате
            'Q': 12,  # Когда пришёл
            'R': 12,  # Последняя активность
            'S': 15,  # Онлайн (увеличено для "2 ч 30 мин назад")
            'T': 12   # Заработано 💎
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # === ЗАКРЕПИТЬ ЗАГОЛОВКИ ===
        ws.freeze_panes = 'A3'
        
        # === СОХРАНЕНИЕ ===
        logging.info(f"💾 Saving users Excel file to: {filename}")
        wb.save(filename)
        logging.info(f"✅ Users Excel created successfully: {len(users_data)} users, 20 columns")
        return filename
        
    except Exception as e:
        import logging
        logging.error(f"❌ Error creating users Excel: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return None

# Пресс релиз автоформатирование 
def parse_flexible_datetime(text: str) -> str:
    """
    Преобразует текст с любыми разделителями в формат 'DD.MM.YYYY HH:MM'.
    Возвращает строку формата 'DD.MM.YYYY HH:MM' или None, если формат не распознан.
    """
    from datetime import datetime
    
    # Находим все группы цифр в тексте
    numbers = re.findall(r'\d+', text)
    
    # Если ввели 4 числа: День, Месяц, Часы, Минуты (без года)
    if len(numbers) == 4:
        d, m, h, min_ = numbers
        y = str(datetime.now().year)
        
    # Если ввели 5 чисел: День, Месяц, Год, Часы, Минуты
    elif len(numbers) == 5:
        d, m, y, h, min_ = numbers
        # Если год указан двумя цифрами (26 вместо 2026)
        if len(y) == 2:
            y = "20" + y
    else:
        return None # Неверное количество параметров

    # Добавляем нули слева, если ввели одну цифру (чтобы 5 превратилось в 05)
    d = d.zfill(2)
    m = m.zfill(2)
    h = h.zfill(2)
    min_ = min_.zfill(2)
    
    try:
        # Проверяем, существует ли такая дата (чтобы не ввели 32.13.2026)
        dt = datetime(int(y), int(m), int(d), int(h), int(min_))
        return dt.strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return None # Дата некорректна