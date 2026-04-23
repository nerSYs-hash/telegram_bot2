#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Экспорт статистики в Excel/PDF."""

import os
import logging
import calendar
import traceback
from datetime import datetime, timedelta, date as date_type
from decimal import Decimal

from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from utils.helpers import (
    format_number, get_moscow_time, get_today_date_msk,
    calculate_days_in_chat, export_stats_to_excel,
    export_users_stats_to_excel, to_decimal, round_decimal,
)
from utils.helpers.export_pdf import export_stats_to_pdf

def _add_rate_sheet_to_excel(filepath, db, start_date, end_date, period_name):
    """
    Добавить лист "Курс" с таблицей + графиком в существующий Excel-файл.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils import get_column_letter

        # Получаем историю курса за период
        db.cursor.execute('''
            SELECT rate, ai_value, total_members, avg_active,
                   denominator, is_manual, timestamp
            FROM exchange_rate_history
            WHERE timestamp >= ? AND timestamp <= ?
            ORDER BY timestamp ASC
        ''', (start_date.strftime('%Y-%m-%d %H:%M:%S'),
              end_date.strftime('%Y-%m-%d %H:%M:%S')))
        rows = db.cursor.fetchall()

        if not rows:
            logging.info("No rate history for this period — skipping Курс sheet")
            return

        wb = load_workbook(filepath)
        ws = wb.create_sheet("Курс")

        # ═══ СТИЛИ ═══
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
        even_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin', color="BDC3C7"),
            right=Side(style='thin', color="BDC3C7"),
            top=Side(style='thin', color="BDC3C7"),
            bottom=Side(style='thin', color="BDC3C7")
        )
        green_font = Font(color="27AE60", bold=True)
        red_font = Font(color="E74C3C", bold=True)

        # ═══ ЗАГОЛОВОК ═══
        row = 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"💱 ДИНАМИКА КУРСА ПУЛЬСА — {period_name}"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = center
        ws.row_dimensions[row].height = 30
        row += 2

        # ═══ СВОДКА ═══
        rates = [r['rate'] for r in rows]
        min_rate = min(rates)
        max_rate = max(rates)
        avg_rate = sum(rates) / len(rates)
        first_rate = rates[0]
        last_rate = rates[-1]
        change = last_rate - first_rate
        change_pct = (change / first_rate * 100) if first_rate > 0 else 0

        summary = [
            ("Записей в периоде", str(len(rows))),
            ("Минимальный курс", f"{min_rate:.6f} ₽"),
            ("Максимальный курс", f"{max_rate:.6f} ₽"),
            ("Средний курс", f"{avg_rate:.6f} ₽"),
            ("Курс на начало", f"{first_rate:.6f} ₽"),
            ("Курс на конец", f"{last_rate:.6f} ₽"),
            ("Изменение", f"{'+' if change >= 0 else ''}{change:.6f} ₽ ({'+' if change_pct >= 0 else ''}{change_pct:.2f}%)"),
        ]

        for i, (label, value) in enumerate(summary):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].alignment = left
            ws[f'B{row}'].alignment = left
            fill = even_fill if i % 2 == 0 else odd_fill
            ws[f'A{row}'].fill = fill
            ws[f'B{row}'].fill = fill
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            # Цвет для изменения
            if label == "Изменение":
                ws[f'B{row}'].font = green_font if change >= 0 else red_font
            row += 1

        row += 1

        # ═══ ГРАФИК ═══
        chart_start_row = row
        # Сначала данные для графика (дата + курс)
        ws[f'A{row}'] = "Дата/Время"
        ws[f'B{row}'] = "Курс (₽)"
        ws[f'C{row}'] = "AI"
        ws[f'D{row}'] = "Участников"
        ws[f'E{row}'] = "Активных (ср.)"
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center
            ws[f'{col}{row}'].border = border
        row += 1

        data_start_row = row
        for i, r in enumerate(rows):
            ts = str(r['timestamp'])
            try:
                dt = datetime.strptime(ts[:19], '%Y-%m-%d %H:%M:%S')
                date_str = dt.strftime('%d.%m %H:%M')
            except (ValueError, TypeError):
                date_str = ts[:16]

            ws[f'A{row}'] = date_str
            ws[f'B{row}'] = round(r['rate'], 6)
            ws[f'C{row}'] = round(r['ai_value'], 2) if r['ai_value'] else 0
            ws[f'D{row}'] = r['total_members'] or 0
            ws[f'E{row}'] = round(r['avg_active'], 1) if r['avg_active'] else 0

            fill = even_fill if i % 2 == 0 else odd_fill
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].alignment = center

            row += 1

        data_end_row = row - 1

        # ═══ СОЗДАЁМ ГРАФИК ═══
        if data_end_row > data_start_row:
            chart = LineChart()
            chart.title = f"Динамика курса Пульса — {period_name}"
            chart.y_axis.title = "Курс (₽)"
            chart.x_axis.title = "Время"
            chart.style = 10
            chart.width = 28
            chart.height = 14

            # Данные курса
            data_ref = Reference(ws, min_col=2, min_row=chart_start_row,
                                 max_row=data_end_row)
            cats_ref = Reference(ws, min_col=1, min_row=data_start_row,
                                 max_row=data_end_row)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            # Стиль линии
            s = chart.series[0]
            s.graphicalProperties.line.width = 25000
            s.graphicalProperties.line.solidFill = "3498DB"
            s.smooth = True

            # Размещаем график справа от данных
            ws.add_chart(chart, f"G{chart_start_row}")

        # ═══ ШИРИНА КОЛОНОК ═══
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16

        ws.freeze_panes = f'A{data_start_row}'

        wb.save(filepath)
        logging.info(f"📊 Rate sheet added to {filepath}: {len(rows)} records")

    except Exception as e:
        logging.error(f"Error adding rate sheet: {e}")
        traceback.print_exc()


# ── Decimal-хелперы для вычислений ───────────────────────────────────────────

def _d(val) -> Decimal:
    """Алиас to_decimal для краткости внутри модуля."""
    return to_decimal(val)


def _calc_index(val, coeff, norm=Decimal('1')) -> Decimal:
    """
    Расчёт компонента индекса через Decimal.
    coeff * (val / norm) — без ошибок плавающей точки.
    """
    return Decimal(str(coeff)) * (_d(val) / _d(norm))


async def generate_export_file(query, data, user, context, db, admin_id, target_chat_id):
    """Генерация полного аналитического отчета (Все данные и индексы на месте)."""
    if user.id != admin_id:
        await query.answer("У вас нет доступа.", show_alert=True)
        return

    parts = data.split('_')
    file_format = parts[1]
    period = parts[2]

    await query.edit_message_text("⏳ Генерирую расширенный аналитический отчёт...")

    try:
        now = get_moscow_time()
        if period == 'yesterday':
            yesterday = (now - timedelta(days=1))
            start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date   = yesterday.replace(hour=23, minute=59, second=59, microsecond=0)
            period_name = "За вчера"
        elif period == 'day':
            start_date  = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date    = now
            period_name = "За сегодня"
        elif period == 'week':
            start_date  = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=6)
            end_date    = now
            period_name = "За неделю"
        elif period == 'month':
            start_date  = now - timedelta(days=30)
            end_date    = now
            period_name = "За месяц"
        elif period == 'year':
            start_date  = now - timedelta(days=365)
            end_date    = now
            period_name = "За год"
        else:
            await query.edit_message_text("❌ Неизвестный период")
            return

        # 1. Знаменатель для индексов активности
        try:
            m_count = await context.bot.get_chat_member_count(target_chat_id)
        except Exception:
            db.cursor.execute('SELECT COUNT(*) as total FROM users')
            m_count = db.cursor.fetchone()['total']

        b_count = int(os.getenv('BOT_COUNT', 1))
        divisor = Decimal(max(m_count - b_count - 1, 1))   # ← Decimal

        # 1. Агрегированные данные из user_stats (заменяем chat_stats)
        db.cursor.execute('''
            SELECT
                COALESCE(SUM(total_chars), 0)                as chars,
                COALESCE(SUM(total_messages), 0)             as msgs,
                COALESCE(SUM(total_words), 0)                as words,
                COALESCE(SUM(media_sent), 0)                 as media,
                COALESCE(SUM(pulses_mined), 0)               as pulses,
                CAST(SUM(total_chars) AS REAL) / NULLIF(SUM(total_messages), 0) as avg_len
            FROM user_stats WHERE date >= ? AND date <= ?
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        raw = db.cursor.fetchone()

        # 2. Детальные данные из user_stats (для индексов)
        db.cursor.execute('''
            SELECT
                COALESCE(SUM(reactions_given), 0)    as korp,
                COALESCE(SUM(reactions_received), 0) as kprp,
                COALESCE(SUM(replies_received), 0)   as kopyup,
                COALESCE(SUM(replies_sent), 0)       as kopyap,
                COALESCE(SUM(mentions_received), 0)  as kupp,
                COALESCE(SUM(other_threads_posts), 0) as pivdvp
            FROM user_stats WHERE date >= ? AND date <= ?
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        det = db.cursor.fetchone()

        # 3. Структура отчёта
        stats_data = {
            'title': f'Аналитический отчет чата - {period_name}',
            'period': period_name, 'period_type': period,
            'start_date': start_date.strftime('%d.%m.%Y'),
            'end_date':   end_date.strftime('%d.%m.%Y'),
            'general': {}, 'top_messages': [], 'top_earners': [],
            'detailed_stats': {}, 'daily_stats': []
        }

        # 4. Наполнение GENERAL через Decimal
        g = stats_data['general']

        oksp_idx  = _calc_index(raw['chars'],    Decimal('0.05'), Decimal('100')) / divisor
        sdsp_idx  = _calc_index(raw['avg_len'],  Decimal('0.05'), Decimal('100')) / divisor
        cho_idx   = _calc_index(raw['words'],     Decimal('0.05'))                 / divisor
        media_idx = _calc_index(raw['media'],    Decimal('0.07'))                 / divisor
        korp_idx  = _calc_index(det['korp'],     Decimal('0.08'))                 / divisor
        kprp_idx  = _calc_index(det['kprp'],     Decimal('0.10'))                 / divisor
        kopyup_idx= _calc_index(det['kopyup'],   Decimal('0.18'))                 / divisor
        kopyap_idx= _calc_index(det['kopyap'],   Decimal('0.15'))                 / divisor
        kupp_idx  = _calc_index(det['kupp'],     Decimal('0.15'))                 / divisor
        pivdvp_idx= _calc_index(det['pivdvp'],   Decimal('0.12'))                 / divisor

        def _fmt_idx(idx_d: Decimal, raw_val) -> str:
            """Формат для general: 'индекс (значение)' — строка только для отображения."""
            return f"{float(round_decimal(idx_d, 3)):.3f} ({format_number(raw_val)})"

        g['ОКС(Ч) - Общее количество символов']        = _fmt_idx(oksp_idx,   raw['chars'])
        g['СДС(Ч) - Средняя длина сообщения']          = _fmt_idx(sdsp_idx,   raw['avg_len'])
        g['КС(Ч) - Количество слов (Всего сообщ.)']    = _fmt_idx(cho_idx,    raw['words'])
        g['Медиа(Ч) - Медиа контент']                  = _fmt_idx(media_idx,  raw['media'])
        g['КОР(Ч) - Реакции оставленные']              = _fmt_idx(korp_idx,   det['korp'])
        g['КПР(Ч) - Реакции полученные']               = _fmt_idx(kprp_idx,   det['kprp'])
        g['КОтв(Ч) - Ответы полученные']               = _fmt_idx(kopyup_idx, det['kopyup'])
        g['КОтп(Ч) - Ответы отправленные']             = _fmt_idx(kopyap_idx, det['kopyap'])
        g['КУП(Ч) - Упоминания @']                     = _fmt_idx(kupp_idx,   det['kupp'])
        g['ПДВ(Ч) - Публ. в других ветках']            = _fmt_idx(pivdvp_idx, det['pivdvp'])

        g['💬 Сообщений с администраторами']   = format_number(raw['with_adm'])
        g['💬 Сообщений без администраторов']  = format_number(raw['no_adm'])

        db.cursor.execute(
            'SELECT COUNT(DISTINCT user_id) as act FROM messages WHERE timestamp >= ? AND timestamp <= ?',
            (start_date, end_date)
        )
        stats_data['act_count'] = db.cursor.fetchone()['act']
        g['👥 Активных пользователей'] = stats_data['act_count']
        g['💎 Добыто Пульсов']         = format_number(raw['pulses'])

        for t_type, t_label in [
            ('donate_to_user',   '🎁 Донатов пользователям'),
            ('donate_to_bank',   '🏦 Донатов в банк'),
            ('reactor_donation', '🔋 Донатов в реактор'),
        ]:
            db.cursor.execute(
                'SELECT COALESCE(SUM(amount),0) as tot, COUNT(*) as cnt FROM transactions '
                'WHERE transaction_type=? AND timestamp>=? AND timestamp<=?',
                (t_type, start_date, end_date)
            )
            r_don = db.cursor.fetchone()
            g[t_label] = f"{format_number(r_don['tot'])} 💎 ({r_don['cnt']} шт.)"

        db.cursor.execute(
            'SELECT COUNT(*) as j FROM users WHERE joined_at >= ? AND joined_at <= ? AND is_admin=0 AND is_owner=0',
            (start_date, end_date)
        )
        g['🆕 Вступило за период'] = db.cursor.fetchone()['j']

        db.cursor.execute(
            "SELECT COUNT(*) as l FROM transactions WHERE transaction_type='return_on_leave' "
            "AND timestamp >= ? AND timestamp <= ?",
            (start_date, end_date)
        )
        g['👋 Вышло за период'] = db.cursor.fetchone()['l']

        er = round_decimal(
            _d(stats_data['act_count']) / _d(m_count) * Decimal('100') if m_count > 0 else Decimal('0'),
            1
        )
        g['📊 Коэффициент вовлеченности (ER)'] = f"{float(er):.1f}%"

        # Итоговый индекс здоровья чата — Decimal сумма
        health_idx = (
            oksp_idx + sdsp_idx + cho_idx + media_idx +
            korp_idx + kprp_idx + kopyup_idx + kopyap_idx +
            kupp_idx + pivdvp_idx
        )
        g['🛡️ ИТОГОВЫЙ ИНДЕКС ЗДОРОВЬЯ ЧАТА'] = f"{float(round_decimal(health_idx, 3)):.3f}"

        # 5. ТОП 10 АКТИВИСТОВ (по индексу активности)
        from utils.exchange_rate import ACTIVITY_INDEX_SQL
        db.cursor.execute(f'''
            SELECT u.username, u.first_name,
                   ({ACTIVITY_INDEX_SQL}) as activity_index,
                   COALESCE(SUM(us.pulses_mined), 0) as earned
            FROM users u
            LEFT JOIN user_stats us ON u.user_id = us.user_id AND us.date >= ? AND us.date <= ?
            WHERE (u.is_admin=0 AND u.is_owner=0)
            GROUP BY u.user_id
            HAVING SUM(us.total_messages) > 0 OR SUM(us.reactions_given) > 0
                   OR SUM(us.reactions_received) > 0 OR SUM(us.replies_sent) > 0
            ORDER BY activity_index DESC LIMIT 10
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        for i, r_msg in enumerate(db.cursor.fetchall()):
            stats_data['top_messages'].append({
                'rank': i + 1,
                'username': f"@{r_msg['username'] or r_msg['first_name']}",
                'activity_score': float(_d(r_msg['activity_index'])),
                'earned_raw': float(_d(r_msg['earned'])),
            })

        # 5b. ТОП 10 ПО ЗАРАБОТКУ
        db.cursor.execute('''
            SELECT u.username, u.first_name,
                   COALESCE(SUM(us.pulses_mined), 0) as earned
            FROM users u
            LEFT JOIN user_stats us ON u.user_id = us.user_id AND us.date >= ? AND us.date <= ?
            WHERE (u.is_admin=0 AND u.is_owner=0)
            GROUP BY u.user_id HAVING earned > 0
            ORDER BY earned DESC LIMIT 10
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        for i, r_earn in enumerate(db.cursor.fetchall()):
            stats_data['top_earners'].append({
                'rank': i + 1,
                'username': f"@{r_earn['username'] or r_earn['first_name']}",
                'earned': float(_d(r_earn['earned'])),
            })

        # 6. ДЕТАЛИЗАЦИЯ с персональным индексом активности
        db.cursor.execute('''
            SELECT u.user_id, u.username, u.first_name, u.joined_at,
                   COALESCE(SUM(us.total_messages), 0)    as msgs,
                   COALESCE(SUM(us.total_chars), 0)       as chars,
                   COALESCE(SUM(us.total_words), 0)       as words,
                   COALESCE(SUM(us.reactions_given), 0)   as reacts_g,
                   COALESCE(SUM(us.reactions_received), 0) as reacts_r,
                   COALESCE(SUM(us.replies_sent), 0)      as replies_s,
                   COALESCE(SUM(us.replies_received), 0)  as replies_r,
                   COALESCE(SUM(us.mentions_received), 0) as kupp_r,
                   COALESCE(SUM(us.media_sent), 0)        as media_s,
                   COALESCE(SUM(us.other_threads_posts), 0) as threads,
                   COALESCE(SUM(us.pulses_mined), 0)      as earned
            FROM users u LEFT JOIN user_stats us ON u.user_id = us.user_id AND us.date >= ? AND us.date <= ?
            WHERE (u.is_admin=0 AND u.is_owner=0)
            GROUP BY u.user_id HAVING msgs > 0 ORDER BY earned DESC
        ''', (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

        CHARS_NORM = Decimal('100')
        for r_user in db.cursor.fetchall():
            _msgs  = _d(r_user['msgs'])
            _chars = _d(r_user['chars'])
            _avg_l = _chars / _msgs if _msgs > 0 else Decimal('0')

            # Индекс активности — Decimal арифметика (ЕДИНАЯ ФОРМУЛА)
            u_idx = (
                Decimal('0.05') * (_chars / CHARS_NORM) +
                Decimal('0.05') * (_avg_l / CHARS_NORM) +
                Decimal('0.05') * _msgs +
                Decimal('0.08') * _d(r_user['reacts_g']) +
                Decimal('0.10') * _d(r_user['reacts_r']) +
                Decimal('0.18') * _d(r_user['replies_r']) +
                Decimal('0.15') * _d(r_user['replies_s']) +
                Decimal('0.15') * _d(r_user['kupp_r']) +
                Decimal('0.07') * _d(r_user['media_s']) +
                Decimal('0.12') * _d(r_user['threads'])
            ) / divisor

            u_name = (
                f"@{r_user['username']}" if r_user['username']
                else (r_user['first_name'] or f"ID:{r_user['user_id']}")
            )
            stats_data['detailed_stats'][u_name] = {
                'user_id':           r_user['user_id'],
                'messages':          int(_msgs),
                'chars':             int(_chars),
                'words':             int(_d(r_user['words'])),
                'reactions':         int(_d(r_user['reacts_g'])),
                'reactions_received': int(_d(r_user['reacts_r'])),
                'replies':           int(_d(r_user['replies_s'])),
                'replies_received':  int(_d(r_user['replies_r'])),
                'mentions_received': int(_d(r_user['kupp_r'])),
                'media':             int(_d(r_user['media_s'])),
                'other_threads_posts': int(_d(r_user['threads'])),
                'days_in_chat':      calculate_days_in_chat(r_user['joined_at']),
                'earned':            float(round_decimal(_d(r_user['earned']), 2)),  # ← Decimal
                'activity_index':    float(round_decimal(u_idx, 3)),                 # ← Decimal
            }

        if period in ['week', 'month', 'year']:

            def collect_day_stats(day):
                """Собирает статистические данные за один день."""
                db.cursor.execute('''
                    SELECT
                        COALESCE(SUM(total_messages), 0)             as messages,
                        COALESCE(SUM(total_chars), 0)                as chars,
                        COALESCE(SUM(total_words), 0)                as words,
                        COALESCE(SUM(total_reactions), 0)            as reactions,
                        COALESCE(SUM(total_media), 0)                as media,
                        COALESCE(SUM(total_pulses_mined), 0)         as pulses,
                        COALESCE(SUM(total_messages_with_admins), 0) as msgs_with_admins,
                        COALESCE(SUM(total_messages_without_admins), 0) as msgs_without_admins,
                        COALESCE(AVG(avg_message_length), 0)         as avg_msg_length
                    FROM chat_stats WHERE date = ?
                ''', (day.strftime('%Y-%m-%d'),))
                day_data = db.cursor.fetchone()

                db.cursor.execute('''
                    SELECT COUNT(DISTINCT user_id) as active_users
                    FROM user_stats WHERE date = ? AND total_messages > 0
                ''', (day.strftime('%Y-%m-%d'),))
                row = db.cursor.fetchone()
                active_users = row['active_users'] if row else 0

                db.cursor.execute(
                    'SELECT COUNT(*) as joined FROM users WHERE DATE(joined_at) = ?',
                    (day.strftime('%Y-%m-%d'),)
                )
                row = db.cursor.fetchone()
                joined = row['joined'] if row else 0

                db.cursor.execute(
                    "SELECT COUNT(*) as left_users FROM transactions "
                    "WHERE transaction_type = 'return_on_leave' AND DATE(timestamp) = ?",
                    (day.strftime('%Y-%m-%d'),)
                )
                row = db.cursor.fetchone()
                left_users = row['left_users'] if row else 0

                db.cursor.execute(
                    "SELECT COUNT(*) as total_users FROM users WHERE joined_at <= ?",
                    (day.strftime('%Y-%m-%d 23:59:59'),)
                )
                row = db.cursor.fetchone()
                total_users = row['total_users'] if row else 0

                # Вовлечённость — Decimal
                engagement = float(
                    round_decimal(_d(active_users) / _d(total_users) * Decimal('100'), 2)
                    if total_users > 0 else Decimal('0')
                )

                db.cursor.execute('''
                    SELECT
                        COALESCE(SUM(reactions_given), 0)    as reactions_given,
                        COALESCE(SUM(reactions_received), 0) as reactions_received,
                        COALESCE(SUM(replies_sent), 0)       as replies_sent,
                        COALESCE(SUM(replies_received), 0)   as replies_received,
                        COALESCE(SUM(mentions_received), 0)  as mentions,
                        COALESCE(SUM(other_threads_posts), 0) as other_threads
                    FROM user_stats WHERE date = ?
                ''', (day.strftime('%Y-%m-%d'),))
                us = db.cursor.fetchone()

                return {
                    'messages':            int(_d(day_data['messages']      if day_data else 0)),
                    'chars':               int(_d(day_data['chars']         if day_data else 0)),
                    'words':               int(_d(day_data['words']         if day_data else 0)),
                    'reactions':           int(_d(day_data['reactions']     if day_data else 0)),
                    'media':               int(_d(day_data['media']         if day_data else 0)),
                    'pulses':              float(round_decimal(_d(day_data['pulses'] if day_data else 0), 2)),
                    'msgs_with_admins':    int(_d(day_data['msgs_with_admins']    if day_data else 0)),
                    'msgs_without_admins': int(_d(day_data['msgs_without_admins'] if day_data else 0)),
                    'avg_msg_length':      float(round_decimal(_d(day_data['avg_msg_length'] if day_data else 0), 2)),
                    'active_users':        int(active_users),
                    'total_users':         int(total_users),
                    'joined':              int(joined),
                    'left_users':          int(left_users),
                    'engagement':          engagement,
                    'reactions_given':     int(_d(us['reactions_given']    if us else 0)),
                    'reactions_received':  int(_d(us['reactions_received'] if us else 0)),
                    'replies_sent':        int(_d(us['replies_sent']       if us else 0)),
                    'replies_received':    int(_d(us['replies_received']   if us else 0)),
                    'mentions':            int(_d(us['mentions']            if us else 0)),
                    'other_threads':       int(_d(us['other_threads']      if us else 0)),
                }

            def collect_month_stats(first_day, last_day):
                """Собирает статистические данные за месяц."""
                db.cursor.execute('''
                    SELECT
                        COALESCE(SUM(total_messages), 0)             as messages,
                        COALESCE(SUM(total_chars), 0)                as chars,
                        COALESCE(SUM(total_words), 0)                as words,
                        COALESCE(SUM(total_reactions), 0)            as reactions,
                        COALESCE(SUM(total_media), 0)                as media,
                        COALESCE(SUM(total_pulses_mined), 0)         as pulses,
                        COALESCE(SUM(total_messages_with_admins), 0) as msgs_with_admins,
                        COALESCE(SUM(total_messages_without_admins), 0) as msgs_without_admins,
                        COALESCE(AVG(avg_message_length), 0)         as avg_msg_length
                    FROM chat_stats WHERE date >= ? AND date <= ?
                ''', (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d')))
                month_data = db.cursor.fetchone()

                db.cursor.execute('''
                    SELECT COUNT(DISTINCT user_id) as active_users
                    FROM user_stats WHERE date >= ? AND date <= ? AND total_messages > 0
                ''', (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d')))
                row = db.cursor.fetchone()
                active_users = row['active_users'] if row else 0

                db.cursor.execute(
                    'SELECT COUNT(*) as joined FROM users WHERE DATE(joined_at) >= ? AND DATE(joined_at) <= ?',
                    (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d'))
                )
                row = db.cursor.fetchone()
                joined = row['joined'] if row else 0

                db.cursor.execute(
                    "SELECT COUNT(*) as left_users FROM transactions "
                    "WHERE transaction_type = 'return_on_leave' AND DATE(timestamp) >= ? AND DATE(timestamp) <= ?",
                    (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d'))
                )
                row = db.cursor.fetchone()
                left_users = row['left_users'] if row else 0

                db.cursor.execute(
                    "SELECT COUNT(*) as total_users FROM users WHERE joined_at <= ?",
                    (last_day.strftime('%Y-%m-%d 23:59:59'),)
                )
                row = db.cursor.fetchone()
                total_users = row['total_users'] if row else 0

                engagement = float(
                    round_decimal(_d(active_users) / _d(total_users) * Decimal('100'), 2)
                    if total_users > 0 else Decimal('0')
                )

                db.cursor.execute('''
                    SELECT
                        COALESCE(SUM(reactions_given), 0)    as reactions_given,
                        COALESCE(SUM(reactions_received), 0) as reactions_received,
                        COALESCE(SUM(replies_sent), 0)       as replies_sent,
                        COALESCE(SUM(replies_received), 0)   as replies_received,
                        COALESCE(SUM(mentions_received), 0)  as mentions,
                        COALESCE(SUM(other_threads_posts), 0) as other_threads
                    FROM user_stats WHERE date >= ? AND date <= ?
                ''', (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d')))
                us = db.cursor.fetchone()

                return {
                    'messages':            int(_d(month_data['messages']      if month_data else 0)),
                    'chars':               int(_d(month_data['chars']         if month_data else 0)),
                    'words':               int(_d(month_data['words']         if month_data else 0)),
                    'reactions':           int(_d(month_data['reactions']     if month_data else 0)),
                    'media':               int(_d(month_data['media']         if month_data else 0)),
                    'pulses':              float(round_decimal(_d(month_data['pulses'] if month_data else 0), 2)),
                    'msgs_with_admins':    int(_d(month_data['msgs_with_admins']    if month_data else 0)),
                    'msgs_without_admins': int(_d(month_data['msgs_without_admins'] if month_data else 0)),
                    'avg_msg_length':      float(round_decimal(_d(month_data['avg_msg_length'] if month_data else 0), 2)),
                    'active_users':        int(active_users),
                    'total_users':         int(total_users),
                    'joined':              int(joined),
                    'left_users':          int(left_users),
                    'engagement':          engagement,
                    'reactions_given':     int(_d(us['reactions_given']    if us else 0)),
                    'reactions_received':  int(_d(us['reactions_received'] if us else 0)),
                    'replies_sent':        int(_d(us['replies_sent']       if us else 0)),
                    'replies_received':    int(_d(us['replies_received']   if us else 0)),
                    'mentions':            int(_d(us['mentions']            if us else 0)),
                    'other_threads':       int(_d(us['other_threads']      if us else 0)),
                }

            weekdays = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВСК']

            if period == 'week':
                current = start_date.date()
                while current <= end_date.date():
                    day_stats = collect_day_stats(current)
                    day_stats['date'] = f"{current.strftime('%d.%m.%y')} {weekdays[current.weekday()]}"
                    stats_data['daily_stats'].append(day_stats)
                    current += timedelta(days=1)

            elif period == 'month':
                current = start_date.date()
                while current <= end_date.date():
                    day_stats = collect_day_stats(current)
                    day_stats['date'] = f"{current.strftime('%d.%m.%y')} {weekdays[current.weekday()]}"
                    stats_data['daily_stats'].append(day_stats)
                    current += timedelta(days=1)

            elif period == 'year':
                cy, cm_ = start_date.year, start_date.month
                while (cy < end_date.year) or (cy == end_date.year and cm_ <= end_date.month):
                    first_day = date_type(cy, cm_, 1)
                    last_day  = date_type(cy, cm_, calendar.monthrange(cy, cm_)[1])
                    month_stats = collect_month_stats(first_day, last_day)
                    month_stats['date']  = f"{cm_:02d}"
                    month_stats['month'] = cm_
                    month_stats['year']  = cy
                    stats_data['daily_stats'].append(month_stats)
                    cm_ += 1
                    if cm_ > 12:
                        cm_ = 1
                        cy += 1

        # ── Детальная статистика майнинга (для листа "СтатМайнинг") ────
        from handlers.messages.mining_logic import (
            COMBO_LABELS, SPRINTS_CONFIG, PENALTY_LABELS,
        )

        MINING_DESCRIPTIONS = {
            # Комбо
            'writer':       {'label': '✍️ Писатель',       'desc': 'Текст > 50 символов'},
            'illustrator':  {'label': '🖼 Иллюстратор',    'desc': 'Текст > 50 символов + Фото'},
            'reviewer':     {'label': '🎬 Обозреватель',    'desc': 'Видео + Текст > 100 слов'},
            'dj':           {'label': '🎧 Диджей',          'desc': 'Ссылка на плейлист'},
            'sharp_tongue': {'label': '🗡 Острый язык',     'desc': '> 2 ответов на пост'},
            'viral_post':   {'label': '📢 Вирусный пост',   'desc': '> 2 лайков'},
            'hit_post':     {'label': '💥 Хит-пост',        'desc': '4+ лайка'},
            'legend_post':  {'label': '👑 Легенда',          'desc': '6+ лайков'},
            # Спринты
            'chat_core':    {'label': '💬 Основа чата',      'desc': '10 сообщений за 24ч'},
            'emotional':    {'label': '😍 Эмоциональный',    'desc': '10 эмодзи за 24ч'},
            'photographer': {'label': '📸 Фотограф',         'desc': '3+ фото за 24ч'},
            'director':     {'label': '🎬 Режиссёр',         'desc': '2 видео за 24ч (спец. ветки)'},
            'music_lover':  {'label': '🎧 Меломан',          'desc': '5 аудио за 24ч (Музыка)'},
            'face_seller':  {'label': '🫧 Лицом торгуешь',   'desc': '5 кружков за 12ч'},
            'center':       {'label': '🎯 Центр внимания',   'desc': '20 ответов за 12ч'},
            'radio':        {'label': '🎙 Радио',             'desc': '4 голосовых за 1ч'},
            'gif_room':     {'label': '🎞 Гифошная',          'desc': '10 гифок за 1ч'},
            'chatterbox':   {'label': '🗣 Болтун',            'desc': '20 ответов за 1ч'},
            'generous':     {'label': '💝 Щедрая душа',      'desc': '5 лайков за 1ч'},
            'favorite':     {'label': '⭐ Любимчик',          'desc': '10 полученных лайков за 1ч'},
            # Штрафы
            'copypaste':    {'label': '📋 Копипаст',          'desc': 'Дубликат сообщения за час'},
            'wrong_door':   {'label': '🚪 Не та дверь',      'desc': 'Контент не по теме ветки'},
            'toxic':        {'label': '☠️ Токсик',            'desc': 'Токсичное поведение'},
        }

        try:
            sd_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
            ed_str = end_date.strftime('%Y-%m-%d %H:%M:%S')

            # 1. Базовая добыча
            db.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM transactions
                WHERE transaction_type = 'message_reward'
                  AND timestamp >= ? AND timestamp <= ?
            ''', (sd_str, ed_str))
            base_total = float(db.cursor.fetchone()['total'])

            # 2. Комбо (поимённо)
            combos = []
            try:
                db.cursor.execute('''
                    SELECT combo_name, COALESCE(SUM(reward), 0) AS total
                    FROM combo_claims
                    WHERE claimed_at >= ? AND claimed_at <= ?
                    GROUP BY combo_name ORDER BY total DESC
                ''', (sd_str, ed_str))
                combos_raw = db.cursor.fetchall()
            except Exception:
                combos_raw = []
            for r in combos_raw:
                name = r['combo_name']
                info = MINING_DESCRIPTIONS.get(name, {'label': name, 'desc': ''})
                combos.append({'name': name, 'label': info['label'],
                               'description': info['desc'], 'sum': float(r['total'])})

            # 3. Спринты (поимённо)
            sprints = []
            try:
                db.cursor.execute('''
                    SELECT sprint_name, COALESCE(SUM(reward), 0) AS total
                    FROM sprint_claims
                    WHERE claimed_at >= ? AND claimed_at <= ?
                    GROUP BY sprint_name ORDER BY total DESC
                ''', (sd_str, ed_str))
                sprints_raw = db.cursor.fetchall()
            except Exception:
                sprints_raw = []
            for r in sprints_raw:
                name = r['sprint_name']
                info = MINING_DESCRIPTIONS.get(name, {'label': name, 'desc': ''})
                sprints.append({'name': name, 'label': info['label'],
                                'description': info['desc'], 'sum': float(r['total'])})

            # 4. Штрафы (поимённо)
            penalties = []
            try:
                db.cursor.execute('''
                    SELECT description, COALESCE(SUM(amount), 0) AS total
                    FROM transactions
                    WHERE transaction_type = 'penalty_deduct'
                      AND timestamp >= ? AND timestamp <= ?
                    GROUP BY description ORDER BY total DESC
                ''', (sd_str, ed_str))
                key_map = {'копипаст': 'copypaste', 'не в ту дверь (нарушение)': 'wrong_door',
                           'удаление (токсик)': 'toxic'}
                for r in db.cursor.fetchall():
                    desc = r['description'] or ''
                    raw_key = desc.replace('Штраф: ', '').strip().lower()
                    mapped = key_map.get(raw_key, raw_key)
                    info = MINING_DESCRIPTIONS.get(mapped, {'label': desc, 'desc': ''})
                    penalties.append({'name': mapped, 'label': info['label'],
                                      'description': info['desc'], 'sum': float(r['total'])})
            except Exception:
                pass

            stats_data['mining_detailed'] = {
                'base_total': base_total,
                'combos': combos,
                'sprints': sprints,
                'penalties': penalties,
            }
        except Exception as e:
            logging.warning(f"Mining detailed stats failed: {e}")
            stats_data['mining_detailed'] = None

        # Генерация файла
        timestamp = get_moscow_time().strftime('%Y%m%d_%H%M%S')
        os.makedirs('logs', exist_ok=True)

        try:
            joins_data = db.get_user_joins(
                start_date=start_date.strftime('%Y-%m-%d %H:%M:%S'),
                end_date=end_date.strftime('%Y-%m-%d %H:%M:%S')
            )
            stats_data['user_joins'] = [dict(j) for j in joins_data] if joins_data else []
        except Exception as e:
            logging.error(f"Error collecting joins data: {e}")
            stats_data['user_joins'] = []

        if file_format == 'pdf':
            filename = f'stats_{period}_{timestamp}.pdf'
            filepath = os.path.join('logs', filename)
            logging.info(f"Generating PDF export to: {filepath}")
            result = export_stats_to_pdf(stats_data, filepath)
            if result is None:
                logging.error("export_stats_to_pdf returned None")
                await query.edit_message_text(
                    "❌ Ошибка при создании PDF. Проверьте логи.\n"
                    "Убедитесь что установлен пакет: pip install reportlab",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("🔙 Назад в меню", callback_data="back_to_menu")
                    ]])
                )
                return
            caption = f"📕 Статистика чата ({period_name}) — PDF"
        else:
            filename = f'stats_{period}_{timestamp}.xlsx'
            filepath = os.path.join('logs', filename)
            logging.info(f"Generating Excel export to: {filepath}")
            result = export_stats_to_excel(stats_data, filepath)
            if result is None:
                logging.error("export_stats_to_excel returned None - file not created")
                await query.edit_message_text(
                    "❌ Ошибка при создании файла. Проверьте логи.",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("🔙 Назад в меню", callback_data="back_to_menu")
                    ]])
                )
                return
            # Добавляем лист "Курс" с графиком
            _add_rate_sheet_to_excel(filepath, db, start_date, end_date, period_name)
            caption = f"📊 Статистика чата ({period_name}) — Excel"

        if os.path.exists(filepath):
            with open(filepath, 'rb') as file:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=file,
                    filename=filename,
                    caption=caption,
                    reply_to_message_id=query.message.message_id
                )
            os.remove(filepath)
            await query.edit_message_text(
                "✅ Отчёт успешно сгенерирован и отправлен!",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад в меню", callback_data="back_to_menu")
                ]])
            )
        else:
            await query.edit_message_text(
                "❌ Ошибка при создании файла",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Назад в меню", callback_data="back_to_menu")
                ]])
            )

    except Exception as e:
        logging.error(f"Error generating export: {e}")
        traceback.print_exc()
        await query.edit_message_text(
            f"❌ Произошла ошибка: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔙 Назад в меню", callback_data="back_to_menu")
            ]])
        )


# ═══════════════════════════════════════════════════════════════
# ТОП, ЛИДЕРБОРДЫ, МЕНЮ СТАТИСТИКИ
# ═══════════════════════════════════════════════════════════════

