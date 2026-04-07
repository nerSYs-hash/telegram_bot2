#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Генерация Excel-файла личной детализации пользователя.
Содержит: все транзакции (доходы/расходы), курс на дату выгрузки,
разбивку по типам, итоги.
"""

import logging
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.helpers import format_number

logger = logging.getLogger(__name__)

# ═══ СТИЛИ ═══
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECTION_FONT = Font(bold=True, size=13, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
EVEN_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GREEN_FONT = Font(color="27AE60")
RED_FONT = Font(color="E74C3C")
BOLD_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color="BDC3C7"),
    right=Side(style='thin', color="BDC3C7"),
    top=Side(style='thin', color="BDC3C7"),
    bottom=Side(style='thin', color="BDC3C7")
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Маппинг типов транзакций на человекочитаемые названия
TX_TYPE_NAMES = {
    'message_reward': '⛏ Майнинг',
    'transfer': '💸 Перевод',
    'referral_reward': '👥 Реферал',
    'lottery_win': '🎰 Лотерея (выигрыш)',
    'lottery_ticket': '🎟 Лотерея (билет)',
    'monthly_gift': '🎁 Подарок месяца',
    'reactor_donation': '🔋 Реактор',
    'admin_give': '🏦 От Центробанка',
    'return_on_leave': '↩️ Возврат в банк',
    'unfreeze_on_return': '🔓 Разморозка',
    'bank_transfer': '🏦 Перевод из банка',
    'combo_reward': '🎭 Тайное Комбо',
    'sprint_reward': '⚡ Тайный Спринт',
    'penalty_deduct': '🚫 Реальный Штраф',
    'secret_event': '🔮 Тайное событие',
    'easter_egg': '🔮 Тайная Пасхалка',
    'wandering_anomaly': '🔮 Блуждающая Аномалия',
    # BBS
    'bbs_popularity': '💫 Популярность BBS',
    # Реакции
    'reaction_given_reward': '❤️ Реакция (поставлена)',
    'reaction_received_reward': '❤️ Реакция (получена)',
    # Донаты
    'donate_to_user': '💝 Донат участнику',
    'donate_to_bank': '🏦 Донат в банк',
    # Бинго
    'bingo_win': '🎯 Бинго (выигрыш)',
    'bingo_ticket': '🎯 Бинго (билет)',
    'bingo_refund': '🎯 Бинго (возврат)',
    'bingo_penalty': '🎯 Бинго (штраф)',
    # Магазин
    'shop_purchase': '🛒 Покупка в магазине',
    'lootbox_win': '📦 Лутбокс (выигрыш)',
    # Лотерея
    'lottery_commission': '🎰 Комиссия лотереи',
    # Прочее
    'compensation_reward': '🔧 Компенсация',
}

# Маппинг описаний на чистые русские (для старых записей в БД)
TX_DESC_RU = {
    'Message reward': 'Награда',
    'Награда за сообщение': 'Награда',
    'User left chat, balance frozen for 30 days and returned to bank': 'Покинул чат, баланс заморожен на 30 дней и возвращён в банк',
    'User returned to chat, frozen balance restored': 'Вернулся в чат, замороженный баланс восстановлен',
    'Transfer from bank by owner': 'Перевод из банка владельцем',
    'Admin gift from': 'Выдача от администратора',
    'Direct transfer': 'Прямой перевод',
    'Donation to community reactor': 'Пожертвование в реактор',
    'Referral reward for user': 'Реферальная награда за пользователя',
    'Won lottery': 'Выигрыш в лотерее',
    'Monthly gift winner': 'Победитель подарка месяца',
}

def _get_tx_name(tx_type):
    return TX_TYPE_NAMES.get(tx_type, tx_type)

def _fmt(amount):
    """Форматирование: 4 знака для микро-сумм, иначе format_number."""
    if 0 < abs(amount) < 0.01:
        return f"{amount:.4f}"
    return format_number(amount)

def _translate_description(desc):
    """Перевод и очистка описаний (убираем суммы, переводим английские)."""
    if not desc:
        return ''
    for key, rus in TX_DESC_RU.items():
        if desc.startswith(key):
            return rus
    return desc


def _apply_row_style(ws, row, cols, idx):
    fill = EVEN_FILL if idx % 2 == 0 else ODD_FILL
    for c in cols:
        cell = ws[f'{c}{row}']
        cell.fill = fill
        cell.border = THIN_BORDER


def generate_user_detalization(db, user_id, period, filename):
    """
    Генерирует Excel-файл личной детализации пользователя.

    Args:
        db: Database instance
        user_id: ID пользователя
        period: 'day' | 'week' | 'month' | 'year'
        filename: путь к выходному файлу

    Returns:
        filepath или None при ошибке
    """
    try:
        from utils.helpers import format_number, get_moscow_time

        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)

        now = get_moscow_time()
        user = db.get_user(user_id)
        if not user:
            return None

        username = user['username'] or user['first_name'] or f"User{user_id}"
        balance = float(user['balance'])
        rate = db.get_exchange_rate()

        # Определяем диапазон дат
        if period == 'day':
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            period_name = f"За сегодня ({now.strftime('%d.%m.%Y')})"
        elif period == 'week':
            start = now - timedelta(days=7)
            period_name = "За последние 7 дней"
        elif period == 'month':
            start = now - timedelta(days=30)
            period_name = "За последние 30 дней"
        elif period == 'year':
            start = now - timedelta(days=365)
            period_name = "За последний год"
        else:
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            period_name = "За сегодня"

        start_str = start.strftime('%Y-%m-%d %H:%M:%S')
        end_str = now.strftime('%Y-%m-%d %H:%M:%S')

        # ═══ ПОЛУЧАЕМ ТРАНЗАКЦИИ ═══
        db.cursor.execute('''
            SELECT t.id, t.from_user_id, t.to_user_id, t.amount, 
                   t.transaction_type, t.description, t.timestamp,
                   u_from.username AS from_username, u_from.first_name AS from_name,
                   u_to.username AS to_username, u_to.first_name AS to_name
            FROM transactions t
            LEFT JOIN users u_from ON t.from_user_id = u_from.user_id
            LEFT JOIN users u_to ON t.to_user_id = u_to.user_id
            WHERE (t.from_user_id = ? OR t.to_user_id = ?)
              AND t.timestamp >= ?
            ORDER BY t.timestamp DESC
        ''', (user_id, user_id, start_str))

        transactions = db.cursor.fetchall()

        # ═══ СОЗДАЁМ WORKBOOK ═══
        wb = Workbook()
        ws = wb.active
        ws.title = "Детализация"

        row = 1

        # ═══ ЗАГОЛОВОК ═══
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"💎 ДЕТАЛИЗАЦИЯ ПУЛЬСОВ — @{username}"
        ws[f'A{row}'].font = TITLE_FONT
        ws[f'A{row}'].fill = TITLE_FILL
        ws[f'A{row}'].alignment = CENTER
        ws.row_dimensions[row].height = 30
        row += 1

        # Информация о пользователе
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = (
            f"📅 Период: {period_name}  |  "
            f"💰 Баланс: {format_number(balance)} 💎  |  "
            f"💱 Курс: {rate:.6f} ₽/Пульс  |  "
            f"💵 ≈ {balance * rate:.4f} ₽  |  "
            f"🕐 Выгружено: {now.strftime('%d.%m.%Y %H:%M')} МСК"
        )
        ws[f'A{row}'].font = Font(size=9, italic=True)
        ws[f'A{row}'].alignment = CENTER
        ws.row_dimensions[row].height = 18
        row += 2

        # ═══ СЕКЦИЯ: СВОДКА ═══
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "📊 СВОДКА ЗА ПЕРИОД"
        ws[f'A{row}'].font = SECTION_FONT
        ws[f'A{row}'].fill = SECTION_FILL
        ws[f'A{row}'].alignment = LEFT
        ws.row_dimensions[row].height = 25
        row += 1

        # Считаем итоги по категориям
        income_by_type = {}
        expense_by_type = {}
        total_in = 0.0
        total_out = 0.0

        for tx in transactions:
            amount = float(tx['amount'])
            tx_type = tx['transaction_type']

            if tx['to_user_id'] == user_id and tx['from_user_id'] != user_id:
                income_by_type[tx_type] = income_by_type.get(tx_type, 0) + amount
                total_in += amount
            elif tx['from_user_id'] == user_id and tx['to_user_id'] != user_id:
                expense_by_type[tx_type] = expense_by_type.get(tx_type, 0) + amount
                total_out += amount
            elif tx['to_user_id'] == user_id and tx['from_user_id'] is None:
                income_by_type[tx_type] = income_by_type.get(tx_type, 0) + amount
                total_in += amount
            elif tx['from_user_id'] == user_id and tx['to_user_id'] is None:
                expense_by_type[tx_type] = expense_by_type.get(tx_type, 0) + amount
                total_out += amount

        # Заголовки сводки
        summary_headers = ['Категория', 'Сумма 💎', '₽ по курсу']
        for ci, h in enumerate(summary_headers):
            c = get_column_letter(ci + 1)
            ws[f'{c}{row}'] = h
            ws[f'{c}{row}'].font = HEADER_FONT
            ws[f'{c}{row}'].fill = HEADER_FILL
            ws[f'{c}{row}'].alignment = CENTER
            ws[f'{c}{row}'].border = THIN_BORDER
        row += 1

        idx = 0
        # Доходы
        for tx_type, amount in sorted(income_by_type.items(), key=lambda x: -x[1]):
            ws[f'A{row}'] = f"📥 {_get_tx_name(tx_type)}"
            ws[f'B{row}'] = f"+{_fmt(amount)}"
            ws[f'B{row}'].font = GREEN_FONT
            ws[f'C{row}'] = f"{amount * rate:.4f} ₽"
            _apply_row_style(ws, row, ['A', 'B', 'C'], idx)
            ws[f'A{row}'].alignment = LEFT
            ws[f'B{row}'].alignment = RIGHT
            ws[f'C{row}'].alignment = RIGHT
            row += 1
            idx += 1

        # Расходы
        for tx_type, amount in sorted(expense_by_type.items(), key=lambda x: -x[1]):
            ws[f'A{row}'] = f"📤 {_get_tx_name(tx_type)}"
            ws[f'B{row}'] = f"-{_fmt(amount)}"
            ws[f'B{row}'].font = RED_FONT
            ws[f'C{row}'] = f"-{amount * rate:.4f} ₽"
            _apply_row_style(ws, row, ['A', 'B', 'C'], idx)
            ws[f'A{row}'].alignment = LEFT
            ws[f'B{row}'].alignment = RIGHT
            ws[f'C{row}'].alignment = RIGHT
            row += 1
            idx += 1

        # Итого
        net = total_in - total_out
        sign = "+" if net >= 0 else ""
        ws[f'A{row}'] = "ИТОГО"
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'] = f"{sign}{_fmt(net)}"
        ws[f'B{row}'].font = Font(bold=True, color="27AE60" if net >= 0 else "E74C3C")
        ws[f'C{row}'] = f"{sign}{net * rate:.4f} ₽"
        ws[f'C{row}'].font = BOLD_FONT
        for c in ['A', 'B', 'C']:
            ws[f'{c}{row}'].fill = TOTAL_FILL
            ws[f'{c}{row}'].border = THIN_BORDER
        ws[f'A{row}'].alignment = LEFT
        ws[f'B{row}'].alignment = RIGHT
        ws[f'C{row}'].alignment = RIGHT
        row += 2

        # ═══ СЕКЦИЯ: ВСЕ ТРАНЗАКЦИИ ═══
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "📋 ВСЕ ОПЕРАЦИИ"
        ws[f'A{row}'].font = SECTION_FONT
        ws[f'A{row}'].fill = SECTION_FILL
        ws[f'A{row}'].alignment = LEFT
        ws.row_dimensions[row].height = 25
        row += 1

        # Заголовки транзакций
        tx_headers = ['Дата', 'Время', 'Тип', 'Контрагент', 'Приход 💎', 'Расход 💎', 'Описание']
        tx_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        for ci, h in enumerate(tx_headers):
            c = tx_cols[ci]
            ws[f'{c}{row}'] = h
            ws[f'{c}{row}'].font = HEADER_FONT
            ws[f'{c}{row}'].fill = HEADER_FILL
            ws[f'{c}{row}'].alignment = CENTER
            ws[f'{c}{row}'].border = THIN_BORDER
        ws.row_dimensions[row].height = 20
        row += 1

        # Данные транзакций
        for idx, tx in enumerate(transactions):
            amount = float(tx['amount'])
            tx_type = tx['transaction_type']

            # Форматирование с поддержкой микро-сумм
            display_amount = _fmt(amount)

            # Пропускаем строки где сумма = 0
            if amount == 0 or display_amount in ('0.00', '0', '0.0', '0.0000'):
                continue

            # Дата и время
            ts = str(tx['timestamp'])
            try:
                dt = datetime.strptime(ts[:19], '%Y-%m-%d %H:%M:%S')
                date_str = dt.strftime('%d.%m.%Y')
                time_str = dt.strftime('%H:%M')
            except (ValueError, TypeError):
                date_str = ts[:10] if len(ts) >= 10 else ts
                time_str = ts[11:16] if len(ts) >= 16 else ''

            # Определяем направление
            is_income = (tx['to_user_id'] == user_id and tx['from_user_id'] != user_id) or \
                        (tx['to_user_id'] == user_id and tx['from_user_id'] is None)

            # Контрагент
            if is_income:
                if tx['from_user_id'] is None:
                    counterparty = "🏦 Центробанк"
                else:
                    name = tx['from_username'] or tx['from_name'] or f"ID:{tx['from_user_id']}"
                    counterparty = f"@{name}"
                income_val = f"+{display_amount}"
                expense_val = ""
            else:
                if tx['to_user_id'] is None:
                    counterparty = "🏦 Центробанк"
                else:
                    name = tx['to_username'] or tx['to_name'] or f"ID:{tx['to_user_id']}"
                    counterparty = f"@{name}"
                income_val = ""
                expense_val = f"-{display_amount}"

            ws[f'A{row}'] = date_str
            ws[f'B{row}'] = time_str
            ws[f'C{row}'] = _get_tx_name(tx_type)
            ws[f'D{row}'] = counterparty
            ws[f'E{row}'] = income_val
            ws[f'F{row}'] = expense_val
            ws[f'G{row}'] = _translate_description(tx['description'])

            # Цвет для приход/расход
            if income_val:
                ws[f'E{row}'].font = GREEN_FONT
            if expense_val:
                ws[f'F{row}'].font = RED_FONT

            _apply_row_style(ws, row, tx_cols, idx)
            ws[f'A{row}'].alignment = CENTER
            ws[f'B{row}'].alignment = CENTER
            ws[f'C{row}'].alignment = LEFT
            ws[f'D{row}'].alignment = LEFT
            ws[f'E{row}'].alignment = RIGHT
            ws[f'F{row}'].alignment = RIGHT
            ws[f'G{row}'].alignment = LEFT

            row += 1

        # Итоговая строка транзакций
        ws[f'A{row}'] = "ИТОГО"
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'E{row}'] = f"+{_fmt(total_in)}"
        ws[f'E{row}'].font = Font(bold=True, color="27AE60")
        ws[f'F{row}'] = f"-{_fmt(total_out)}"
        ws[f'F{row}'].font = Font(bold=True, color="E74C3C")
        for c in tx_cols:
            ws[f'{c}{row}'].fill = TOTAL_FILL
            ws[f'{c}{row}'].border = THIN_BORDER
        ws[f'E{row}'].alignment = RIGHT
        ws[f'F{row}'].alignment = RIGHT
        row += 2

        # ═══ СЕКЦИЯ: КУРС ═══
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "💱 КУРС НА ДАТУ ВЫГРУЗКИ"
        ws[f'A{row}'].font = SECTION_FONT
        ws[f'A{row}'].fill = SECTION_FILL
        ws[f'A{row}'].alignment = LEFT
        ws.row_dimensions[row].height = 25
        row += 1

        rate_info = [
            ("Курс 1 Пульса", f"{rate:.6f} ₽"),
            ("Ваш баланс", f"{format_number(balance)} 💎"),
            ("Эквивалент в рублях", f"{balance * rate:.4f} ₽"),
            ("Дата выгрузки", now.strftime('%d.%m.%Y %H:%M МСК')),
        ]

        for ri, (label, value) in enumerate(rate_info):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            _apply_row_style(ws, row, ['A', 'B'], ri)
            ws[f'A{row}'].alignment = LEFT
            ws[f'B{row}'].alignment = RIGHT
            ws[f'A{row}'].font = BOLD_FONT
            row += 1

        # ═══ ШИРИНА КОЛОНОК ═══
        widths = {'A': 14, 'B': 10, 'C': 22, 'D': 20, 'E': 16, 'F': 16, 'G': 30}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        ws.freeze_panes = 'A5'

        # ═══ СОХРАНЕНИЕ ═══
        wb.save(filename)
        logger.info(f"✅ User detalization Excel created: {filename} ({len(transactions)} tx)")
        return filename

    except Exception as e:
        logger.error(f"❌ Error creating user detalization: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
