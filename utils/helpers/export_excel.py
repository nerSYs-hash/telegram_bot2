#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Экспорт статистики в Excel (openpyxl).

Data Integrity: все числовые значения записываются в ячейки как Decimal/float,
а НЕ как отформатированные строки. Визуальное форматирование реализовано
через number_format ячейки, что позволяет Excel корректно:
  - суммировать и фильтровать данные
  - строить графики
  - сортировать по числовым значениям
"""

import os
import re
import logging
from datetime import datetime
from decimal import Decimal

from utils.helpers.formatting import (
    format_number,        # только для строк-отображения
    format_month_column,
    to_decimal,
    to_float,
    round_decimal,
)

# ── Числовые форматы Excel (визуальное представление) ────────────────────────
_FMT_INT     = '#,##0'        # целые:    1 234
_FMT_FLOAT2  = '#,##0.00'     # 2 знака:  1 234.56
_FMT_FLOAT3  = '#,##0.000'    # индексы:  0.941
_FMT_PCT     = '0.0"%"'       # проценты: 12.5%


def _n(val, places: int = 2) -> float:
    """Конвертирует val в float через Decimal для записи в Excel-ячейку."""
    return float(round_decimal(val, places))


def export_stats_to_excel(stats_data, filename):
    """Export statistics to Excel file - beautiful single sheet."""
    try:
        logging.info(f"Starting Excel export to: {filename}")
        logging.info(f"Stats data keys: {stats_data.keys() if stats_data else 'None'}")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"

        # ── СТИЛИ ─────────────────────────────────────────────────────────────
        title_font    = Font(bold=True, size=18, color="FFFFFF")
        title_fill    = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        title_align   = Alignment(horizontal='center', vertical='center')

        section_font  = Font(bold=True, size=14, color="FFFFFF")
        section_fill  = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        section_align = Alignment(horizontal='left', vertical='center')

        header_font   = Font(bold=True, size=11, color="FFFFFF")
        header_fill   = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
        header_align  = Alignment(horizontal='center', vertical='center')

        even_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin',   color="BDC3C7"),
            right=Side(style='thin',  color="BDC3C7"),
            top=Side(style='thin',    color="BDC3C7"),
            bottom=Side(style='thin', color="BDC3C7"),
        )
        thick_border = Border(
            left=Side(style='medium',   color="34495E"),
            right=Side(style='medium',  color="34495E"),
            top=Side(style='medium',    color="34495E"),
            bottom=Side(style='medium', color="34495E"),
        )

        left_align   = Alignment(horizontal='left',   vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')
        right_align  = Alignment(horizontal='right',  vertical='center')

        # ── ХЕЛПЕР: разбить "индекс (значение)" → (float, float) ─────────────
        def parse_stat_value(val):
            """
            '0.941 (6 275.00)' → (0.941, 6275.0)  — оба float
            Если формат не совпадает → (None, исходная_строка)
            """
            s = str(val).strip()
            m = re.match(r'^([\d.]+)\s*\(([\d\s.,]+)\)$', s)
            if m:
                try:
                    idx_f = _n(m.group(1).strip(), 3)
                    raw_f = _n(m.group(2).strip().replace(' ', ''), 2)
                    return idx_f, raw_f
                except Exception:
                    pass
            return None, s

        # ── ГЛАВНЫЙ ЗАГОЛОВОК ─────────────────────────────────────────────────
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = stats_data['title']
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = title_align
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 30

        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"Период: {stats_data['start_date']} - {stats_data['end_date']}"
        ws[f'A{row}'].font = Font(size=11, italic=True)
        ws[f'A{row}'].alignment = center_align
        ws.row_dimensions[row].height = 20

        row += 2

        # ── СЕКЦИЯ 1: ОБЩАЯ СТАТИСТИКА ────────────────────────────────────────
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "📋 ОБЩАЯ СТАТИСТИКА"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = section_align
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 25
        row += 1

        for col, hdr in zip(['A', 'B', 'C'], ['Параметр', 'Значение', 'Индекс']):
            ws[f'{col}{row}'] = hdr
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, (key, value) in enumerate(stats_data['general'].items()):
            index_val, raw_val = parse_stat_value(value)
            fill = even_fill if idx % 2 == 0 else odd_fill

            ws[f'A{row}'] = key
            ws[f'B{row}'] = raw_val                         # float если удалось, иначе str
            if isinstance(raw_val, float):
                ws[f'B{row}'].number_format = _FMT_FLOAT2

            if index_val is not None:
                ws[f'C{row}'] = index_val                   # ← float, не строка
                ws[f'C{row}'].number_format = _FMT_FLOAT3

            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = thin_border
            ws[f'A{row}'].alignment = left_align
            ws[f'B{row}'].alignment = right_align
            ws[f'C{row}'].alignment = right_align
            row += 1

        row += 1

        # ── СЕКЦИЯ 2: ТОП-10 АКТИВИСТОВ ──────────────────────────────────────
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "⚡ ТОП-10 АКТИВИСТОВ"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = section_align
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 25
        row += 1

        for col, hdr in zip(['A', 'B', 'C', 'D'], ['№', 'Пользователь', 'Активность 🏅', 'Заработано 💎']):
            ws[f'{col}{row}'] = hdr
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, user_data in enumerate(stats_data.get('top_messages', [])[:10]):
            fill = even_fill if idx % 2 == 0 else odd_fill

            ws[f'A{row}'] = int(to_decimal(user_data.get('rank', idx + 1)))
            ws[f'B{row}'] = user_data.get('username', user_data.get('Пользователь', 'Unknown'))

            ws[f'C{row}'] = _n(user_data.get('activity_score', 0), 3)  # float ← было строка
            ws[f'C{row}'].number_format = _FMT_FLOAT3

            ws[f'D{row}'] = _n(user_data.get('earned_raw', user_data.get('earned', 0)), 2)  # float ← было строка
            ws[f'D{row}'].number_format = _FMT_FLOAT2

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = thin_border
            ws[f'A{row}'].alignment = center_align
            ws[f'B{row}'].alignment = left_align
            ws[f'C{row}'].alignment = center_align
            ws[f'D{row}'].alignment = center_align

            if idx == 0:   ws[f'A{row}'].font = Font(bold=True, color="FFD700", size=12)
            elif idx == 1: ws[f'A{row}'].font = Font(bold=True, color="C0C0C0", size=12)
            elif idx == 2: ws[f'A{row}'].font = Font(bold=True, color="CD7F32", size=12)
            row += 1

        row += 1

        # ── СЕКЦИЯ 3: ТОП-10 ПО ЗАРАБОТКУ ────────────────────────────────────
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "💎 ТОП-10 ПО ЗАРАБОТКУ"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = section_align
        ws[f'A{row}'].border = thick_border
        ws.row_dimensions[row].height = 25
        row += 1

        for col, hdr in zip(['A', 'B', 'C'], ['№', 'Пользователь', 'Заработано 💎']):
            ws[f'{col}{row}'] = hdr
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = thin_border
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, user_data in enumerate(stats_data.get('top_earners', [])[:10]):
            fill = even_fill if idx % 2 == 0 else odd_fill
            ws[f'A{row}'] = int(to_decimal(user_data.get('rank', idx + 1)))
            ws[f'B{row}'] = user_data.get('Пользователь', user_data.get('username', 'Unknown'))

            earned_val = user_data.get('💎 Добыто Пульсов', user_data.get('earned', 0))
            ws[f'C{row}'] = _n(earned_val, 2)               # float ← было строка "1 234.56"
            ws[f'C{row}'].number_format = _FMT_FLOAT2

            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = thin_border
            ws[f'A{row}'].alignment = center_align
            ws[f'B{row}'].alignment = left_align
            ws[f'C{row}'].alignment = center_align

            if idx == 0:   ws[f'A{row}'].font = Font(bold=True, color="FFD700", size=12)
            elif idx == 1: ws[f'A{row}'].font = Font(bold=True, color="C0C0C0", size=12)
            elif idx == 2: ws[f'A{row}'].font = Font(bold=True, color="CD7F32", size=12)
            row += 1

        row += 1

        # ── СЕКЦИЯ 4: ДЕТАЛЬНАЯ СТАТИСТИКА ───────────────────────────────────
        if stats_data.get('detailed_stats'):
            ws.merge_cells(f'A{row}:P{row}')
            ws[f'A{row}'] = "📊 ДЕТАЛЬНАЯ СТАТИСТИКА ВСЕХ ПОЛЬЗОВАТЕЛЕЙ"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].alignment = section_align
            ws[f'A{row}'].border = thick_border
            ws.row_dimensions[row].height = 25
            row += 1

            detail_headers = [
                '№', 'Пользователь',
                'Кол-во\n Сообщений', 'Всего\n Символов', 'Всего\n Слов', 'Средняя длина\n Сообщения',
                'Реакции\n Оставлено', 'Реакции\n Получено',
                'Ответов\n Получено', 'Ответов\n Отправлено',
                'Упоминания\n @username', 'Медиа\n Отправлено', 'Публик.\n в Ветках',
                'Дней\n в Чате', 'Заработано\n💎 Пульсов', 'Активность 🎖'
            ]
            detail_cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
            for col, hdr in zip(detail_cols, detail_headers):
                ws[f'{col}{row}'] = hdr
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].alignment = center_align
                ws[f'{col}{row}'].border = thin_border
            ws.row_dimensions[row].height = 35
            row += 1

            _member_count = stats_data.get('member_count', 10)
            _bot_count    = stats_data.get('bot_count', 1)
            _divisor      = max(_member_count - _bot_count - 1, 1)
            _CHARS_NORM   = Decimal('100')

            for idx, (user, data) in enumerate(stats_data['detailed_stats'].items()):
                _msgs  = int(to_decimal(data.get('messages', 0)))
                _chars = int(to_decimal(data.get('chars', 0)))
                _avg_len = _n(Decimal(_chars) / Decimal(_msgs), 2) if _msgs > 0 else 0.0

                fill = even_fill if idx % 2 == 0 else odd_fill

                ws[f'A{row}'] = idx + 1
                ws[f'B{row}'] = user
                ws[f'C{row}'] = _msgs                                        # int
                ws[f'D{row}'] = _chars                                       # int
                ws[f'E{row}'] = int(to_decimal(data.get('words', 0)))       # int
                ws[f'F{row}'] = _avg_len                                     # float ← было строка
                ws[f'F{row}'].number_format = _FMT_FLOAT2
                ws[f'G{row}'] = int(to_decimal(data.get('reactions', 0)))
                ws[f'H{row}'] = int(to_decimal(data.get('reactions_received', 0)))
                ws[f'I{row}'] = int(to_decimal(data.get('replies_received', 0)))
                ws[f'J{row}'] = int(to_decimal(data.get('replies', 0)))
                ws[f'K{row}'] = int(to_decimal(data.get('mentions_received', 0)))
                ws[f'L{row}'] = int(to_decimal(data.get('media', 0)))
                ws[f'M{row}'] = int(to_decimal(data.get('other_threads_posts', 0)))
                ws[f'N{row}'] = int(to_decimal(data.get('days_in_chat', 0)))
                ws[f'O{row}'] = _n(data.get('earned', 0), 2)                # float ← было строка
                ws[f'O{row}'].number_format = _FMT_FLOAT2

                if data.get('activity_index'):
                    act = _n(data['activity_index'], 3)
                else:
                    act = float(round_decimal((
                        Decimal('0.03') * (Decimal(_chars) / _CHARS_NORM) +
                        Decimal('0.03') * (to_decimal(_avg_len) / _CHARS_NORM) +
                        Decimal('0.04') * Decimal(_msgs) +
                        Decimal('0.07') * to_decimal(data.get('reactions', 0)) +
                        Decimal('0.07') * to_decimal(data.get('reactions_received', 0)) +
                        Decimal('0.15') * to_decimal(data.get('replies_received', 0)) +
                        Decimal('0.16') * to_decimal(data.get('replies', 0)) +
                        Decimal('0.15') * to_decimal(data.get('mentions_received', 0)) +
                        Decimal('0.05') * to_decimal(data.get('media', 0)) +
                        Decimal('0.10') * to_decimal(data.get('other_threads_posts', 0))
                    ) / Decimal(_divisor), 3))

                ws[f'P{row}'] = act                                          # float ← было строка
                ws[f'P{row}'].number_format = _FMT_FLOAT3

                for col in detail_cols:
                    ws[f'{col}{row}'].fill = fill
                    ws[f'{col}{row}'].border = thin_border
                ws[f'A{row}'].alignment = center_align
                ws[f'B{row}'].alignment = left_align
                for col in detail_cols[2:]:
                    ws[f'{col}{row}'].alignment = center_align
                row += 1

        # ── ШИРИНА КОЛОНОК ────────────────────────────────────────────────────
        col_widths = {
            'A': 40, 'B': 20, 'C': 14, 'D': 19, 'E': 16, 'F': 25,
            'G': 22, 'H': 20, 'I': 20, 'J': 23, 'K': 23, 'L': 22,
            'M': 16, 'N': 16, 'O': 26, 'P': 20,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = 'A3'

        # ── ДЕТАЛИЗАЦИЯ ПО ДНЯМ/МЕСЯЦАМ (новый лист) ─────────────────────────
        if 'daily_stats' in stats_data and stats_data['daily_stats'] and 'period_type' in stats_data:
            period_type = stats_data['period_type']
            daily_data  = stats_data['daily_stats']

            ws_d = wb.create_sheet("Детализация")

            row = 1
            last_col = get_column_letter(len(daily_data) + 2)
            ws_d.merge_cells(f'A{row}:{last_col}{row}')
            titles_map = {
                'week':  "📊 ДЕТАЛИЗАЦИЯ ПО ДНЯМ - НЕДЕЛЯ",
                'month': "📊 ДЕТАЛИЗАЦИЯ ПО ДНЯМ - МЕСЯЦ",
                'year':  "📊 ДЕТАЛИЗАЦИЯ ПО МЕСЯЦАМ - ГОД",
            }
            ws_d[f'A{row}'] = titles_map.get(period_type, "📊 ДЕТАЛЬНАЯ СТАТИСТИКА")
            ws_d[f'A{row}'].font = title_font
            ws_d[f'A{row}'].fill = title_fill
            ws_d[f'A{row}'].alignment = title_align
            ws_d[f'A{row}'].border = thick_border
            ws_d.row_dimensions[row].height = 30

            row += 1
            ws_d.merge_cells(f'A{row}:{last_col}{row}')
            ws_d[f'A{row}'] = f"Период: {stats_data['start_date']} - {stats_data['end_date']}"
            ws_d[f'A{row}'].font = Font(size=11, italic=True)
            ws_d[f'A{row}'].alignment = center_align
            ws_d.row_dimensions[row].height = 20

            row += 2

            ws_d[f'A{row}'] = "Параметр"
            ws_d[f'B{row}'] = "ИТОГО"
            for idx2, day_stat in enumerate(daily_data):
                col = get_column_letter(idx2 + 3)
                date_str = day_stat['date']
                if period_type == 'year' and 'month' in day_stat and 'year' in day_stat:
                    date_str = format_month_column(day_stat['month'], day_stat['year'], datetime.now().year)
                ws_d[f'{col}{row}'] = date_str
                ws_d[f'{col}{row}'].font = header_font
                ws_d[f'{col}{row}'].fill = header_fill
                ws_d[f'{col}{row}'].alignment = center_align
                ws_d[f'{col}{row}'].border = thin_border

            for hc in ['A', 'B']:
                ws_d[f'{hc}{row}'].font = header_font
                ws_d[f'{hc}{row}'].fill = header_fill
                ws_d[f'{hc}{row}'].alignment = center_align
                ws_d[f'{hc}{row}'].border = thin_border
            ws_d.row_dimensions[row].height = 20
            row += 1

            # (название, daily_key, excel_format, decimal_places)
            params = [
                ('ОКС(Ч) - Общее количество символов',        'chars',               _FMT_INT,    0),
                ('СДС(Ч) - Средняя длина сообщения',          'avg_msg_length',      _FMT_FLOAT2, 2),
                ('ЧО - Частота общения (Всего сообщ.)',     'messages',            _FMT_INT,    0),
                ('КСП - Количество слов',                   'words',               _FMT_INT,    0),
                ('Медиа(Ч) - Медиа контент',                  'media',               _FMT_INT,    0),
                ('КОР(Ч) - Реакции оставленные',              'reactions_given',     _FMT_INT,    0),
                ('КПР(Ч) - Реакции полученные',               'reactions_received',  _FMT_INT,    0),
                ('КОтв(Ч) - Ответы пользователю',             'replies_received',    _FMT_INT,    0),
                ('КОтп(Ч) - Ответы пользователя',             'replies_sent',        _FMT_INT,    0),
                ('КУП(Ч) - Упоминания @',                     'mentions',            _FMT_INT,    0),
                ('ПДВ(Ч) - Публ. в других ветках',          'other_threads',       _FMT_INT,    0),
                ('💬 Сообщений с администраторами',         'msgs_with_admins',    _FMT_INT,    0),
                ('💬 Сообщений без администраторов',        'msgs_without_admins', _FMT_INT,    0),
                ('👥 Активных пользователей',               'active_users',        _FMT_INT,    0),
                ('💎 Добыто Пульсов',                       'pulses',              _FMT_FLOAT2, 2),
                ('👥 Пользователей в чате',                 'total_users',         _FMT_INT,    0),
                ('🆕 Вступило за период',                   'joined',              _FMT_INT,    0),
                ('👋 Вышло за период',                      'left_users',          _FMT_INT,    0),
                ('📊 Коэффициент вовлеченности (ER) %',     'engagement',          _FMT_FLOAT2, 2),
            ]

            for p_idx, (p_name, d_key, fmt, places) in enumerate(params):
                fill = even_fill if p_idx % 2 == 0 else odd_fill

                ws_d[f'A{row}'] = p_name
                ws_d[f'A{row}'].fill = fill
                ws_d[f'A{row}'].alignment = left_align
                ws_d[f'A{row}'].border = thin_border

                # ИТОГО — float, не строка
                if d_key in ('engagement', 'avg_msg_length'):
                    total_v = _n(
                        sum(to_decimal(d.get(d_key, 0)) for d in daily_data) / len(daily_data)
                        if daily_data else 0, places
                    )
                elif d_key == 'total_users':
                    total_v = int(max((to_decimal(d.get(d_key, 0)) for d in daily_data), default=Decimal(0)))
                else:
                    total_v = _n(sum(to_decimal(d.get(d_key, 0)) for d in daily_data), places)

                ws_d[f'B{row}'] = total_v                   # ← float/int, не строка
                ws_d[f'B{row}'].number_format = fmt
                ws_d[f'B{row}'].fill = fill
                ws_d[f'B{row}'].alignment = center_align
                ws_d[f'B{row}'].border = thin_border

                for idx2, day_stat in enumerate(daily_data):
                    col = get_column_letter(idx2 + 3)
                    val = _n(day_stat.get(d_key, 0), places)  # ← float, не строка
                    ws_d[f'{col}{row}'] = val
                    ws_d[f'{col}{row}'].number_format = fmt
                    ws_d[f'{col}{row}'].fill = fill
                    ws_d[f'{col}{row}'].alignment = center_align
                    ws_d[f'{col}{row}'].border = thin_border

                row += 1

            ws_d.column_dimensions['A'].width = 30
            ws_d.column_dimensions['B'].width = 15
            for i in range(len(daily_data)):
                ws_d.column_dimensions[get_column_letter(i + 3)].width = 13
            ws_d.freeze_panes = 'C4'

        # ── ЛИСТ "ВСТУПЛЕНИЯ В ЧАТ" ──────────────────────────────────────────
        if 'user_joins' in stats_data and stats_data['user_joins']:
            ws_j = wb.create_sheet("Вступления в чат")

            row = 1
            ws_j.merge_cells(f'A{row}:G{row}')
            ws_j[f'A{row}'] = "👥 ВСТУПЛЕНИЯ В ЧАТ"
            ws_j[f'A{row}'].font = title_font
            ws_j[f'A{row}'].fill = title_fill
            ws_j[f'A{row}'].alignment = title_align
            ws_j[f'A{row}'].border = thick_border
            ws_j.row_dimensions[row].height = 30

            row += 1
            ws_j.merge_cells(f'A{row}:G{row}')
            ws_j[f'A{row}'] = f"Период: {stats_data['start_date']} - {stats_data['end_date']}"
            ws_j[f'A{row}'].font = Font(size=11, italic=True)
            ws_j[f'A{row}'].alignment = center_align
            ws_j.row_dimensions[row].height = 20

            row += 2

            method_labels = {
                'referral_link':                '🔗 По реферальной ссылке',
                'referral_link_legacy':         '🔗 По реферальной ссылке (старая)',
                'direct':                       '➡️ Прямой /start',
                'direct_invite':                '👤 Прямое приглашение от админа',
                'chat_invite_link':             '🔗 По инвайт-ссылке чата',
                'chat_member_joined':           '📥 Вступил в чат сам',
                'expired_referral_link':        '⛔ Истёкшая ссылка',
                'referral_link_existing_user':  '🔗 Ссылка (уже был в боте)',
                'unknown':                      '❓ Неизвестно',
            }

            join_headers = ['№', 'Имя', '@username', 'ID', 'Метод вступления', 'Пригласил', 'Дата вступления']
            join_cols    = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
            for col, hdr in zip(join_cols, join_headers):
                ws_j[f'{col}{row}'] = hdr
                ws_j[f'{col}{row}'].font = header_font
                ws_j[f'{col}{row}'].fill = header_fill
                ws_j[f'{col}{row}'].alignment = center_align
                ws_j[f'{col}{row}'].border = thin_border
            ws_j.row_dimensions[row].height = 20
            row += 1

            for idx, join in enumerate(stats_data['user_joins'], 1):
                fill = even_fill if idx % 2 == 0 else odd_fill

                joined_at = join.get('joined_at', '')
                if joined_at:
                    try:
                        dt = datetime.strptime(str(joined_at)[:19], '%Y-%m-%d %H:%M:%S')
                        joined_at = dt.strftime('%d.%m.%Y %H:%M')
                    except Exception:
                        joined_at = str(joined_at)[:16]

                referrer = ''
                if join.get('referrer_username'):    referrer = f"@{join['referrer_username']}"
                elif join.get('referrer_first_name'): referrer = join['referrer_first_name']
                elif join.get('referrer_id'):         referrer = f"ID: {join['referrer_id']}"

                values = [
                    idx,
                    join.get('first_name', 'Unknown'),
                    f"@{join['username']}" if join.get('username') else '-',
                    int(to_decimal(join.get('user_id', 0))),    # int ← целостность
                    method_labels.get(join.get('join_method', 'unknown'), '❓'),
                    referrer or '-',
                    joined_at,
                ]
                for c_idx, (col, val) in enumerate(zip(join_cols, values)):
                    cell = ws_j[f'{col}{row}']
                    cell.value = val
                    cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = left_align if c_idx in (1, 4, 5) else center_align
                row += 1

            row += 1
            ws_j.merge_cells(f'A{row}:G{row}')
            method_counts = {}
            for j in stats_data['user_joins']:
                m = j.get('join_method', 'unknown')
                method_counts[m] = method_counts.get(m, 0) + 1
            parts = [f"{method_labels.get(m, m)}: {c}" for m, c in method_counts.items()]
            ws_j[f'A{row}'] = f"Всего вступлений: {len(stats_data['user_joins'])}  |  " + "  |  ".join(parts)
            ws_j[f'A{row}'].font = Font(bold=True, size=10, color="2C3E50")
            ws_j[f'A{row}'].alignment = left_align

            for col, w in zip(['A','B','C','D','E','F','G'], [5,20,18,14,32,22,18]):
                ws_j.column_dimensions[col].width = w
            ws_j.freeze_panes = 'A5'

        # ── ЛИСТ "СТАТ МАЙНИНГ" (детализация по каждой активности) ────────
        if 'mining_detailed' in stats_data and stats_data['mining_detailed']:
            md = stats_data['mining_detailed']
            ws_m = wb.create_sheet("СтатМайнинг")

            # ── Стили ─────────────────────────────────────────────────────
            green_font   = Font(bold=False, size=11, color="27AE60")
            green_bold   = Font(bold=True,  size=11, color="27AE60")
            red_font     = Font(bold=False, size=11, color="E74C3C")
            red_bold     = Font(bold=True,  size=11, color="E74C3C")
            bold_font    = Font(bold=True,  size=12)
            cat_font     = Font(bold=True,  size=12, color="FFFFFF")
            cat_fill_grn = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            cat_fill_org = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            cat_fill_blu = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cat_fill_red = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            total_fill   = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            m_cols = ['A', 'B', 'C', 'D']

            row = 1

            # ── Заголовок ─────────────────────────────────────────────────
            ws_m.merge_cells(f'A{row}:D{row}')
            ws_m[f'A{row}'] = "📊 ДЕТАЛЬНАЯ СТАТИСТИКА МАЙНИНГА"
            ws_m[f'A{row}'].font = title_font
            ws_m[f'A{row}'].fill = title_fill
            ws_m[f'A{row}'].alignment = title_align
            ws_m[f'A{row}'].border = thick_border
            ws_m.row_dimensions[row].height = 30
            row += 1

            ws_m.merge_cells(f'A{row}:D{row}')
            ws_m[f'A{row}'] = f"Период: {stats_data.get('start_date', '?')} — {stats_data.get('end_date', '?')}"
            ws_m[f'A{row}'].font = Font(size=11, italic=True)
            ws_m[f'A{row}'].alignment = center_align
            ws_m.row_dimensions[row].height = 20
            row += 2

            # ── Заголовки таблицы ─────────────────────────────────────────
            tbl_headers = ['Категория', 'Название активности', 'Описание условия', 'Сумма Пульсов (💎)']
            for col, hdr in zip(m_cols, tbl_headers):
                ws_m[f'{col}{row}'] = hdr
                ws_m[f'{col}{row}'].font = header_font
                ws_m[f'{col}{row}'].fill = header_fill
                ws_m[f'{col}{row}'].alignment = center_align
                ws_m[f'{col}{row}'].border = thin_border
            ws_m.row_dimensions[row].height = 22
            row += 1

            idx_row = 0  # для чередования
            chart_data_rows = []  # (label, sum) для PieChart

            # ── Хелпер: строка-категория (цветная полоса) ─────────────────
            def _cat_row(label, cfill):
                nonlocal row
                ws_m.merge_cells(f'A{row}:D{row}')
                ws_m[f'A{row}'] = label
                ws_m[f'A{row}'].font = cat_font
                ws_m[f'A{row}'].fill = cfill
                ws_m[f'A{row}'].alignment = left_align
                ws_m[f'A{row}'].border = thick_border
                ws_m.row_dimensions[row].height = 24
                row += 1

            # ── Хелпер: строка данных ─────────────────────────────────────
            def _data_row(cat, name, desc, val, font_style=green_font):
                nonlocal row, idx_row
                fill = even_fill if idx_row % 2 == 0 else odd_fill
                ws_m[f'A{row}'] = cat
                ws_m[f'B{row}'] = name
                ws_m[f'C{row}'] = desc
                ws_m[f'D{row}'] = _n(val, 4)
                ws_m[f'D{row}'].font = font_style
                ws_m[f'D{row}'].number_format = _FMT_FLOAT2
                for c in m_cols:
                    ws_m[f'{c}{row}'].fill = fill
                    ws_m[f'{c}{row}'].border = thin_border
                ws_m[f'A{row}'].alignment = left_align
                ws_m[f'B{row}'].alignment = left_align
                ws_m[f'C{row}'].alignment = left_align
                ws_m[f'D{row}'].alignment = right_align
                row += 1
                idx_row += 1

            # ══════════════════════════════════════════════════════════════
            #  1. БАЗОВАЯ ДОБЫЧА
            # ══════════════════════════════════════════════════════════════
            base_sum = float(md.get('base_total', 0))
            _cat_row("🔹 БАЗОВАЯ ДОБЫЧА", cat_fill_grn)
            _data_row("", "Базовые начисления",
                      "За сообщения, медиа, реакции и ответы",
                      base_sum, green_font)
            chart_data_rows.append(('Базовая добыча', base_sum))

            # ══════════════════════════════════════════════════════════════
            #  2. КОМБО (каждое поимённо)
            # ══════════════════════════════════════════════════════════════
            combos = md.get('combos', [])
            combo_total = sum(float(c.get('sum', 0)) for c in combos)
            _cat_row("🔥 КОМБО-БОНУСЫ", cat_fill_org)
            if combos:
                for c in combos:
                    _data_row("", c.get('label', c.get('name', '?')),
                              c.get('description', ''),
                              c.get('sum', 0), green_font)
            else:
                _data_row("", "(нет данных)", "Комбо не выполнялись за период", 0, green_font)
            chart_data_rows.append(('Комбо-бонусы', combo_total))

            # ══════════════════════════════════════════════════════════════
            #  3. СПРИНТЫ (каждый поимённо)
            # ══════════════════════════════════════════════════════════════
            sprints = md.get('sprints', [])
            sprint_total = sum(float(s.get('sum', 0)) for s in sprints)
            _cat_row("🚀 СПРИНТЫ", cat_fill_blu)
            if sprints:
                for s in sprints:
                    _data_row("", s.get('label', s.get('name', '?')),
                              s.get('description', ''),
                              s.get('sum', 0), green_font)
            else:
                _data_row("", "(нет данных)", "Спринты не завершались за период", 0, green_font)
            chart_data_rows.append(('Спринты', sprint_total))

            # ══════════════════════════════════════════════════════════════
            #  4. ШТРАФЫ (каждый поимённо)
            # ══════════════════════════════════════════════════════════════
            penalties = md.get('penalties', [])
            penalty_total = sum(float(p.get('sum', 0)) for p in penalties)
            _cat_row("⛔ ШТРАФЫ", cat_fill_red)
            if penalties:
                for p in penalties:
                    _data_row("", p.get('label', p.get('name', '?')),
                              p.get('description', ''),
                              -float(p.get('sum', 0)), red_font)
            else:
                _data_row("", "(нет штрафов)", "Штрафы не применялись за период", 0, red_font)
            chart_data_rows.append(('Штрафы', penalty_total))

            # ══════════════════════════════════════════════════════════════
            #  ИТОГО ЭМИССИЯ
            # ══════════════════════════════════════════════════════════════
            net = _n(Decimal(str(base_sum)) + Decimal(str(combo_total))
                     + Decimal(str(sprint_total)) - Decimal(str(penalty_total)), 4)

            row += 1
            for c in m_cols:
                ws_m[f'{c}{row}'].fill = total_fill
                ws_m[f'{c}{row}'].border = thick_border
            ws_m[f'A{row}'] = "ИТОГО ЭМИССИЯ"
            ws_m[f'A{row}'].font = bold_font
            ws_m[f'A{row}'].alignment = left_align
            ws_m.merge_cells(f'B{row}:C{row}')
            ws_m[f'B{row}'] = f"База + Комбо + Спринты − Штрафы"
            ws_m[f'B{row}'].font = Font(size=10, italic=True)
            ws_m[f'B{row}'].alignment = center_align
            ws_m[f'B{row}'].fill = total_fill
            ws_m[f'D{row}'] = net
            ws_m[f'D{row}'].font = Font(bold=True, size=13,
                                         color="27AE60" if net >= 0 else "E74C3C")
            ws_m[f'D{row}'].number_format = _FMT_FLOAT2
            ws_m[f'D{row}'].alignment = right_align
            ws_m[f'D{row}'].fill = total_fill
            ws_m[f'D{row}'].border = thick_border
            row += 2

            # ── Данные для PieChart (вспомогательные ячейки F2:G5) ─────────
            try:
                from openpyxl.chart import PieChart, Reference
                from openpyxl.chart.series import DataPoint
                from openpyxl.chart.label import DataLabelList

                # Фильтруем нулевые доли — не показываем на пироге
                active_rows = [(lbl, abs(val)) for lbl, val in chart_data_rows if abs(val) > 0]
                if not active_rows:
                    active_rows = [('Нет данных', 1)]

                pie_start = 2
                for pi, (lbl, val) in enumerate(active_rows):
                    ws_m[f'F{pie_start + pi}'] = lbl
                    ws_m[f'G{pie_start + pi}'] = _n(val, 4)
                    ws_m[f'G{pie_start + pi}'].number_format = _FMT_FLOAT2
                    # Скрываем вспомогательные ячейки
                    ws_m[f'F{pie_start + pi}'].font = Font(size=1, color="FFFFFF")
                    ws_m[f'G{pie_start + pi}'].font = Font(size=1, color="FFFFFF")

                chart = PieChart()
                chart.title = "Источники эмиссии Пульсов"
                chart.style = 26
                chart.width = 20
                chart.height = 14

                labels = Reference(ws_m, min_col=6, min_row=pie_start,
                                   max_row=pie_start + len(active_rows) - 1)
                values = Reference(ws_m, min_col=7, min_row=pie_start,
                                   max_row=pie_start + len(active_rows) - 1)
                chart.add_data(values, titles_from_data=False)
                chart.set_categories(labels)

                pie_colors = ['27AE60', 'E67E22', '3498DB', 'E74C3C']
                for i in range(len(active_rows)):
                    pt = DataPoint(idx=i)
                    pt.graphicalProperties.solidFill = pie_colors[i % len(pie_colors)]
                    chart.series[0].data_points.append(pt)

                chart.dataLabels = DataLabelList()
                chart.dataLabels.showPercent = True
                chart.dataLabels.showVal = False
                chart.dataLabels.showCatName = True
                chart.dataLabels.showSerName = False
                chart.dataLabels.showLeaderLines = True
                chart.dataLabels.numFmt = '0.0%'

                # Легенда справа
                from openpyxl.chart.layout import Layout, ManualLayout
                chart.legend.position = 'r'

                ws_m.add_chart(chart, "F8")

            except Exception as chart_err:
                logging.warning(f"Mining PieChart skipped: {chart_err}")

            # ── Ширина колонок ────────────────────────────────────────────
            ws_m.column_dimensions['A'].width = 22
            ws_m.column_dimensions['B'].width = 28
            ws_m.column_dimensions['C'].width = 48
            ws_m.column_dimensions['D'].width = 22
            ws_m.freeze_panes = 'A6'

        # ── СОХРАНЕНИЕ ────────────────────────────────────────────────────────
        logging.info(f"Saving Excel file to: {filename}")
        wb.save(filename)
        logging.info(f"✅ Excel file created successfully: {filename}")
        return filename

    except Exception as e:
        logging.error(f"❌ Error creating Excel: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return None


def export_users_stats_to_excel(users_data, filename, period_name, member_count=10, bot_count=1):
    """
    Export ALL users statistics to Excel.

    Data Integrity: все числовые поля — Decimal/float, не строки.
    """
    CHARS_NORM = Decimal('100')

    def _activity(user) -> float:
        """АктП по формуле — Decimal арифметика, возвращает float для ячейки."""
        total_msgs  = int(to_decimal(user.get('total_messages', 0)))
        total_chars = int(to_decimal(user.get('total_chars', 0)))
        avg_len = to_decimal(total_chars) / to_decimal(total_msgs) if total_msgs > 0 else Decimal('0')
        divisor = max(member_count - bot_count - 1, 1)

        score = (
            Decimal('0.03') * (to_decimal(total_chars) / CHARS_NORM) +
            Decimal('0.03') * (avg_len / CHARS_NORM) +
            Decimal('0.04') * to_decimal(total_msgs) +
            Decimal('0.07') * to_decimal(user.get('reactions_given', 0)) +
            Decimal('0.07') * to_decimal(user.get('reactions_received', 0)) +
            Decimal('0.15') * to_decimal(user.get('replies_received', 0)) +
            Decimal('0.16') * to_decimal(user.get('replies_sent', 0)) +
            Decimal('0.15') * to_decimal(user.get('mentions_received', 0)) +
            Decimal('0.05') * to_decimal(user.get('media_sent', 0)) +
            Decimal('0.10') * to_decimal(user.get('other_threads_posts', 0))
        ) / Decimal(divisor)
        return float(round_decimal(score, 2))

    def _online_time(last_active_str, total_messages: int) -> str:
        from datetime import datetime as dt
        if not last_active_str:
            return 'Не известно'
        try:
            last_active = dt.fromisoformat(str(last_active_str).replace('Z', '+00:00'))
            now = dt.now(last_active.tzinfo) if last_active.tzinfo else dt.now()
            total_sec = int((now - last_active).total_seconds())
            if total_sec < 60:    return "Только что"
            if total_sec < 3600:  return f"{total_sec // 60} мин назад"
            if total_sec < 86400:
                h, m = total_sec // 3600, (total_sec % 3600) // 60
                return f"{h} ч {m} мин назад" if m else f"{h} ч назад"
            days = total_sec // 86400
            if days == 1:         return "Вчера"
            if days < 30:         return f"{days} дн назад"
            est = total_messages * 2
            return f"~{est} мин" if est < 60 else f"~{est // 60} ч"
        except Exception:
            if total_messages > 0:
                est = total_messages * 2
                return f"~{est} мин" if est < 60 else f"~{est // 60} ч"
            return 'Не известно'

    try:
        logging.info(f"📊 Starting users Excel export to: {filename}")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Все пользователи"

        title_font  = Font(bold=True, size=16, color="FFFFFF")
        title_fill  = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align   = Alignment(horizontal='left',   vertical='center')
        even_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color="BDC3C7"), right=Side(style='thin', color="BDC3C7"),
            top=Side(style='thin', color="BDC3C7"),  bottom=Side(style='thin', color="BDC3C7"),
        )

        ws.merge_cells('A1:V1')
        ws['A1'] = f'📊 СТАТИСТИКА ВСЕХ ПОЛЬЗОВАТЕЛЕЙ ЧАТА - {period_name.upper()}'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        headers = [
            '№', 'Имя', '@username', 'ID', 'User ID', 'Сообщений',
            'Общее кол-во\nсимволов', 'Количество\nслов', 'Средняя\nдлина',
            'Реакции\nоставленные', 'Реакции\nполученные',
            'Ответы\nпользователю', 'Ответы\nпользователя',
            'Упоминания\n@', 'Медиа', 'Др. ветки',
            'Дней\n в чате', 'Когда\nпришёл', 'Последняя\nактивность',
            'Онлайн', 'Заработано\n💎', 'Активность\n🏅',
        ]
        for col_num, hdr in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = hdr
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 35

        # Соответствие колонка → (number_format, decimal_places)
        # Колонка 5 = User ID (строка), не форматируем
        num_fmt_map = {
            6:  (_FMT_INT,    0),   # Сообщений
            7:  (_FMT_INT,    0),   # ОКСП
            8:  (_FMT_INT,    0),   # КСП
            9:  (_FMT_FLOAT2, 2),   # СДСП
            10: (_FMT_INT,    0),   # КОРП
            11: (_FMT_INT,    0),   # КПРП
            12: (_FMT_INT,    0),   # КОПЮП
            13: (_FMT_INT,    0),   # КОПЯП
            14: (_FMT_INT,    0),   # КУПП
            15: (_FMT_INT,    0),   # МедиаП
            16: (_FMT_INT,    0),   # ПИВДВП
            17: (_FMT_INT,    0),   # Дней в чате
            21: (_FMT_FLOAT2, 2),   # Заработано
            22: (_FMT_FLOAT3, 3),   # Активность
        }

        for idx, user in enumerate(users_data, 1):
            row = idx + 2
            fill = even_fill if idx % 2 == 0 else odd_fill

            total_msgs  = int(to_decimal(user.get('total_messages', 0)))
            total_chars = int(to_decimal(user.get('total_chars', 0)))
            avg_chars   = _n(Decimal(total_chars) / Decimal(total_msgs), 2) if total_msgs > 0 else 0.0

            joined_at        = user.get('joined_at', '')
            joined_date      = joined_at[:10] if joined_at else '-'
            last_active      = user.get('last_active', '')
            last_active_date = last_active[:10] if last_active else '-'

            values = [
                idx,                                                            # 1  №
                user.get('first_name', 'Unknown'),                              # 2  Имя
                f"@{user['username']}" if user.get('username') else '-',       # 3  @username
                int(to_decimal(user.get('user_id', 0))),                       # 4  ID
                f"#user{int(to_decimal(user.get('user_id', 0)))}",            # 5  User ID
                total_msgs,                                                     # 6  Сообщений   int
                total_chars,                                                    # 7  ОКСП        int
                int(to_decimal(user.get('total_words', 0))),                   # 8  КСП         int
                avg_chars,                                                      # 9  СДСП        float
                int(to_decimal(user.get('reactions_given', 0))),               # 10 КОРП        int
                int(to_decimal(user.get('reactions_received', 0))),            # 11 КПРП        int
                int(to_decimal(user.get('replies_received', 0))),              # 12 КОПЮП       int
                int(to_decimal(user.get('replies_sent', 0))),                  # 13 КОПЯП       int
                int(to_decimal(user.get('mentions_received', 0))),             # 14 КУПП        int
                int(to_decimal(user.get('media_sent', 0))),                    # 15 МедиаП      int
                int(to_decimal(user.get('other_threads_posts', 0))),           # 16 ПИВДВП      int
                int(to_decimal(user.get('days_in_chat', 0))),                  # 17 Дней        int
                joined_date,                                                    # 18 Когда пришёл  str
                last_active_date,                                               # 19 Последняя акт str
                _online_time(last_active, total_msgs),                         # 20 Онлайн        str
                _n(user.get('pulses_mined', 0), 2),                            # 21 Заработано  float
                _activity(user),                                                # 22 Активность  float
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = value
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = left_align if col_num == 2 else center_align
                if col_num in num_fmt_map:
                    cell.number_format = num_fmt_map[col_num][0]

        widths = {
            'A': 30, 'B': 20, 'C': 16, 'D': 16, 'E': 22, 'F': 16,
            'G': 18, 'H': 16, 'I': 16, 'J': 16, 'K': 16,
            'L': 18, 'M': 12, 'N': 20, 'O': 16, 'P': 16,
            'Q': 16, 'R': 18, 'S': 18, 'T': 18, 'U': 16, 'V': 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = 'A3'

        logging.info(f"💾 Saving users Excel file to: {filename}")
        wb.save(filename)
        logging.info(f"✅ Users Excel created: {len(users_data)} users, 22 columns")
        return filename

    except Exception as e:
        logging.error(f"❌ Error creating users Excel: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return None
