import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

def create_economy_calculator():
    wb = openpyxl.Workbook()
    
    # ==========================================
    # СТИЛИ И ВАЛИДАЦИЯ
    # ==========================================
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="2C3E50")
    subheader_font = Font(bold=True, color="2C3E50", italic=True)
    
    fill_text = PatternFill(start_color="34495E", fill_type="solid")
    fill_media = PatternFill(start_color="8E44AD", fill_type="solid")
    fill_social = PatternFill(start_color="D35400", fill_type="solid")
    fill_biomes = PatternFill(start_color="16A085", fill_type="solid")
    fill_quests = PatternFill(start_color="2980B9", fill_type="solid") # Синий для квестов
    fill_subheader = PatternFill(start_color="E5E7E9", fill_type="solid")
    fill_input = PatternFill(start_color="FFF2CC", fill_type="solid")
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True) # Для примечаний

    dv = DataValidation(type="list", formula1='"1 час,12 часов,24 часа,Неделя,Рандом"', allow_blank=True)
    
    # ==========================================
    # ДАННЫЕ ЛИСТА 1: НАСТРОЙКИ (4 параметра: Имя, Коэфф, Частота, Примечание)
    # ==========================================
    data_structure =[
        {
            "title": "📝 ТЕКСТ И ОБЩЕНИЕ",
            "start_col": 1,
            "fill": fill_text,
            "blocks": {
                "🔹 БАЗА":[
                    ("Текст (сообщение)", 1, "-", "Текст длиннее 2 слов"),
                    ("Эмодзи (без текста)", 1, "-", "Только смайлики"),
                ],
                "🔥 КОМБО":[
                    ("Писатель (> 50 симв)", 10, "24 часа", "Длинный осмысленный текст"),
                ],
                "🚀 СПРИНТЫ":[
                    ("Основа чата (10 сообщ)", 25, "24 часа", "Написать 10 сообщений за сутки"),
                    ("Эмоциональный (10 эмодзи)", 10, "24 часа", "Отправить 10 отдельных эмодзи"),
                ],
                "🎁 СЕКРЕТЫ (Пасхалки)":[
                    ("Магия чисел (11:11 и т.д.)", 5500, "24 часа", "Сообщение в красивое время"), 
                    ("Ранняя пташка (05:00)", 7500, "24 часа", "Первое сообщение чата утром"),      
                    ("Лев Толстой (>1000 симв)", 15000, "24 часа", "Мега-лонгрид"),  
                    ("Юбилейное сообщение", 25000, "24 часа", "100-е, 500-е сообщение за день"),       
                    ("Резонанс (Секретное слово)", 15000, "Рандом", "Употребить тайное слово недели"), 
                    ("Критический удар (Шанс 2%)", 5000, "Рандом", "Случайный джекпот при отправке"),  
                    ("Амбассадор вежливости", 100, "1 час", "Слова спасибо/пожалуйста"),         
                ],
                "⛔ ШТРАФЫ":[
                    ("Копипаст", -25000, "-", "Отправка одинакового текста"),
                ]
            }
        },
        {
            "title": "📷 МЕДИА И ВЛОЖЕНИЯ",
            "start_col": 6,
            "fill": fill_media,
            "blocks": {
                "🔹 БАЗА":[
                    ("Фото", 4, "-", "Одно фото"),
                    ("Фото (2+ в спец. ветке)", 6, "-", "Альбом в Питомцах и т.д."),
                    ("Видео обычное", 5, "-", "Загруженный видеофайл"),
                    ("Видео кружок", 7, "-", "Video Note"),
                    ("Аудио / Ссылка", 2, "-", "Трек или URL"),
                    ("Голосовое (Voice)", 3, "-", "Голосовое сообщение"),
                    ("GIF-анимация", 1, "-", "Гифка"),
                ],
                "🔥 КОМБО":[
                    ("Иллюстратор (Текст+Фото)", 20, "24 часа", "Фотография с описанием"),
                    ("Рецензент (Видео+Текст)", 20, "24 часа", "Видео с длинным текстом"),
                    ("Диджей (Плейлист)", 15, "24 часа", "Ссылка на плейлист"),
                ],
                "🚀 СПРИНТЫ":[
                    ("Фотограф (3+ фото)", 20, "24 часа", "3 фото за сутки"),
                    ("Режиссер (2 видео)", 70, "24 часа", "2 видео в спецветках"),
                    ("Лицом торгуешь (5 кружков)", 30, "12 часов", "5 видеокружков за полдня"),
                    ("Меломан (5 треков)", 60, "24 часа", "5 треков в МузON"),
                    ("Радио (4 войса)", 10, "1 час", "4 голосовых за час"),
                    ("Гифошная (10 гиф)", 15, "1 час", "Квота на гифки"),
                ],
                "🎁 СЕКРЕТЫ (Пасхалки)":[],
                "⛔ ШТРАФЫ":[]
            }
        },
        {
            "title": "🤝 СОЦИУМ И РЕАКЦИИ",
            "start_col": 11,
            "fill": fill_social,
            "blocks": {
                "🔹 БАЗА":[
                    ("Ответ (ты ответил)", 1, "-", "Сделал Reply"),
                    ("Ответ (тебе ответили)", 1, "-", "Твой пост процитировали"),
                    ("Лайк (ты поставил)", 1, "-", "Поставил реакцию"),
                    ("Лайк (тебе поставили)", 2, "-", "Получил реакцию"),
                ],
                "🔥 КОМБО":[
                    ("Острый язычок (>2 репостов)", 3, "24 часа", "На твой пост ответили 3+ раз"),
                    ("Вирусный пост (>2 лайков)", 2, "24 часа", "Пост собрал 3 лайка"),
                    ("Хит (4 лайка)", 5, "24 часа", "Пост собрал 4 лайка"),
                    ("Легенда (6+ лайков)", 10, "24 часа", "Пост собрал 6 лайков"),
                ],
                "🚀 СПРИНТЫ":[
                    ("Болтун (20 твоих ответов)", 20, "1 час", "Много реплаев за час"),
                    ("Центр (20 ответов тебе)", 25, "12 часов", "Активное обсуждение тебя"),
                    ("Щедрая душа (5 лайков)", 10, "1 час", "Поставил 5 лайков"),
                    ("Любимчик (10 лайков тебе)", 20, "1 час", "Собрал 10 лайков за час"),
                ],
                "🎁 СЕКРЕТЫ (Пасхалки)":[
                    ("Хлеб-соль (Ответ новичку)", 12500, "24 часа", "Ответил на первое сообщение новенького"), 
                    ("Пинг-понг (Ответ < 10 сек)", 2500, "1 час", "Молниеносный ответ"),   
                    ("Некромант (Ответ на пост >48ч)", 10000, "12 часов", "Реанимация старой темы"), 
                ],
                "⛔ ШТРАФЫ":[
                    ("Не в ту дверь (Нарушение)", -25000, "-", "Текст в ветке для медиа и т.д."),
                    ("Удаление (Токсик)", -50000, "-", "Пост удален админом"),
                ]
            }
        },
        {
            "title": "🌍 БИОМЫ И ИВЕНТЫ",
            "start_col": 16,
            "fill": fill_biomes,
            "blocks": {
                "🟢 ПРОСТЫЕ":[
                    ("Без спойлеров (КиноON)", 1000, "12 часов", "Использован скрытый текст"),
                    ("Организатор (Встречи)", 1500, "24 часа", "Написано точное время и день"),
                ],
                "🔵 СРЕДНИЕ":[
                    ("Сбор Пати (ЗадротыON)", 4000, "24 часа", "Тегнуто 3+ человек в одном посте"),
                    ("Журналист (НьюзON)", 5000, "12 часов", "Ссылка + текст > 20 слов"),
                ],
                "🟣 СЛОЖНЫЕ":[
                    ("Ночной Дозор (ПлюсON)", 12500, "24 часа", "Развернутый ответ с 02:00 до 06:00"),
                    ("Амбассадор Любви (BBS)", 10000, "24 часа", "Реакция ❤️/🔥 на 3 разные анкеты"),
                ],
                "🟡 ЛЕГЕНДАРНЫЕ":[
                    ("Резонанс Толпы (5 одинак. эмодзи)", 25000, "Рандом", "Флешмоб из 5 одинаковых смайлов подряд"),
                    ("Разрыв Интернета (15 лайков)", 75000, "-", "Мем собрал 15 лайков за 1 час"),
                ],
                "🐸 ИВЕНТЫ":[
                    ("Поимка Золотой Лягушки", 25000, "Рандом", "Первым написать в загаданную ветку"),
                ]
            }
        }
    ]

    # ==========================================
    # Вкладка 1: НАСТРОЙКИ
    # ==========================================
    ws_set = wb.active
    ws_set.title = "Настройки"
    ws_set.add_data_validation(dv)
    
    ws_set.merge_cells('A1:C1')
    ws_set['A1'] = "ГЛОБАЛЬНАЯ БАЗА (1x) ="
    ws_set['A1'].font = title_font
    ws_set['A1'].alignment = Alignment(horizontal="right")
    
    ws_set['D1'] = 0.002
    ws_set['D1'].font = Font(bold=True, size=14, color="E74C3C")
    ws_set['D1'].number_format = '0.000'
    ws_set['E1'] = "💎 Пульсов"
    ws_set['E1'].font = title_font

    for cat in data_structure:
        col = cat["start_col"]
        
        # Заголовки (Добавлено Примечание)
        ws_set.cell(row=3, column=col, value=cat["title"]).font = header_font
        ws_set.cell(row=3, column=col).fill = cat["fill"]
        ws_set.cell(row=3, column=col+1, value="Коэф. (x)").font = header_font
        ws_set.cell(row=3, column=col+1).fill = cat["fill"]
        ws_set.cell(row=3, column=col+2, value="Частота").font = header_font
        ws_set.cell(row=3, column=col+2).fill = cat["fill"]
        ws_set.cell(row=3, column=col+3, value="Примечание").font = header_font
        ws_set.cell(row=3, column=col+3).fill = cat["fill"]
        
        ws_set.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
        ws_set.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = 10
        ws_set.column_dimensions[openpyxl.utils.get_column_letter(col+2)].width = 12
        ws_set.column_dimensions[openpyxl.utils.get_column_letter(col+3)].width = 35 # Широкая колонка для текста
        ws_set.column_dimensions[openpyxl.utils.get_column_letter(col+4)].width = 2  # Разделитель
        
        current_row = 4
        for block_name, items in cat["blocks"].items():
            if not items: continue
            
            ws_set.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+3)
            sub_cell = ws_set.cell(row=current_row, column=col, value=block_name)
            sub_cell.font = subheader_font
            sub_cell.fill = fill_subheader
            sub_cell.alignment = align_center
            for c_offset in range(4): ws_set.cell(row=current_row, column=col+c_offset).border = border_thin
            current_row += 1
            
            for name, coeff, freq, desc in items:
                ws_set.cell(row=current_row, column=col, value=name).border = border_thin
                
                c_cell = ws_set.cell(row=current_row, column=col+1, value=coeff)
                c_cell.alignment = align_center
                c_cell.border = border_thin
                
                f_cell = ws_set.cell(row=current_row, column=col+2, value=freq)
                f_cell.alignment = align_center
                f_cell.border = border_thin
                
                d_cell = ws_set.cell(row=current_row, column=col+3, value=desc)
                d_cell.alignment = align_wrap
                d_cell.border = border_thin
                
                if block_name not in["🔹 БАЗА", "⛔ ШТРАФЫ"]:
                    dv.add(f_cell)
                    
                current_row += 1
            current_row += 1

    # ==========================================
    # Вкладка 2: СИМУЛЯТОР (Оставляем как было, сдвигаем колонки)
    # ==========================================
    ws_sim = wb.create_sheet("Симулятор")
    ws_sim.merge_cells('A1:B2')
    ws_sim['A1'] = "ИТОГО ЗА ДЕНЬ:"
    ws_sim['A1'].font = Font(bold=True, size=16)
    ws_sim['A1'].alignment = Alignment(horizontal="right", vertical="center")
    
    ws_sim.merge_cells('C1:D2')
    ws_sim['C1'] = f"=SUM(V3:V6)" # Сдвинуто из-за новых колонок
    ws_sim['C1'].font = Font(bold=True, size=20, color="27AE60")
    ws_sim['C1'].number_format = '0.000 💎'
    ws_sim['C1'].alignment = align_left

    sum_cells = {"Текст":[], "Медиа":[], "Социум":[], "Биомы": []}
    cat_names =["Текст", "Медиа", "Социум", "Биомы"]

    for idx, cat in enumerate(data_structure):
        col = (idx * 5) + 1 # 1, 6, 11, 16
        cat_key = cat_names[idx]
        
        ws_sim.cell(row=4, column=col, value=cat["title"]).font = header_font
        ws_sim.cell(row=4, column=col).fill = cat["fill"]
        ws_sim.cell(row=4, column=col+1, value="Кол-во").font = header_font
        ws_sim.cell(row=4, column=col+1).fill = cat["fill"]
        ws_sim.cell(row=4, column=col+2, value="💎 Итого").font = header_font
        ws_sim.cell(row=4, column=col+2).fill = cat["fill"]
        
        ws_sim.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
        ws_sim.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = 10
        ws_sim.column_dimensions[openpyxl.utils.get_column_letter(col+2)].width = 12
        
        current_row = 5
        set_row = 4 
        
        for block_name, items in cat["blocks"].items():
            if not items: continue
            
            ws_sim.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+2)
            sub_cell = ws_sim.cell(row=current_row, column=col, value=block_name)
            sub_cell.font = subheader_font
            sub_cell.fill = fill_subheader
            sub_cell.alignment = align_center
            for c_offset in range(3): ws_sim.cell(row=current_row, column=col+c_offset).border = border_thin
            
            current_row += 1
            set_row += 1
            
            for name, coeff, freq, desc in items:
                set_col_letter = openpyxl.utils.get_column_letter(cat["start_col"])
                coeff_col_letter = openpyxl.utils.get_column_letter(cat["start_col"]+1)
                
                ws_sim.cell(row=current_row, column=col, value=f"='Настройки'!{set_col_letter}{set_row}").border = border_thin
                
                in_cell = ws_sim.cell(row=current_row, column=col+1, value=0)
                in_cell.fill = fill_input
                in_cell.alignment = align_center
                in_cell.border = border_thin
                
                in_col_letter = openpyxl.utils.get_column_letter(col+1)
                formula = f"={in_col_letter}{current_row} * 'Настройки'!{coeff_col_letter}{set_row} * 'Настройки'!$D$1"
                tot_cell = ws_sim.cell(row=current_row, column=col+2, value=formula)
                tot_cell.number_format = '0.000'
                tot_cell.font = Font(bold=True, color="2C3E50")
                tot_cell.border = border_thin
                
                sum_cells[cat_key].append(f"{openpyxl.utils.get_column_letter(col+2)}{current_row}")
                
                current_row += 1
                set_row += 1
                
            current_row += 1
            set_row += 1

    sum_col = 21 # U
    ws_sim.cell(row=2, column=sum_col, value="Сектор").font = header_font
    ws_sim.cell(row=2, column=sum_col).fill = fill_text
    ws_sim.cell(row=2, column=sum_col+1, value="Заработано").font = header_font
    ws_sim.cell(row=2, column=sum_col+1).fill = fill_text
    ws_sim.column_dimensions['U'].width = 15
    ws_sim.column_dimensions['V'].width = 15

    for idx, cat_name in enumerate(cat_names, start=3):
        ws_sim.cell(row=idx, column=sum_col, value=cat_name).font = Font(bold=True)
        cells_to_sum = ",".join(sum_cells[cat_name])
        ws_sim.cell(row=idx, column=sum_col+1, value=f"=SUM({cells_to_sum})" if cells_to_sum else 0).number_format = '0.000'

    pie = PieChart()
    labels = Reference(ws_sim, min_col=sum_col, min_row=3, max_row=6)
    data = Reference(ws_sim, min_col=sum_col+1, min_row=2, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Источники дохода 💎"
    ws_sim.add_chart(pie, "U8")

    # ==========================================
    # Вкладка 3: КВЕСТЫ И СЕЗОНЫ (Дейлики/Виклики)
    # ==========================================
    ws_quests = wb.create_sheet("Квесты и Сезоны")
    
    quest_headers =["Категория", "Название Квеста", "Описание (Примечание)", "Награда (💎)", "Частота"]
    for col, h in enumerate(quest_headers, 1):
        cell = ws_quests.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_quests
        cell.alignment = align_center
        cell.border = border_thin
        
    ws_quests.column_dimensions['A'].width = 20
    ws_quests.column_dimensions['B'].width = 30
    ws_quests.column_dimensions['C'].width = 50
    ws_quests.column_dimensions['D'].width = 15
    ws_quests.column_dimensions['E'].width = 15

    # Пример данных для квестов (чтобы сохранить твою структуру)
    quests_data =[
        ("🟢 Простые", "Жаворонок", "Написать первое сообщение в чат до 08:00 МСК", 2, "24 часа"),
        ("🟢 Простые", "Эмоциональный отклик", "Поставить 5 любых реакций на чужие сообщения", 1, "24 часа"),
        ("🟢 Простые", "Котопес", "Поставить 3 реакции ❤️ на чужих животных в ПИТОМЦЫ", 1, "24 часа"),
        
        ("🔵 Средние", "Длинная мысль", "Написать осмысленное сообщение от 20 до 40 слов", 3, "24 часа"),
        ("🔵 Средние", "Голос чата", "Записать 1 голосовое сообщение (длиннее 10 секунд)", 3, "24 часа"),
        ("🔵 Средние", "Кружок по интересам", "Записать 1 Video Note (кружочек лицом)", 5, "24 часа"),
        
        ("🟣 Сложные", "Звезда эфира", "Ваше голосовое сообщение или кружок собрал 5+ реакций", 12, "24 часа"),
        ("🟣 Сложные", "Топовый мем", "Ваш мем в МемасON собрал 7+ реакций от разных людей", 15, "24 часа"),
        ("🟣 Сложные", "Лидер мнений", "Текст в Главном чате вызвал дискуссию (4 ответа)", 15, "24 часа"),
        
        ("🟡 Еженедельные", "Марафонец", "Выполнить 20 ежедневных квестов (дейликов) за неделю", 40, "Неделя"),
        ("🟡 Еженедельные", "Меценат недели", "Пожертвовать в Реактор суммарно 500 💎", 50, "Неделя"),
        
        ("🗺 Сезон 1 (Бронза)", "Ритуал Времени", "Говорят, если заговорить с миром, когда часы показывают идеальное отражение...", 15, "24 часа"),
        ("🗺 Сезон 1 (Бронза)", "Слово Древних", "Варвары кричат, а мудрецы благодарят. Покажи манеры в длинной речи.", 10, "24 часа"),
        ("🗺 Сезон 1 (Серебро)", "Некромантия", "Спустись в архивы и вдохни жизнь в дискуссию, мертвую двое суток.", 40, "12 часов"),
        ("🗺 Сезон 1 (Золото)", "Ночной Дозор", "Когда биом ПлюсON во тьме (02-06), стань тем, кто даст развернутый ответ.", 150, "24 часа"),
    ]

    current_row = 2
    for cat, name, desc, reward, freq in quests_data:
        ws_quests.cell(row=current_row, column=1, value=cat).border = border_thin
        ws_quests.cell(row=current_row, column=2, value=name).border = border_thin
        
        d_cell = ws_quests.cell(row=current_row, column=3, value=desc)
        d_cell.border = border_thin
        d_cell.alignment = align_wrap
        
        r_cell = ws_quests.cell(row=current_row, column=4, value=reward)
        r_cell.border = border_thin
        r_cell.alignment = align_center
        r_cell.font = Font(bold=True, color="27AE60")
        
        f_cell = ws_quests.cell(row=current_row, column=5, value=freq)
        f_cell.border = border_thin
        f_cell.alignment = align_center
        
        current_row += 1

    # Сохраняем под НОВЫМ именем, чтобы не затереть твой старый файл!
    wb.save("Pulse_Economy_V6.xlsx")
    print("✅ Файл 'Pulse_Economy_V6.xlsx' успешно создан! Открой его и перенеси свои данные.")

if __name__ == "__main__":
    create_economy_calculator()
    #Триггер срабатывания
