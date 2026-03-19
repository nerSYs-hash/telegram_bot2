#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Функция для экспорта статистики в Excel с расширенными данными
Добавьте эту функцию в utils/helpers.py или создайте отдельный файл
"""

def export_stats_to_excel(stats_data, filepath):
    """
    Экспорт статистики в Excel файл
    
    Args:
        stats_data: Словарь с данными статистики
        filepath: Путь для сохранения файла
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # === ЛИСТ 1: ОБЩАЯ СТАТИСТИКА ===
        ws_general = wb.active
        ws_general.title = "Общая статистика"
        
        # Заголовок
        ws_general['A1'] = stats_data.get('title', 'Статистика чата')
        ws_general['A1'].font = Font(size=16, bold=True)
        ws_general.merge_cells('A1:B1')
        
        # Период
        ws_general['A2'] = f"Период: {stats_data.get('start_date', '')} - {stats_data.get('end_date', '')}"
        ws_general['A2'].font = Font(size=11)
        ws_general.merge_cells('A2:B2')
        
        # Пустая строка
        row = 4
        
        # Заголовки столбцов
        ws_general[f'A{row}'] = "Метрика"
        ws_general[f'B{row}'] = "Значение"
        
        for col in ['A', 'B']:
            cell = ws_general[f'{col}{row}']
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Данные
        for metric, value in stats_data.get('general', {}).items():
            ws_general[f'A{row}'] = metric
            ws_general[f'B{row}'] = value
            ws_general[f'A{row}'].font = Font(size=11)
            ws_general[f'B{row}'].font = Font(size=11, bold=True)
            ws_general[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Авто-ширина столбцов
        ws_general.column_dimensions['A'].width = 50
        ws_general.column_dimensions['B'].width = 25
        
        # === ЛИСТ 2: ТОП ПО СООБЩЕНИЯМ ===
        ws_top_msg = wb.create_sheet("Топ по сообщениям")
        
        # Заголовок
        ws_top_msg['A1'] = "ТОП-10 ПОЛЬЗОВАТЕЛЕЙ ПО СООБЩЕНИЯМ"
        ws_top_msg['A1'].font = Font(size=14, bold=True)
        ws_top_msg.merge_cells('A1:N1')
        
        row = 3
        
        # Заголовки таблицы
        if stats_data.get('top_messages'):
            headers = list(stats_data['top_messages'][0].keys())
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_top_msg.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            
            # Данные
            for user_data in stats_data['top_messages']:
                for col_idx, (key, value) in enumerate(user_data.items(), 1):
                    cell = ws_top_msg.cell(row=row, column=col_idx, value=value)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='center' if col_idx == 1 else 'left')
                row += 1
        
        # Авто-ширина столбцов
        for col_idx in range(1, 15):
            col_letter = get_column_letter(col_idx)
            if col_idx == 2:  # Пользователь
                ws_top_msg.column_dimensions[col_letter].width = 20
            else:
                ws_top_msg.column_dimensions[col_letter].width = 12
        
        # === ЛИСТ 3: ТОП ПО ЗАРАБОТКУ ===
        ws_top_earn = wb.create_sheet("Топ по заработку")
        
        # Заголовок
        ws_top_earn['A1'] = "ТОП-10 ПОЛЬЗОВАТЕЛЕЙ ПО ДОБЫТЫМ ПУЛЬСАМ"
        ws_top_earn['A1'].font = Font(size=14, bold=True)
        ws_top_earn.merge_cells('A1:D1')
        
        row = 3
        
        # Заголовки таблицы
        if stats_data.get('top_earners'):
            headers = list(stats_data['top_earners'][0].keys())
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_top_earn.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            
            # Данные
            for user_data in stats_data['top_earners']:
                for col_idx, (key, value) in enumerate(user_data.items(), 1):
                    cell = ws_top_earn.cell(row=row, column=col_idx, value=value)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='center' if col_idx == 1 else 'left')
                row += 1
        
        # Авто-ширина столбцов
        ws_top_earn.column_dimensions['A'].width = 10
        ws_top_earn.column_dimensions['B'].width = 20
        ws_top_earn.column_dimensions['C'].width = 20
        ws_top_earn.column_dimensions['D'].width = 15
        
        # Сохранение файла
        wb.save(filepath)
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False
